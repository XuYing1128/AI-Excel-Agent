"""Deterministic builders for user-defined tables and uploaded datasets."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .content_plan import infer_column_kind
from .data_cleaner import clean_dataframe
from .io_utils import ensure_output_path, read_table
from .style_library import (
    apply_calculation_style,
    apply_header_style,
    apply_input_style,
    apply_print_settings,
    apply_summary_style,
    apply_title_style,
    freeze_and_filter,
    set_reasonable_column_widths,
    style_instruction_sheet,
)


def build_custom_workbook(
    plan: dict[str, Any],
    output: str | Path,
    *,
    include_charts: bool = False,
) -> Path:
    """Create a workbook that follows explicit columns, records and semantic rules."""

    output_path = ensure_output_path(output, "自定义表格.xlsx")
    columns = [dict(item) for item in plan.get("columns", []) if item.get("name")]
    if not columns:
        raise ValueError("自定义生成方案没有可用列。")
    records = [dict(item) for item in plan.get("records", []) if isinstance(item, dict)]
    title = str(plan.get("title") or "自定义表格").strip()
    year_match = re.search(r"(20\d{2})年", title)
    default_year = int(year_match.group(1)) if year_match else None
    sheet_name = str(plan.get("sheet_name") or "数据").strip()[:31] or "数据"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    max_col = len(columns)
    header_row = 3
    start_row = 4
    minimum_rows = max(1, len(records))
    end_row = start_row + minimum_rows - 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = title
    apply_title_style(ws, 1, 1, max_col)
    ws["A2"] = "浅黄色为录入内容，浅绿色为自动计算。"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    column_index = {item["name"]: index for index, item in enumerate(columns, start=1)}
    for index, item in enumerate(columns, start=1):
        ws.cell(header_row, index, item["name"])
    apply_header_style(ws, header_row)

    formula_rules = {
        str(item.get("target")): item for item in plan.get("formula_rules", [])
    }
    formula_cols: list[int] = []
    input_cols: list[int] = []
    for col_idx, column in enumerate(columns, start=1):
        if column["name"] in formula_rules or column.get("role") == "formula":
            formula_cols.append(col_idx)
        else:
            input_cols.append(col_idx)

    for row_offset in range(minimum_rows):
        excel_row = start_row + row_offset
        record = records[row_offset] if row_offset < len(records) else {}
        for col_idx, column in enumerate(columns, start=1):
            name = column["name"]
            if name in formula_rules:
                value = _formula_for_rule(
                    formula_rules[name],
                    excel_row,
                    column_index,
                )
            else:
                value = _coerce_value(
                    record.get(name, ""),
                    column.get("kind", "text"),
                    default_year=default_year,
                )
            ws.cell(excel_row, col_idx, value)

    apply_input_style(ws, start_row, end_row, input_cols)
    apply_calculation_style(ws, start_row, end_row, formula_cols)
    _apply_custom_number_formats(ws, columns, start_row, end_row)
    freeze_and_filter(
        ws,
        f"A{start_row}",
        f"A{header_row}:{get_column_letter(max_col)}{end_row}",
    )
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)
    _write_inline_summary(ws, plan, column_index, start_row, end_row)
    if include_charts:
        _add_custom_chart(ws, columns, start_row, end_row)
    wb.save(output_path)
    return output_path


def build_inline_tables_workbook(
    plan: dict[str, Any],
    output: str | Path,
    *,
    include_charts: bool = False,
) -> Path:
    """Preserve every table embedded in the request as a usable workbook.

    This is the deterministic fallback for complex multi-table prompts. It is
    deliberately preferable to selecting an unrelated business template when
    a model response is unavailable.
    """

    tables = [
        dict(item)
        for item in plan.get("inline_tables", [])
        if isinstance(item, dict) and item.get("columns")
    ]
    if not tables:
        raise ValueError("需求中没有可生成的内嵌数据表。")
    output_path = ensure_output_path(output, "需求内数据表.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = wb.create_sheet("说明")
    instructions["A1"] = str(plan.get("title") or "表格生成说明")
    instructions["A3"] = "数据来源"
    instructions["B3"] = "用户需求文字中识别出的结构化数据"
    instructions["A4"] = "识别结果"
    instructions["B4"] = f"共识别 {len(tables)} 个数据表，已分别保存为工作表。"
    instructions["A5"] = "计算说明"
    instructions["B5"] = (
        "原始输入数据已完整保留；模型方案不可用时，不会擅自套用无关模板。"
    )
    instructions["A6"] = "复核提醒"
    instructions["B6"] = "复杂计算、薪酬、绩效、财务结果需要人工复核。"
    style_instruction_sheet(instructions)

    used_names = {"说明"}
    primary_name = str(plan.get("primary_table_name") or "")
    for index, table in enumerate(tables, start=1):
        requested_name = str(table.get("name") or f"数据表{index}")
        sheet_name = _unique_sheet_name(requested_name, used_names)
        used_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        title = (
            str(plan.get("title") or requested_name)
            if requested_name == primary_name
            else requested_name
        )
        _write_inline_table_sheet(ws, table, title)

    if include_charts:
        primary_sheet = next(
            (
                wb[name]
                for name in wb.sheetnames
                if name != "说明" and primary_name and primary_name[:31] in name
            ),
            wb[wb.sheetnames[1]],
        )
        headers = [
            {
                "name": str(primary_sheet.cell(3, col).value or ""),
                "kind": infer_column_kind(primary_sheet.cell(3, col).value or ""),
                "role": "input",
            }
            for col in range(1, primary_sheet.max_column + 1)
        ]
        _add_custom_chart(primary_sheet, headers, 4, primary_sheet.max_row)

    wb.save(output_path)
    return output_path


def build_dataset_workbook(
    input_path: str | Path,
    output: str | Path,
    *,
    title: str,
    include_summary: bool = True,
    include_charts: bool = False,
) -> Path:
    """Build from actual uploaded columns instead of falling back to an unrelated template."""

    source_df = read_table(input_path)
    cleaned_df, clean_report = clean_dataframe(source_df)
    output_path = ensure_output_path(output, "数据分析结果.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = wb.create_sheet("说明")
    instructions["A1"] = title
    instructions["A3"] = "数据来源"
    instructions["B3"] = Path(input_path).name
    instructions["A4"] = "处理内容"
    instructions["B4"] = "保留原始数据，另建清洗数据，并按实际字段生成可筛选的结果。"
    instructions["A5"] = "清洗结果"
    instructions["B5"] = (
        f"原始 {clean_report['original_rows']} 行，清洗后 {clean_report['cleaned_rows']} 行，"
        f"删除重复 {clean_report['duplicate_rows_removed']} 行。"
    )
    style_instruction_sheet(instructions)

    _write_dataframe_sheet(wb.create_sheet("原始数据"), source_df, "原始数据")
    clean_ws = wb.create_sheet("清洗数据")
    _write_dataframe_sheet(clean_ws, cleaned_df, "清洗数据")

    if include_summary:
        summary = wb.create_sheet("汇总")
        _write_dataset_summary(summary, cleaned_df, clean_report)
        if include_charts:
            _add_dataset_chart(summary, clean_ws, cleaned_df)

    report_ws = wb.create_sheet("清洗报告")
    _write_clean_report(report_ws, clean_report)
    wb.save(output_path)
    return output_path


def _formula_for_rule(
    rule: dict[str, Any],
    row: int,
    columns: dict[str, int],
) -> str:
    kind = str(rule.get("kind", ""))
    sources = [str(item) for item in rule.get("sources", [])]
    refs = [
        f"{get_column_letter(columns[source])}{row}"
        for source in sources
        if source in columns
    ]
    if kind == "average" and refs:
        return f'=IF(COUNTA({",".join(refs)})=0,"",IFERROR(AVERAGE({",".join(refs)}),""))'
    if kind == "difference" and len(refs) >= 2:
        return f'=IFERROR({refs[0]}-{refs[1]},"")'
    if kind == "product" and len(refs) >= 2:
        return f'=IFERROR({refs[0]}*{refs[1]},"")'
    if kind == "ratio" and len(refs) >= 2:
        return f'=IFERROR({refs[0]}/{refs[1]},"")'
    if kind == "sum" and refs:
        return f'=IFERROR(SUM({",".join(refs)}),"")'
    if kind == "weather_advice" and len(refs) >= 2:
        options = dict(rule.get("options") or {})
        rain_threshold = float(options.get("rain_threshold", 0.5))
        heat_threshold = float(options.get("heat_threshold", 30))
        rain_text = str(options.get("rain_text", "建议携带雨具"))
        heat_text = str(options.get("heat_text", "注意防暑"))
        default_text = str(options.get("default_text", "适宜出行"))
        return (
            f'=IF({refs[0]}="","",IF({refs[0]}>={rain_threshold},'
            f'"{rain_text}",IF({refs[1]}>={heat_threshold},"{heat_text}","{default_text}")))'
        )
    return '=""'


def _write_inline_summary(
    ws,
    plan: dict[str, Any],
    columns: dict[str, int],
    data_start: int,
    data_end: int,
) -> None:
    rules = [item for item in plan.get("summary_rules", []) if isinstance(item, dict)]
    if not rules:
        return
    summary_header = data_end + 2
    ws.cell(summary_header, 1, "汇总项目")
    ws.cell(summary_header, 2, "结果")
    apply_header_style(ws, summary_header)
    for offset, rule in enumerate(rules, start=summary_header + 1):
        ws.cell(offset, 1, str(rule.get("label", "汇总")))
        kind = str(rule.get("kind", ""))
        value_col = str(rule.get("value_col") or rule.get("source_col") or "")
        value_index = columns.get(value_col)
        if kind in {"averageif", "sumif", "countif"}:
            group_col = str(rule.get("group_col", ""))
            group_index = columns.get(group_col)
            group_value = str(rule.get("group_value", "")).replace('"', '""')
            if group_index and value_index:
                group_range = (
                    f"${get_column_letter(group_index)}${data_start}:"
                    f"${get_column_letter(group_index)}${data_end}"
                )
                value_range = (
                    f"${get_column_letter(value_index)}${data_start}:"
                    f"${get_column_letter(value_index)}${data_end}"
                )
                function = {"averageif": "AVERAGEIF", "sumif": "SUMIF", "countif": "COUNTIF"}[kind]
                if kind == "countif":
                    formula = f'=COUNTIF({group_range},"{group_value}")'
                else:
                    formula = (
                        f'=IFERROR({function}({group_range},"{group_value}",{value_range}),0)'
                    )
                ws.cell(offset, 2, formula)
        elif value_index:
            value_range = (
                f"{get_column_letter(value_index)}{data_start}:"
                f"{get_column_letter(value_index)}{data_end}"
            )
            function = {"average": "AVERAGE", "sum": "SUM", "count": "COUNT"}.get(kind)
            if function:
                ws.cell(offset, 2, f"=IFERROR({function}({value_range}),0)")
        ws.cell(offset, 2).number_format = "0.00"
    apply_summary_style(ws, summary_header + 1, summary_header + len(rules), 1, 2)


def _add_custom_chart(
    ws,
    columns: list[dict[str, Any]],
    start_row: int,
    end_row: int,
) -> None:
    if end_row < start_row:
        return
    numeric = [
        index
        for index, item in enumerate(columns, start=1)
        if item.get("kind") in {"number", "money"} and item.get("role") != "formula"
    ]
    if not numeric:
        numeric = [
            index
            for index, item in enumerate(columns, start=1)
            if item.get("kind") in {"number", "money", "percentage"}
        ]
    category = next(
        (
            index
            for index, item in enumerate(columns, start=1)
            if item.get("kind") in {"date", "text"}
        ),
        None,
    )
    if not numeric or not category:
        return
    is_date = columns[category - 1].get("kind") == "date"
    chart = LineChart() if is_date else BarChart()
    if isinstance(chart, BarChart):
        chart.type = "col"
        chart.grouping = "clustered"
        chart.overlap = -10
        chart.gapWidth = 80
    chart.title = str(ws["A1"].value)
    chart.style = 10
    chart.height = 7.5
    chart.width = 15
    cats = Reference(ws, min_col=category, min_row=start_row, max_row=end_row)
    series_cols = numeric[:4]
    for col in series_cols:
        data = Reference(ws, min_col=col, min_row=start_row - 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if len(series_cols) == 1:
        chart.legend = None
    ws.add_chart(chart, f"{get_column_letter(len(columns) + 2)}3")


def _apply_custom_number_formats(
    ws,
    columns: list[dict[str, Any]],
    start_row: int,
    end_row: int,
) -> None:
    formats = {
        "date": "yyyy-mm-dd",
        "time": "hh:mm",
        "number": "0.00",
        "money": '#,##0.00;[Red]-#,##0.00',
        "percentage": "0%",
    }
    for index, item in enumerate(columns, start=1):
        number_format = formats.get(item.get("kind"))
        if not number_format:
            continue
        for row in range(start_row, end_row + 1):
            ws.cell(row, index).number_format = number_format


def _coerce_value(value: Any, kind: str, *, default_year: int | None = None) -> Any:
    if value in (None, ""):
        return ""
    if kind == "percentage":
        if isinstance(value, str) and value.strip().endswith("%"):
            try:
                return float(value.strip().rstrip("%")) / 100
            except ValueError:
                return value
        try:
            numeric = float(value)
            return numeric / 100 if numeric > 1 else numeric
        except (TypeError, ValueError):
            return value
    if kind in {"number", "money"}:
        try:
            return float(value)
        except (TypeError, ValueError):
            return value
    if kind == "date":
        text = str(value).strip()
        if default_year and re.fullmatch(r"\d{1,2}月\d{1,2}日", text):
            match = re.fullmatch(r"(\d{1,2})月(\d{1,2})日", text)
            if match:
                return datetime(default_year, int(match.group(1)), int(match.group(2)))
        parsed = pd.to_datetime(text, errors="coerce")
        return parsed.to_pydatetime() if pd.notna(parsed) else value
    return value


def _write_dataframe_sheet(ws, df: pd.DataFrame, title: str) -> None:
    max_col = max(1, len(df.columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = title
    apply_title_style(ws, 1, 1, max_col)
    for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row_index, col_index, value)
    apply_header_style(ws, 3)
    if ws.max_row >= 3:
        freeze_and_filter(
            ws,
            "A4",
            f"A3:{get_column_letter(ws.max_column)}{ws.max_row}",
        )
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _write_inline_table_sheet(ws, table: dict[str, Any], title: str) -> None:
    columns = [str(item) for item in table.get("columns", [])]
    records = [dict(item) for item in table.get("records", []) if isinstance(item, dict)]
    max_col = max(1, len(columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = title
    apply_title_style(ws, 1, 1, max_col)
    for col, name in enumerate(columns, start=1):
        ws.cell(3, col, name)
    apply_header_style(ws, 3)
    for row_index, record in enumerate(records, start=4):
        for col_index, name in enumerate(columns, start=1):
            ws.cell(row_index, col_index, record.get(name, ""))
    if records:
        freeze_and_filter(
            ws,
            "A4",
            f"A3:{get_column_letter(max_col)}{3 + len(records)}",
        )
    for col_index, name in enumerate(columns, start=1):
        kind = infer_column_kind(name)
        number_format = {
            "money": '#,##0.00;[Red]-#,##0.00',
            "percentage": "0.00%",
            "date": "yyyy-mm-dd",
            "number": '#,##0.00;[Red]-#,##0.00',
        }.get(kind)
        if number_format:
            for row_index in range(4, 4 + len(records)):
                ws.cell(row_index, col_index).number_format = number_format
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _unique_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", str(value)).strip()[:31] or "数据"
    if base not in used:
        return base
    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            return candidate
        counter += 1


def _write_dataset_summary(ws, df: pd.DataFrame, report: dict[str, Any]) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = "数据概览"
    apply_title_style(ws, 1, 1, 4)
    ws["A3"] = "项目"
    ws["B3"] = "结果"
    apply_header_style(ws, 3)
    rows = [
        ("清洗后记录数", "=MAX(COUNTA('清洗数据'!A:A)-1,0)"),
        ("字段数", len(df.columns)),
        ("删除重复行", int(report.get("duplicate_rows_removed", 0))),
        ("缺失值总数", int(sum(report.get("missing_values", {}).values()))),
    ]
    for row, (label, value) in enumerate(rows, start=4):
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
    apply_summary_style(ws, 4, 7, 1, 2)
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _add_dataset_chart(summary_ws, clean_ws, df: pd.DataFrame) -> None:
    if df.empty:
        return
    numeric_cols = [
        index
        for index, column in enumerate(df.columns, start=1)
        if pd.api.types.is_numeric_dtype(df[column])
    ]
    if not numeric_cols:
        return
    category_col = 1
    chart = BarChart()
    chart.title = f"{df.columns[numeric_cols[0] - 1]}概览"
    max_row = min(clean_ws.max_row, 23)
    data = Reference(
        clean_ws,
        min_col=numeric_cols[0],
        min_row=3,
        max_row=max_row,
    )
    cats = Reference(clean_ws, min_col=category_col, min_row=4, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    summary_ws.add_chart(chart, "D3")


def _write_clean_report(ws, report: dict[str, Any]) -> None:
    ws["A1"] = "清洗报告"
    apply_title_style(ws, 1, 1, 3)
    ws["A3"] = "检查项"
    ws["B3"] = "字段"
    ws["C3"] = "结果"
    apply_header_style(ws, 3)
    row = 4
    for key, value in report.items():
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                ws.cell(row, 1, key)
                ws.cell(row, 2, sub_key)
                ws.cell(row, 3, sub_value)
                row += 1
        else:
            ws.cell(row, 1, key)
            ws.cell(row, 3, value)
            row += 1
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)
