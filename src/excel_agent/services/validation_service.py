"""Adapter around the existing deterministic workbook validator."""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..task_paths import TaskPaths, append_run_log_event
from ..task_spec import TaskSpec
from ..validators import validate_workbook


@dataclass
class ValidationResult:
    status: str
    issues: list[dict[str, Any]]
    warnings: list[dict[str, Any]]
    suggestions: list[str]
    report_file: str
    summary: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def validate_generated_workbook(
    output_file: str | Path,
    task_spec: TaskSpec,
    task_paths: TaskPaths,
) -> ValidationResult:
    append_run_log_event(
        task_paths,
        event="validation_started",
        status="running",
        details={"output_file": str(output_file), "validator": "excel_agent.validators.validate_workbook"},
    )
    try:
        report = validate_workbook(output_file, task_paths.validation_report)
        _apply_task_fidelity_checks(report, output_file, task_spec)
        report["summary"]["error_count"] = len(report.get("errors", []))
        report["summary"]["warning_count"] = len(report.get("warnings", []))
        report["status"] = (
            "fail"
            if report["errors"]
            else "warn"
            if report["warnings"]
            else "pass"
        )
        task_paths.validation_report.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        result = ValidationResult(
            status=str(report.get("status", "error")),
            issues=list(report.get("errors", [])),
            warnings=list(report.get("warnings", [])),
            suggestions=[str(item) for item in report.get("suggestions", [])],
            report_file=str(task_paths.validation_report),
            summary=dict(report.get("summary", {})),
        )
        append_run_log_event(
            task_paths,
            event="validation_completed",
            status=result.status,
            details={
                "report_file": result.report_file,
                "error_count": len(result.issues),
                "warning_count": len(result.warnings),
                "task_type": task_spec.task_type,
            },
        )
        return result
    except Exception as exc:
        error_item = {
            "check": "validation_service",
            "message": f"{type(exc).__name__}: {exc}",
        }
        fallback_report = {
            "status": "error",
            "file": str(output_file),
            "summary": {},
            "errors": [error_item],
            "warnings": [],
            "suggestions": ["检查生成文件和本地 Python 依赖后重新运行校验。"],
        }
        task_paths.validation_report.write_text(
            json.dumps(fallback_report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        result = ValidationResult(
            status="error",
            issues=[error_item],
            warnings=[],
            suggestions=fallback_report["suggestions"],
            report_file=str(task_paths.validation_report),
            summary={},
        )
        append_run_log_event(
            task_paths,
            event="validation_failed",
            status="error",
            details=result.to_dict(),
        )
        return result


def _apply_task_fidelity_checks(
    report: dict[str, Any],
    output_file: str | Path,
    task_spec: TaskSpec,
) -> None:
    agent_plan = task_spec.options.get("agent_blueprint")
    content_plan = task_spec.options.get("content_plan", {})
    if isinstance(agent_plan, dict) and agent_plan.get("sheets"):
        sheet_plans = [
            item for item in agent_plan.get("sheets", []) if isinstance(item, dict)
        ]
        primary = max(
            sheet_plans,
            key=lambda item: len(item.get("records", [])),
            default={},
        )
        plan = {
            "title": agent_plan.get("title") or primary.get("title"),
            "columns": [
                {"name": item.get("label")}
                for item in primary.get("columns", [])
                if isinstance(item, dict)
            ],
            "expected_data_rows": len(primary.get("records", [])),
            "expected_sheet_names": [
                str(item.get("sheet_name") or "") for item in sheet_plans
            ],
            "explicit_structure": True,
        }
    elif isinstance(agent_plan, dict) and agent_plan.get("columns"):
        plan = {
            "title": agent_plan.get("title"),
            "columns": [
                {"name": item.get("label")}
                for item in agent_plan.get("columns", [])
                if isinstance(item, dict)
            ],
            "expected_data_rows": len(agent_plan.get("records", [])),
            "explicit_structure": True,
        }
    else:
        plan = content_plan
    try:
        wb = load_workbook(output_file, data_only=False)
    except Exception:
        return
    if task_spec.include_charts and sum(len(ws._charts) for ws in wb.worksheets) == 0:
        _fidelity_error(report, "requirement_chart", "用户要求图表，但工作簿中未找到图表。")
    if "公式" in task_spec.user_goal:
        formula_count = sum(
            1
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        if formula_count == 0:
            _fidelity_error(
                report,
                "requirement_formula",
                "用户明确要求使用公式，但工作簿中没有公式。",
            )
    _suppress_expected_merge_warnings(report, task_spec)
    if not isinstance(plan, dict) or not plan.get("explicit_structure"):
        report["summary"]["requirement_fidelity_checked"] = True
        return
    expected_columns = [
        str(item.get("name", "")).strip()
        for item in plan.get("columns", [])
        if str(item.get("name", "")).strip()
    ]
    expected_title = str(plan.get("title", "")).strip()
    target = None
    detected_headers: list[str] = []
    header_row = None
    best_overlap = 0
    for ws in wb.worksheets:
        if ws.title.strip().lower() in {"说明", "instructions", "readme", "使用说明"}:
            continue
        preferred_sheet = str(plan.get("expected_primary_sheet") or "").strip()
        if preferred_sheet and ws.title != preferred_sheet:
            continue
        for row in range(1, min(ws.max_row, 12) + 1):
            row_values = [
                str(ws.cell(row, column).value).strip()
                for column in range(1, ws.max_column + 1)
                if ws.cell(row, column).value not in (None, "")
            ]
            next_values = (
                [
                    str(ws.cell(row + 1, column).value).strip()
                    for column in range(1, ws.max_column + 1)
                    if ws.cell(row + 1, column).value not in (None, "")
                ]
                if row < ws.max_row
                else []
            )
            values = list(dict.fromkeys([*row_values, *next_values]))
            overlap = sum(
                any(_labels_match(expected, found) for found in values)
                for expected in expected_columns
            )
            required_overlap = min(
                len(expected_columns),
                max(2, min(4, len(expected_columns))),
            )
            if (
                expected_columns
                and overlap >= required_overlap
                and overlap > best_overlap
            ):
                target = ws
                detected_headers = values
                header_row = row
                best_overlap = overlap

    if target is None:
        _fidelity_error(
            report,
            "requirement_columns",
            "生成结果中没有找到用户要求的表头。",
        )
        return

    missing = [
        name
        for name in expected_columns
        if not any(_labels_match(name, found) for found in detected_headers)
    ]
    if missing:
        _fidelity_error(
            report,
            "requirement_columns",
            f"生成结果缺少用户要求的列：{'、'.join(missing)}。",
        )

    if expected_title:
        visible_titles = [
            str(ws["A1"].value or "").strip() for ws in wb.worksheets if ws.sheet_state == "visible"
        ]
        if not any(expected_title in value or value in expected_title for value in visible_titles if value):
            _fidelity_warning(
                report,
                "requirement_title",
                f"工作簿标题与需求标题“{expected_title}”不一致。",
            )

    expected_rows = int(plan.get("expected_data_rows") or 0)
    rows_by_sheet = _normalize_rows_by_sheet(plan.get("expected_data_rows_by_sheet"))
    if rows_by_sheet:
        for sheet_name, sheet_expected_rows in rows_by_sheet.items():
            if sheet_name not in wb.sheetnames:
                continue
            actual_rows = _count_data_rows(wb[sheet_name])
            if actual_rows < sheet_expected_rows:
                _fidelity_error(
                    report,
                    "requirement_rows",
                    f"{sheet_name} 需求为 {sheet_expected_rows} 条数据，但实际只有 {actual_rows} 条。",
                )
        expected_rows = 0
    if expected_rows and header_row:
        actual_rows = sum(
            1
            for row in range(header_row + 1, target.max_row + 1)
            if any(target.cell(row, col).value not in (None, "") for col in range(1, target.max_column + 1))
        )
        if actual_rows < expected_rows:
            _fidelity_error(
                report,
                "requirement_rows",
                f"需求中识别到 {expected_rows} 条数据，但表格只有 {actual_rows} 条。",
            )

    expected_sheet_names = [
        str(item).strip()
        for item in plan.get("expected_sheet_names", [])
        if str(item).strip()
    ]
    should_check_sheet_names = bool(
        plan.get("strict_sheet_names")
        or plan.get("consolidated_inline_tables")
        or plan.get("layout") == "multi_sheet"
        or len(expected_sheet_names) > 1
    )
    if expected_sheet_names and should_check_sheet_names:
        visible_names = [
            ws.title for ws in wb.worksheets if ws.sheet_state == "visible"
        ]
        missing_sheets = [
            name for name in expected_sheet_names if name not in visible_names
        ]
        if missing_sheets:
            _fidelity_error(
                report,
                "requirement_sheets",
                f"生成结果缺少用户要求的工作表：{'、'.join(missing_sheets)}。",
            )
        if len(visible_names) < len(expected_sheet_names):
            _fidelity_error(
                report,
                "requirement_sheets",
                f"需求中识别到 {len(expected_sheet_names)} 个数据表，"
                f"但工作簿只有 {len(visible_names)} 个可见工作表。",
            )
        if plan.get("strict_sheet_names"):
            extra_sheets = [
                name for name in visible_names if name not in expected_sheet_names
            ]
            if extra_sheets:
                _fidelity_error(
                    report,
                    "requirement_sheets",
                    f"生成结果包含未要求的工作表：{'、'.join(extra_sheets)}。",
                )

    report["summary"]["requirement_fidelity_checked"] = True
    report["summary"]["expected_column_count"] = len(expected_columns)


def _normalize_rows_by_sheet(value: Any) -> dict[str, int]:
    if not isinstance(value, dict):
        return {}
    result: dict[str, int] = {}
    for key, raw in value.items():
        name = str(key or "").strip()
        if not name:
            continue
        try:
            count = int(raw)
        except (TypeError, ValueError):
            continue
        if count > 0:
            result[name] = count
    return result


def _count_data_rows(ws) -> int:
    header_row = None
    if ws.auto_filter.ref:
        try:
            header_row = int(str(ws.auto_filter.ref).split(":")[0][1:])
        except (ValueError, IndexError):
            header_row = None
    if header_row is None:
        for row in range(1, min(ws.max_row, 12) + 1):
            non_empty = sum(
                1
                for col in range(1, ws.max_column + 1)
                if ws.cell(row, col).value not in (None, "")
            )
            if non_empty >= 2:
                header_row = row
                break
    if header_row is None:
        return 0
    return sum(
        1
        for row in range(header_row + 1, ws.max_row + 1)
        if any(
            ws.cell(row, col).value not in (None, "")
            for col in range(1, ws.max_column + 1)
        )
    )


def _suppress_expected_merge_warnings(
    report: dict[str, Any],
    task_spec: TaskSpec,
) -> None:
    goal = str(task_spec.user_goal or "")
    if not ("合并" in goal and "专业" in goal):
        return
    report["warnings"] = [
        item
        for item in report.get("warnings", [])
        if not (
            item.get("check") == "merged_filter_area"
            and str(item.get("message", "")).startswith("合并单元格 C")
            and item.get("sheet") in {"学生信息", "成绩录入", "学期总评"}
        )
    ]


def _fidelity_error(report: dict[str, Any], check: str, message: str) -> None:
    report["errors"].append({"check": check, "message": message})
    report["status"] = "fail"
    if "请根据检查结果修改后重新生成。" not in report["suggestions"]:
        report["suggestions"].append("请根据检查结果修改后重新生成。")


def _fidelity_warning(report: dict[str, Any], check: str, message: str) -> None:
    report["warnings"].append({"check": check, "message": message})
    if report.get("status") == "pass":
        report["status"] = "warn"


def _labels_match(expected: str, actual: str) -> bool:
    def normalize(value: str) -> str:
        return (
            str(value)
            .replace("（", "(")
            .replace("）", ")")
            .replace(" ", "")
            .lower()
        )

    wanted = normalize(expected)
    found = normalize(actual)
    if wanted == found:
        return True
    if not wanted or not found:
        return False
    if max(len(wanted), len(found)) > min(len(wanted), len(found)) * 2 + 4:
        return False
    return wanted in found or found in wanted
