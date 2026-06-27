"""用本机 LibreOffice 无头真算工作簿，暴露运行时才发作的公式错误。

静态校验（validators）只看公式长相和已有缓存，抓不到“真算才出现”的错误：跨表统计漏了
``工作表名!`` 前缀导致在本表内自引用、单元格条件引用错位成循环引用等，写出来都像模像样，
一算才报 #VALUE!。这里把工作簿交给 LibreOffice 真算一遍、读回算出的值，挑出错误单元格，
反馈给智能体修复。没装 LibreOffice、或真算超时/异常时一律优雅降级（不阻塞出表）。
"""

from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..validators import ERROR_VALUES as _STATIC_ERROR_VALUES

# LibreOffice 对循环引用会给出 Err:522；其余沿用 Excel 标准错误值。
ERROR_VALUES = set(_STATIC_ERROR_VALUES) | {"Err:522", "#CIRCULAR!"}

_SOFFICE_CANDIDATES = (
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/usr/bin/soffice",
    "/usr/local/bin/soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
)


def find_soffice() -> str | None:
    """定位 LibreOffice 可执行文件；找不到返回 None（调用方据此降级）。"""

    for name in ("soffice", "soffice.exe", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found
    for path in _SOFFICE_CANDIDATES:
        if Path(path).exists():
            return path
    return None


def recalc_available() -> bool:
    return find_soffice() is not None


def recalc_workbook(path: str | Path, *, timeout: int = 120) -> dict[str, Any]:
    """真算 ``path`` 并返回 {available, ok, error_cells, detail}。

    ``ok`` 的语义偏“放行”：只有在真算确实跑通、且发现了错误单元格时才为 False；
    没装 LibreOffice、超时、异常、未产出文件等情况一律 ``ok=True``（宁可放过也不卡死出表）。
    """

    path = Path(path)
    soffice = find_soffice()
    if soffice is None:
        return {"available": False, "ok": True, "error_cells": [], "detail": "未检测到 LibreOffice，跳过真算。"}
    if not path.exists():
        return {"available": True, "ok": True, "error_cells": [], "detail": "文件不存在，跳过真算。"}

    work = Path(tempfile.mkdtemp(prefix="excel_recalc_"))
    profile = (work / "profile").resolve().as_uri()
    outdir = work / "out"
    outdir.mkdir(parents=True, exist_ok=True)
    try:
        proc = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile}",
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(outdir),
                str(path),
            ],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        produced = list(outdir.glob("*.xlsx"))
        if not produced:
            return {
                "available": True,
                "ok": True,
                "error_cells": [],
                "detail": f"真算未产出文件，跳过（returncode={proc.returncode}）。",
            }
        cached = load_workbook(produced[0], data_only=True)
        error_cells: list[dict[str, str]] = []
        for sheet_name in cached.sheetnames:
            for row in cached[sheet_name].iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value in ERROR_VALUES:
                        error_cells.append(
                            {"sheet": sheet_name, "cell": cell.coordinate, "value": cell.value}
                        )
        return {
            "available": True,
            "ok": not error_cells,
            "error_cells": error_cells,
            "detail": "真算通过，无错误值。" if not error_cells else f"真算发现 {len(error_cells)} 个错误单元格。",
        }
    except subprocess.TimeoutExpired:
        return {"available": True, "ok": True, "error_cells": [], "detail": "真算超时，跳过。"}
    except Exception as exc:  # 真算本身失败绝不阻塞出表
        return {"available": True, "ok": True, "error_cells": [], "detail": f"真算异常，跳过：{exc}"}
    finally:
        shutil.rmtree(work, ignore_errors=True)


def describe_error_cells(error_cells: list[dict[str, str]], limit: int = 20) -> str:
    """把错误单元格列成给智能体看的简短清单。"""

    shown = error_cells[:limit]
    parts = [f"{item['sheet']}!{item['cell']}={item['value']}" for item in shown]
    text = "、".join(parts)
    if len(error_cells) > limit:
        text += f" 等共 {len(error_cells)} 处"
    return text
