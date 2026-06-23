"""Deterministic bridge from a confirmed TaskSpec to the existing Excel core."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from ..api_settings import ApiSettings
from ..custom_workbook_builder import (
    build_custom_workbook,
    build_dataset_workbook,
    build_inline_tables_workbook,
)
from ..domain_builders import (
    build_performance_compensation_workbook,
    can_build_performance_compensation,
)
from ..io_utils import read_table
from ..model_registry import get_role_api_settings, load_model_settings
from ..task_paths import TaskPaths, append_run_log_event, stage_input_files
from ..task_spec import TaskSpec, save_task_spec
from ..template_adapter import (
    apply_template_mode,
    inspect_template,
    prepare_template_file,
)
from ..workbook_builder import analyze_sales_file, create_workbook
from .llm_workbook_agent import (
    generate_blueprint_via_json,
    generate_with_llm_agent,
)
from .agent.orchestrator import run_agent


GENERATION_API_VERSION = 2


@dataclass
class GenerationResult:
    success: bool
    output_file: str | None
    message: str
    error: str | None
    used_command: str | None
    mode: str
    notices: list[str]
    agent_tool_calls: int = 0
    agent_rounds: int = 0
    blueprint_file: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def generate_from_task_spec(
    task_spec: TaskSpec,
    task_paths: TaskPaths,
    api_settings: ApiSettings | None = None,
    progress: Callable[[str, str], None] | None = None,
) -> GenerationResult:
    """Generate one workbook without allowing a model to touch cells."""

    append_run_log_event(
        task_paths,
        event="generation_started",
        status="running",
        details={
            "task_type": task_spec.task_type,
            "policy": task_spec.options.get("generation_policy"),
            "model_may_edit_cells": False,
        },
    )
    notices: list[str] = []
    try:
        save_task_spec(task_spec, task_paths.task_spec_file)
        if task_spec.input_files:
            task_spec.input_files = stage_input_files(task_spec.input_files, task_paths)
        if task_spec.template_files:
            task_spec.template_files = stage_input_files(
                task_spec.template_files,
                task_paths,
            )
            prepared_template = prepare_template_file(
                task_spec.template_files[0],
                task_paths.task_dir,
            )
            task_spec.options["prepared_template_file"] = str(prepared_template)
            task_spec.options["template_summary"] = inspect_template(prepared_template)
            if task_spec.options.get("use_template_data"):
                task_spec.input_files.append(str(prepared_template))
                task_spec.options["template_data_added_to_inputs"] = True
        if task_spec.input_files:
            profiles = []
            total_rows = 0
            all_columns: list[str] = []
            for source_path in task_spec.input_files:
                source_frame = read_table(source_path)
                columns = [str(item) for item in source_frame.columns]
                profiles.append(
                    {
                        "file_name": Path(source_path).name,
                        "row_count": int(len(source_frame)),
                        "columns": columns,
                    }
                )
                total_rows += int(len(source_frame))
                all_columns.extend(columns)
            task_spec.options["input_data_profile"] = {
                "files": profiles,
                "row_count": total_rows,
                "columns": list(dict.fromkeys(all_columns)),
                "values_sent_to_model": False,
            }
        save_task_spec(task_spec, task_paths.task_spec_file)

        settings = api_settings or ApiSettings()
        if not (settings.configured and settings.use_for_generation):
            registry_settings = load_model_settings()
            role_settings = get_role_api_settings(
                "builder",
                registry_settings,
                use_for_intent=False,
                use_for_review=False,
                use_for_generation=registry_settings.agent_enabled,
            )
            if role_settings is not None:
                settings = role_settings
        if settings.configured and settings.use_for_generation:
            return _generate_with_model(
                task_spec, task_paths, settings, progress, notices
            )
        return _local_generate(task_spec, task_paths, progress, notices)
    except Exception as exc:
        result = GenerationResult(
            success=False,
            output_file=str(task_paths.output_file) if task_paths.output_file.exists() else None,
            message="电子表格生成失败。",
            error=f"{type(exc).__name__}: {exc}",
            used_command=None,
            mode="error",
            notices=notices,
        )
        append_run_log_event(
            task_paths,
            event="generation_failed",
            status="error",
            details=result.to_dict(),
        )
        return result


def _generate_with_model(
    task_spec: TaskSpec,
    task_paths: TaskPaths,
    settings: ApiSettings,
    progress: Callable[[str, str], None] | None,
    notices: list[str],
) -> GenerationResult:
    """Let the model design the workbook, with two automatic fallbacks.

    1. tool-calling agent (best when the endpoint supports function calls)
    2. plain JSON blueprint (best for models that return JSON, not tool calls)
    3. deterministic local generation, with a visible notice (never silent)

    A usable workbook is always produced unless even local generation cannot
    run (for example, a missing input file).
    """

    # Fast path: when the request already contains structured data tables that we
    # parsed locally, the deterministic builders reproduce them exactly in
    # seconds. A slow reasoning model would only truncate its output on a request
    # this large (and then we'd fall back to exactly this build anyway), so skip
    # the round-trip and build locally up front.
    content_plan = task_spec.options.get("content_plan") or {}
    inline_tables = (
        content_plan.get("inline_tables", []) if isinstance(content_plan, dict) else []
    )
    if inline_tables and not task_spec.input_files:
        if progress:
            progress("build", "已识别需求中的数据表，正在用本地规则快速精确生成……")
        notices.append(
            "已识别需求中的结构化数据表，直接用本地规则精确生成（比让大模型重述更快更稳）。"
        )
        return _local_generate(task_spec, task_paths, progress, notices)

    # If the user gave us a template plus data, the only correct output is one
    # that matches that template (often required for the file to be importable).
    # Skip the model and clone the template directly — fast and exact.
    template_paths_for_fast = list(task_spec.template_files or [])
    template_mode_for_fast = str(task_spec.options.get("template_mode") or "").lower()
    if (
        template_paths_for_fast
        and task_spec.input_files
        and template_mode_for_fast in {"flexible", "strict"}
        ):
        if progress:
            progress("build", "已识别上传的模板与数据，正在按模板精确填入……")
        return _local_generate(task_spec, task_paths, progress, notices)

    registry_settings = load_model_settings()
    builder_settings = get_role_api_settings(
        "builder",
        registry_settings,
        use_for_intent=False,
        use_for_review=False,
        use_for_generation=True,
    )
    if (
        registry_settings.agent_enabled
        and builder_settings is not None
        and _same_model_endpoint(settings, builder_settings)
    ):
        agent_result = run_agent(task_spec, task_paths, progress=progress)
        if agent_result.success:
            template_notices = _apply_selected_template(task_spec, task_paths)
            save_task_spec(task_spec, task_paths.task_spec_file)
            result = GenerationResult(
                success=True,
                output_file=agent_result.output_file,
                message=agent_result.message,
                error=None,
                used_command="excel_agent.services.agent.orchestrator.run_agent",
                mode=agent_result.mode,
                notices=[*notices, *(agent_result.notices or []), *template_notices],
                agent_tool_calls=agent_result.tool_calls,
                agent_rounds=agent_result.steps,
                blueprint_file=agent_result.blueprint_file,
            )
            append_run_log_event(
                task_paths,
                event="generation_completed",
                status="success",
                details=result.to_dict(),
            )
            return result
        notices.append(
            "多步智能体未能完成本次任务，已自动切换到兼容生成路径。"
            f"（反馈：{(agent_result.error or '未知原因')[:120]}）"
        )

    agent = generate_with_llm_agent(task_spec, task_paths, settings, progress=progress)
    if not agent.success:
        if progress:
            progress("model", "换一种方式请大模型给出整体方案……")
        agent = generate_blueprint_via_json(
            task_spec, task_paths, settings, progress=progress
        )

    if agent.success:
        template_notices = _apply_selected_template(task_spec, task_paths)
        save_task_spec(task_spec, task_paths.task_spec_file)
        result = GenerationResult(
            success=True,
            output_file=agent.output_file,
            message=agent.message,
            error=None,
            used_command="excel_agent.services.llm_workbook_agent",
            mode="llm_tool_agent",
            notices=[*notices, *agent.notices, *template_notices],
            agent_tool_calls=agent.tool_calls,
            agent_rounds=agent.rounds,
            blueprint_file=agent.blueprint_file,
        )
        append_run_log_event(
            task_paths,
            event="generation_completed",
            status="success",
            details=result.to_dict(),
        )
        return result

    # Both model paths failed; fall back to local generation, but tell the user.
    if progress:
        progress("build", "大模型这次没能给出可用方案，正在用本地规则稳定生成一版……")
    reason = (agent.error or "未知原因").strip()
    notices.append(
        "这次大模型没有返回可用的表格方案，已自动改用本地规则生成可用表格。"
        f"（模型反馈：{reason[:160]}）"
        "可在“接口设置”确认接口地址和模型名称是否正确，或直接在“继续修改”里调整。"
    )
    append_run_log_event(
        task_paths,
        event="generation_model_fallback",
        status="warning",
        details={"reason": reason},
    )
    result = _local_generate(task_spec, task_paths, progress, notices)
    result.mode = f"local_fallback:{result.mode}"
    return result


def _local_generate(
    task_spec: TaskSpec,
    task_paths: TaskPaths,
    progress: Callable[[str, str], None] | None,
    notices: list[str],
) -> GenerationResult:
    """Deterministic, no-model generation. Raises only if no file can be made."""

    if progress:
        progress("build", "正在用本地规则生成表格……")
    mode = "standard_template"
    used_command = "excel_agent.workbook_builder.create_workbook"
    content_plan = task_spec.options.get("content_plan", {})
    inline_tables = (
        content_plan.get("inline_tables", [])
        if isinstance(content_plan, dict)
        else []
    )
    # Template-faithful branch: the user gave us a template plus data, and
    # expects the result to match the template exactly (often required so the
    # file can be imported into another system). Don't invent a structure -
    # clone the template and fill its sheets.
    template_paths = list(task_spec.template_files or [])
    template_mode = str(task_spec.options.get("template_mode") or "").lower()
    if (
        template_paths
        and task_spec.input_files
        and template_mode in {"flexible", "strict"}
    ):
        from ..template_filler import fill_template

        mode = "template_faithful"
        used_command = "excel_agent.template_filler.fill_template"
        if progress:
            progress("build", "正在按模板结构填入数据……")
        result = fill_template(
            template_paths[0],
            task_spec.input_files,
            task_paths.output_file,
            prompt=task_spec.user_goal,
            task_dir=task_paths.task_dir,
        )
        notices.append(
            "已严格按上传模板生成结果（保留全部工作表、表头和样式），"
            "数据按列名自动填入；这样得到的文件可以直接用于原系统导入。"
        )
        for sheet in result.get("sheets_filled", [])[:6]:
            notices.append(f"已填充工作表：{sheet}")
        notices.extend(result.get("notices", []))
        return _wrap_result(
            task_spec, task_paths, mode, used_command, notices, progress
        )

    if (
        inline_tables
        and not task_spec.input_files
        and can_build_performance_compensation(task_spec.user_goal, content_plan)
    ):
        mode = "domain_compiler:performance_compensation"
        used_command = (
            "excel_agent.domain_builders.build_performance_compensation_workbook"
        )
        build_performance_compensation_workbook(
            content_plan,
            task_spec.user_goal,
            task_paths.output_file,
        )
        content_plan["expected_sheet_names"] = ["参数表", "明细表"]
        content_plan["consolidated_inline_tables"] = True
        task_spec.options["content_plan"] = content_plan
        save_task_spec(task_spec, task_paths.task_spec_file)
        notices.append(
            "模型方案不可用时，已使用本地绩效薪酬业务编译器生成参数表、"
            "明细表、跨表公式、排名和汇总。结果需要人工复核。"
        )
    elif inline_tables and not task_spec.input_files:
        mode = "inline_multi_table"
        used_command = (
            "excel_agent.custom_workbook_builder.build_inline_tables_workbook"
        )
        build_inline_tables_workbook(
            content_plan,
            task_paths.output_file,
            include_charts=task_spec.include_charts,
        )
        notices.append(
            f"已从需求文字中识别并保留 {len(inline_tables)} 个数据表；"
            "未使用无关标准模板。"
        )
    elif (
        isinstance(content_plan, dict)
        and content_plan.get("explicit_structure")
        and content_plan.get("columns")
        and not task_spec.input_files
    ):
        mode = "custom_content"
        used_command = "excel_agent.custom_workbook_builder.build_custom_workbook"
        build_custom_workbook(
            content_plan,
            task_paths.output_file,
            include_charts=task_spec.include_charts,
        )
    elif task_spec.task_type == "sales_report" and task_spec.input_files:
        mode = "sales_input_analysis"
        used_command = "excel_agent.workbook_builder.analyze_sales_file"
        analyze_sales_file(task_spec.input_files[0], task_paths.output_file)
    elif task_spec.input_files:
        mode = "input_dataset"
        used_command = "excel_agent.custom_workbook_builder.build_dataset_workbook"
        build_dataset_workbook(
            task_spec.input_files[0],
            task_paths.output_file,
            title=content_plan.get("title") or task_spec.user_goal[:60],
            include_summary=task_spec.include_summary,
            include_charts=task_spec.include_charts,
        )
        notices.append("已按上传文件的实际字段生成结果。")
    else:
        create_workbook(task_spec.task_type, task_paths.output_file)
        if task_spec.task_type == "sales_report":
            notices.append(
                "未提供销售数据文件，已生成销售报表示例，可在其中填入真实数据。"
            )

    notices.extend(_apply_task_options(task_paths.output_file, task_spec))
    notices.extend(_apply_selected_template(task_spec, task_paths))
    if not task_paths.output_file.exists():
        raise FileNotFoundError(f"生成内核未产生预期文件: {task_paths.output_file}")

    result = GenerationResult(
        success=True,
        output_file=str(task_paths.output_file),
        message="表格已生成。",
        error=None,
        used_command=used_command,
        mode=mode,
        notices=notices,
    )
    append_run_log_event(
        task_paths,
        event="generation_completed",
        status="success",
        details=result.to_dict(),
    )
    return result


def _wrap_result(
    task_spec: TaskSpec,
    task_paths: TaskPaths,
    mode: str,
    used_command: str,
    notices: list[str],
    progress: Callable[[str, str], None] | None,
) -> GenerationResult:
    """Common tail used by template-faithful and other early-return paths."""

    if not task_paths.output_file.exists():
        raise FileNotFoundError(f"生成内核未产生预期文件: {task_paths.output_file}")
    result = GenerationResult(
        success=True,
        output_file=str(task_paths.output_file),
        message="表格已生成。",
        error=None,
        used_command=used_command,
        mode=mode,
        notices=notices,
    )
    append_run_log_event(
        task_paths,
        event="generation_completed",
        status="success",
        details=result.to_dict(),
    )
    return result


def _apply_task_options(output_file: Path, task_spec: TaskSpec) -> list[str]:
    """Apply small deterministic presentation choices after core generation."""

    notices: list[str] = []
    wb = load_workbook(output_file)
    changed = False

    if not task_spec.include_charts:
        for ws in wb.worksheets:
            if ws._charts:
                ws._charts = []
                changed = True

    for summary_name in ("Summary", "汇总"):
        if task_spec.include_summary or summary_name not in wb.sheetnames:
            continue
        if task_spec.task_type == "dashboard":
            notices.append("综合仪表盘依赖汇总页，已为避免公式失效保留汇总页。")
        else:
            wb.remove(wb[summary_name])
            changed = True

    for instruction_name in ("Instructions", "说明"):
        if (
            not task_spec.include_instructions_sheet
            and instruction_name in wb.sheetnames
            and len(wb.sheetnames) > 1
        ):
            wb.remove(wb[instruction_name])
            changed = True

    if changed:
        wb.save(output_file)
    return notices


def _same_model_endpoint(left: ApiSettings, right: ApiSettings) -> bool:
    """Avoid accidentally using a different saved provider than the caller chose."""

    return (
        left.configured
        and right.configured
        and left.base_url == right.base_url
        and left.model == right.model
        and left.api_key == right.api_key
    )


def _apply_selected_template(
    task_spec: TaskSpec,
    task_paths: TaskPaths,
) -> list[str]:
    template_path = str(task_spec.options.get("prepared_template_file") or "")
    if not template_path:
        return []
    mode = str(task_spec.options.get("template_mode") or "reference")
    apply_template_mode(
        task_paths.output_file,
        template_path,
        mode=mode,
        blueprint=task_spec.options.get("agent_blueprint"),
        use_template_data=bool(task_spec.options.get("use_template_data")),
    )
    labels = {
        "reference": "已参考模板的字体、表头样式、列宽和页面设置。",
        "flexible": "已灵活套用模板形式；与本次需求冲突的内容以需求为准。",
        "strict": "已按严格模板模式写入；模板字段顺序和工作表结构保持不变。",
    }
    notice = labels.get(mode, labels["reference"])
    if not task_spec.options.get("use_template_data"):
        notice += " 模板中的示例数据未带入结果。"
    else:
        notice += " 模板中的现有数据已作为额外数据源。"
    return [notice]
