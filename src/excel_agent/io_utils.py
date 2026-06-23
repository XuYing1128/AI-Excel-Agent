"""File and dataframe IO helpers."""

from __future__ import annotations

import json
import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd


INVALID_SHEET_CHARS = r"[\[\]\:\*\?\/\\]"


def project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def ensure_output_path(path: str | Path | None, default_name: str = "workbook.xlsx") -> Path:
    out = Path(path) if path else project_root() / "outputs" / default_name
    if not out.is_absolute():
        out = project_root() / out
    out.parent.mkdir(parents=True, exist_ok=True)
    return out


def backup_file(path: str | Path, suffix: str = ".bak") -> Path:
    source = Path(path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = source.with_name(f"{source.stem}.{timestamp}{suffix}{source.suffix}")
    shutil.copy2(source, backup)
    return backup


def safe_sheet_name(name: str, fallback: str = "Sheet") -> str:
    clean = re.sub(INVALID_SHEET_CHARS, "_", str(name)).strip("' ")
    clean = clean or fallback
    return clean[:31]


def read_table(path: str | Path, sheet_name: str | int | None = 0) -> pd.DataFrame:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(source)
    suffix = source.suffix.lower()
    if suffix in {".csv", ".txt"}:
        return _read_delimited_text(source)
    if suffix in {".tsv"}:
        return pd.read_csv(source, sep="\t")
    if suffix in {".xlsx", ".xlsm"}:
        raw = pd.read_excel(
            source,
            sheet_name=sheet_name,
            engine="openpyxl",
            header=None,
        )
        return _promote_detected_header(raw)
    if suffix == ".xls":
        try:
            raw = pd.read_excel(
                source,
                sheet_name=sheet_name,
                engine="xlrd",
                header=None,
            )
            return _promote_detected_header(raw)
        except ImportError as exc:
            raise RuntimeError(
                "读取 .xls 需要 xlrd。请重新运行 install.bat 安装最新依赖。"
            ) from exc
    raise ValueError(f"Unsupported table file: {source}")


def convert_legacy_xls(path: str | Path, output_dir: str | Path) -> Path:
    """Convert a legacy .xls workbook to .xlsx for template-preserving operations.

    LibreOffice gives the highest-fidelity conversion (styles, merges, print
    setup). When it is not installed, fall back to a basic xlrd -> openpyxl
    conversion that preserves sheet names, headers, values and layout so the
    template can still be used as a reference instead of failing outright.
    """

    source = Path(path).resolve()
    if source.suffix.lower() != ".xls":
        return source
    destination_dir = Path(output_dir).resolve()
    destination_dir.mkdir(parents=True, exist_ok=True)
    soffice = _find_soffice()
    if soffice:
        try:
            return _convert_xls_with_soffice(source, destination_dir, soffice)
        except (subprocess.SubprocessError, OSError, RuntimeError):
            pass  # Fall through to the dependency-free basic conversion.
    return _convert_xls_basic(source, destination_dir)


def _convert_xls_with_soffice(source: Path, destination_dir: Path, soffice: str) -> Path:
    with tempfile.TemporaryDirectory(prefix="ai_excel_xls_") as temporary:
        completed = subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                temporary,
                str(source),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        converted = Path(temporary) / f"{source.stem}.xlsx"
        if not converted.exists():
            detail = (completed.stdout or completed.stderr).strip()
            raise RuntimeError(f".xls 转换失败：{detail}")
        target = destination_dir / converted.name
        shutil.copy2(converted, target)
        return target


def _convert_xls_basic(source: Path, destination_dir: Path) -> Path:
    """Dependency-free .xls -> .xlsx conversion using xlrd (values and layout)."""

    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - xlrd is a declared dependency
        raise RuntimeError(
            "读取 .xls 模板需要 xlrd。请重新运行 install.bat 安装最新依赖。"
        ) from exc
    from openpyxl import Workbook

    destination_dir.mkdir(parents=True, exist_ok=True)
    book = xlrd.open_workbook(str(source))
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in book.sheets():
        ws = wb.create_sheet(safe_sheet_name(sheet.name or "Sheet"))
        for row_index in range(sheet.nrows):
            for col_index in range(sheet.ncols):
                cell = sheet.cell(row_index, col_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                    except (ValueError, OverflowError):
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)
                if value not in (None, ""):
                    ws.cell(row_index + 1, col_index + 1, value)
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    target = destination_dir / f"{source.stem}.xlsx"
    wb.save(target)
    return target


def _read_delimited_text(source: Path) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return pd.read_csv(
                source,
                sep=None,
                engine="python",
                encoding=encoding,
            )
        except (UnicodeDecodeError, pd.errors.ParserError, ValueError) as exc:
            last_error = exc
    raise ValueError(f"无法识别文本数据的编码或分隔符：{source.name}；{last_error}")


def _promote_detected_header(raw: pd.DataFrame) -> pd.DataFrame:
    if raw.empty:
        return raw
    scan = raw.head(40)
    candidates: list[tuple[int, int]] = []
    for index, row in scan.iterrows():
        values = [value for value in row.tolist() if not pd.isna(value) and str(value).strip()]
        if len(values) >= 2:
            candidates.append((len(values), int(index)))
    header_index = min(
        (index for count, index in candidates if count == max(item[0] for item in candidates)),
        default=0,
    ) if candidates else 0
    headers = []
    seen: dict[str, int] = {}
    for position, value in enumerate(raw.iloc[header_index].tolist(), start=1):
        name = str(value).strip() if not pd.isna(value) and str(value).strip() else f"Column_{position}"
        seen[name] = seen.get(name, 0) + 1
        headers.append(name if seen[name] == 1 else f"{name}_{seen[name]}")
    result = raw.iloc[header_index + 1 :].copy()
    result.columns = headers
    return result.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)


def _find_soffice() -> str | None:
    discovered = shutil.which("soffice") or shutil.which("libreoffice")
    if discovered:
        return discovered
    candidates = (
        Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
        Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
    )
    return str(next((item for item in candidates if item.exists()), "")) or None


def save_json(data: dict[str, Any], path: str | Path) -> Path:
    output = ensure_output_path(path, "report.json")
    output.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return output


def normalize_output_extension(path: str | Path, suffix: str = ".xlsx") -> Path:
    p = Path(path)
    return p if p.suffix else p.with_suffix(suffix)
