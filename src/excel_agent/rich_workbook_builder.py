"""Execute a model-authored workbook blueprint with deterministic openpyxl code."""

from __future__ import annotations

import ast
import re
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .io_utils import ensure_output_path


PROHIBITED_FORMULA_TOKENS = {
    "WEBSERVICE",
    "HYPERLINK",
    "RTD",
    "DDE",
    "CALL",
    "EXEC",
    "REGISTER",
    "FILTERXML",
}
ALLOWED_COLUMN_TYPES = {"text", "number", "money", "percentage", "date", "integer"}
ALLOWED_CHART_TYPES = {"bar", "column", "line", "pie"}
THIN_GRAY = Side(style="thin", color="B7C2D0")


def build_rich_workbook(
    blueprint: dict[str, Any],
    output: str | Path,
    *,
    require_charts: bool = False,
) -> Path:
    plan = normalize_blueprint(blueprint)
    output_path = ensure_output_path(output, "智能生成表格.xlsx")
    columns = plan["columns"]
    records = _sort_records(plan["records"], plan.get("sort", []), columns)

    wb = Workbook()
    ws = wb.active
    ws.title = plan["sheet_name"]
    column_map = {item["key"]: index for index, item in enumerate(columns, start=1)}
    max_col = len(columns)

    title_row = 1
    header_top = 3
    header_bottom = 4 if plan.get("header_groups") else 3
    data_start = header_bottom + 1

    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_col)
    title_cell = ws.cell(title_row, 1, plan["title"])
    title_cell.font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="24588A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_row].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    note = "填写说明：基础数据为输入项，自动计算列、汇总和图表会随数据更新。"
    if plan.get("notes"):
        note = f"填写说明：{plan['notes'][0]}"
    ws.cell(2, 1, note)
    ws.cell(2, 1).font = Font(name="Microsoft YaHei", size=9, color="5F6F7F")
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")

    _write_headers(ws, columns, plan.get("header_groups", []), column_map, header_top, header_bottom)

    body_rows: list[int] = []
    body_records: list[dict[str, Any]] = []
    group_rows: dict[str, list[int]] = {}
    current_row = data_start
    group_config = plan.get("group_subtotals")
    group_key = str(group_config.get("group_key", "")) if group_config else ""
    grouped: list[tuple[str | None, list[dict[str, Any]]]]
    if group_config and group_key:
        grouped = _group_records(records, group_key)
    else:
        grouped = [(None, records)]

    for group_value, group_records in grouped:
        group_data_rows: list[int] = []
        for record in group_records:
            _write_record_row(ws, current_row, columns, record, column_map, plan)
            body_rows.append(current_row)
            body_records.append(record)
            group_data_rows.append(current_row)
            current_row += 1
        if group_value is not None and group_config:
            _write_subtotal_row(
                ws,
                current_row,
                group_value,
                group_data_rows,
                columns,
                column_map,
                group_config,
            )
            group_rows[str(group_value)] = group_data_rows
            current_row += 1

    grand_total = plan.get("grand_total")
    if grand_total:
        _write_grand_total_row(
            ws,
            current_row,
            body_rows,
            columns,
            column_map,
            grand_total,
        )
        current_row += 1

    last_row = current_row - 1
    _apply_body_styles(ws, columns, data_start, last_row, plan)
    _apply_conditional_formats(ws, plan.get("conditional_formats", []), column_map, body_rows)

    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_bottom}:{get_column_letter(max_col)}{last_row}"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = f"1:{header_bottom}"
    ws.print_area = f"A1:{get_column_letter(max_col)}{last_row}"
    ws.page_setup.orientation = "landscape" if max_col >= 8 else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "第 &P 页 / 共 &N 页"

    charts = list(plan.get("charts", []))
    if require_charts and not charts:
        charts = [_default_chart(columns)]
    for chart_spec in charts:
        _add_chart(
            ws,
            chart_spec,
            column_map,
            data_start,
            body_rows,
            body_records,
            columns,
        )
    if charts:
        chart_edge = get_column_letter(min(max_col + 14, 40))
        ws.print_area = f"A1:{chart_edge}{max(last_row, 20)}"

    wb.save(output_path)
    return output_path


def normalize_blueprint(blueprint: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(blueprint, dict):
        raise ValueError("工作簿方案必须是 JSON 对象。")
    title = str(blueprint.get("title") or "智能生成表格").strip()[:100]
    sheet_name = re.sub(
        r"[\[\]:*?/\\]",
        "_",
        str(blueprint.get("sheet_name") or title or "数据"),
    )[:31]
    raw_columns = blueprint.get("columns")
    if not isinstance(raw_columns, list) or not raw_columns:
        raise ValueError("工作簿方案必须包含 columns。")
    columns: list[dict[str, Any]] = []
    keys: set[str] = set()
    for index, raw in enumerate(raw_columns[:40], start=1):
        if not isinstance(raw, dict):
            continue
        key = str(raw.get("key") or f"col_{index}").strip()
        if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", key):
            key = f"col_{index}"
        if key in keys:
            raise ValueError(f"列 key 重复: {key}")
        keys.add(key)
        label = str(raw.get("label") or key).strip()[:60]
        column_type = str(raw.get("type") or "text").lower()
        if column_type not in ALLOWED_COLUMN_TYPES:
            column_type = "text"
        columns.append(
            {
                "key": key,
                "label": label,
                "type": column_type,
                "width": min(max(float(raw.get("width") or _default_width(label)), 8), 35),
                "formula": _sanitize_formula_template(raw.get("formula")),
                "number_format": str(raw.get("number_format") or "").strip(),
            }
        )
    records = blueprint.get("records")
    if not isinstance(records, list):
        records = []
    normalized = {
        "title": title,
        "sheet_name": sheet_name or "数据",
        "columns": columns,
        "header_groups": _normalize_header_groups(
            blueprint.get("header_groups"), keys
        ),
        "records": [dict(item) for item in records[:10000] if isinstance(item, dict)],
        "sort": list(blueprint.get("sort") or []),
        "group_subtotals": _normalize_summary_config(
            blueprint.get("group_subtotals"), keys
        ),
        "grand_total": _normalize_summary_config(blueprint.get("grand_total"), keys),
        "conditional_formats": list(blueprint.get("conditional_formats") or []),
        "charts": _normalize_charts(blueprint.get("charts"), keys),
        "notes": [str(item) for item in blueprint.get("notes", []) if str(item).strip()],
    }
    return normalized


def inspect_rich_workbook(path: str | Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    result = {"sheet_count": len(wb.sheetnames), "sheets": []}
    for ws in wb.worksheets:
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
        result["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
                "freeze_panes": str(ws.freeze_panes or ""),
                "auto_filter": ws.auto_filter.ref,
                "chart_count": len(ws._charts),
                "conditional_format_count": sum(
                    len(item.rules) for item in ws.conditional_formatting
                ),
                "formulas": formulas[:80],
                "values": [
                    [ws.cell(row, col).value for col in range(1, min(ws.max_column, 15) + 1)]
                    for row in range(1, min(ws.max_row, 30) + 1)
                ],
            }
        )
    return result


def _write_headers(
    ws,
    columns: list[dict[str, Any]],
    groups: list[dict[str, Any]],
    column_map: dict[str, int],
    top: int,
    bottom: int,
) -> None:
    grouped_keys: set[str] = set()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    second_fill = PatternFill("solid", fgColor="EAF3F9")
    for group in groups:
        start = column_map[group["start_key"]]
        end = column_map[group["end_key"]]
        grouped_keys.update(
            item["key"] for item in columns[start - 1 : end]
        )
        ws.merge_cells(start_row=top, start_column=start, end_row=top, end_column=end)
        cell = ws.cell(top, start, group["label"])
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", bold=True, color="17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start, end + 1):
            child = ws.cell(bottom, col, columns[col - 1]["label"])
            child.fill = second_fill
            child.font = Font(name="Microsoft YaHei", bold=True, color="17324D")
            child.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, column in enumerate(columns, start=1):
        if column["key"] in grouped_keys:
            continue
        if bottom > top:
            ws.merge_cells(start_row=top, start_column=col, end_row=bottom, end_column=col)
        cell = ws.cell(top, col, column["label"])
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", bold=True, color="17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(top, bottom + 1):
        ws.row_dimensions[row].height = 25
        for col in range(1, len(columns) + 1):
            ws.cell(row, col).border = Border(
                left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY
            )


def _write_record_row(
    ws,
    row: int,
    columns: list[dict[str, Any]],
    record: dict[str, Any],
    column_map: dict[str, int],
    plan: dict[str, Any],
) -> None:
    for col, column in enumerate(columns, start=1):
        if column.get("formula"):
            value = _render_formula(column["formula"], row, column_map)
        else:
            value = record.get(column["key"], "")
        ws.cell(row, col, value)


def _write_subtotal_row(
    ws,
    row: int,
    group_value: Any,
    data_rows: list[int],
    columns: list[dict[str, Any]],
    column_map: dict[str, int],
    config: dict[str, Any],
) -> None:
    label_keys = config.get("merge_label_keys") or [config.get("group_key")]
    label_indices = [column_map[key] for key in label_keys if key in column_map]
    if label_indices:
        start, end = min(label_indices), max(label_indices)
        if end > start:
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        template = str(config.get("label_template", "{group}小计"))
        label = template.replace("{group}", str(group_value)).replace(
            "{group_key}", str(group_value)
        ).replace(
            "{group_value}", str(group_value)
        )
        ws.cell(row, start, label)
    _write_aggregate_cells(ws, row, data_rows, columns, column_map, config)
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="17324D")
        cell.border = Border(top=Side(style="thin", color="8095AA"), bottom=THIN_GRAY)


def _write_grand_total_row(
    ws,
    row: int,
    data_rows: list[int],
    columns: list[dict[str, Any]],
    column_map: dict[str, int],
    config: dict[str, Any],
) -> None:
    label_keys = config.get("merge_label_keys") or _default_total_label_keys(
        columns,
        config,
    )
    label_indices = [column_map[key] for key in label_keys if key in column_map]
    if label_indices:
        start, end = min(label_indices), max(label_indices)
        if end > start:
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        ws.cell(row, start, str(config.get("label", "总计")))
    _write_aggregate_cells(ws, row, data_rows, columns, column_map, config)
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor="B4C6E7")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="10243A")
        cell.border = Border(top=Side(style="medium", color="50677F"), bottom=THIN_GRAY)


def _default_total_label_keys(
    columns: list[dict[str, Any]],
    config: dict[str, Any],
) -> list[str]:
    """Choose label cells that cannot overlap aggregate/formula targets."""

    protected = {
        *config.get("sum_keys", []),
        *config.get("average_keys", []),
        *config.get("average_from", {}).keys(),
        *config.get("value_map", {}).keys(),
    }
    candidates: list[str] = []
    for column in columns:
        key = column["key"]
        if key in protected or column.get("formula"):
            break
        candidates.append(key)
    if not candidates:
        candidates = [
            column["key"]
            for column in columns
            if column["key"] not in protected
        ][:1]
    return candidates


def _write_aggregate_cells(
    ws,
    row: int,
    data_rows: list[int],
    columns: list[dict[str, Any]],
    column_map: dict[str, int],
    config: dict[str, Any],
) -> None:
    if not data_rows:
        return
    ranges = _contiguous_ranges(data_rows)
    for key in config.get("sum_keys", []):
        if key not in column_map:
            continue
        letter = get_column_letter(column_map[key])
        args = ",".join(f"{letter}{start}:{letter}{end}" for start, end in ranges)
        ws.cell(row, column_map[key], f"=SUM({args})")
    for key in config.get("average_keys", []):
        if key not in column_map:
            continue
        letter = get_column_letter(column_map[key])
        args = ",".join(f"{letter}{start}:{letter}{end}" for start, end in ranges)
        ws.cell(row, column_map[key], f"=ROUND(AVERAGE({args}),0)")
    for target_key, source_key in config.get("average_from", {}).items():
        if target_key not in column_map or source_key not in column_map:
            continue
        letter = get_column_letter(column_map[source_key])
        args = ",".join(f"{letter}{start}:{letter}{end}" for start, end in ranges)
        ws.cell(row, column_map[target_key], f"=ROUND(AVERAGE({args}),0)")
    for key, value in config.get("value_map", {}).items():
        if key in column_map:
            ws.cell(row, column_map[key], value)
    for key in config.get("blank_keys", []):
        if key in column_map:
            ws.cell(row, column_map[key], "")


def _apply_body_styles(
    ws,
    columns: list[dict[str, Any]],
    start_row: int,
    end_row: int,
    plan: dict[str, Any],
) -> None:
    for col, column in enumerate(columns, start=1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = column["width"]
        number_format = column.get("number_format") or _number_format(column["type"])
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(
                horizontal="left" if column["type"] == "text" else "right",
                vertical="center",
            )
            cell.border = Border(
                left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY
            )
            if number_format:
                cell.number_format = number_format
        ws.column_dimensions[letter].bestFit = False
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = 22


def _apply_conditional_formats(
    ws,
    rules: list[dict[str, Any]],
    column_map: dict[str, int],
    body_rows: list[int],
) -> None:
    if not body_rows:
        return
    start, end = min(body_rows), max(body_rows)
    for raw in rules:
        if not isinstance(raw, dict):
            continue
        key = str(raw.get("column_key", ""))
        if key not in column_map:
            continue
        letter = get_column_letter(column_map[key])
        cell_range = f"{letter}{start}:{letter}{end}"
        font_color = str(raw.get("font_color") or "000000").lstrip("#")
        font = Font(
            bold=bool(raw.get("font_bold")),
            color=font_color,
        )
        fill_color = str(raw.get("fill_color") or "").lstrip("#")
        fill = PatternFill("solid", fgColor=fill_color) if fill_color else None
        kind = str(raw.get("kind") or "cell_is")
        if kind == "formula":
            template = str(raw.get("formula") or "")
            formula = template.replace("{cell}", f"{letter}{start}")
            ws.conditional_formatting.add(
                cell_range,
                FormulaRule(formula=[formula.lstrip("=")], font=font, fill=fill),
            )
        else:
            operator = str(raw.get("operator") or "greaterThanOrEqual")
            value = raw.get("value", 0)
            formula_value = (
                f'"{str(value).replace(chr(34), chr(34) * 2)}"'
                if isinstance(value, str)
                else str(value)
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator=operator,
                    formula=[formula_value],
                    font=font,
                    fill=fill,
                ),
            )


def _add_chart(
    ws,
    spec: dict[str, Any],
    column_map: dict[str, int],
    data_start: int,
    body_rows: list[int],
    body_records: list[dict[str, Any]],
    columns: list[dict[str, Any]],
) -> None:
    if not body_rows:
        return
    chart_type = spec.get("type", "column")
    chart = (
        LineChart()
        if chart_type == "line"
        else PieChart()
        if chart_type == "pie"
        else BarChart()
    )
    if chart_type == "column" and isinstance(chart, BarChart):
        chart.type = "col"
    chart.title = spec.get("title") or "数据图表"
    chart.height = float(spec.get("height") or 8)
    chart.width = float(spec.get("width") or 15)
    category_key = spec.get("category_key")
    value_keys = [
        key for key in spec.get("value_keys", []) if key in column_map
    ]
    if category_key not in column_map or not value_keys:
        return
    helper = _chart_helper_sheet(ws.parent)
    helper.cell(1, 1, ws.cell(data_start - 1, column_map[category_key]).value)
    for index, key in enumerate(value_keys, start=2):
        helper.cell(1, index, ws.cell(data_start - 1, column_map[key]).value)
    for helper_row, (source_row, record) in enumerate(
        zip(body_rows, body_records),
        start=2,
    ):
        category_value = record.get(category_key)
        if category_value in (None, ""):
            category_value = ws.cell(source_row, column_map[category_key]).value
        helper.cell(helper_row, 1, category_value)
        for index, key in enumerate(value_keys, start=2):
            value = _chart_record_value(record, key, columns)
            if value is None:
                raw = ws.cell(source_row, column_map[key]).value
                value = raw if not (isinstance(raw, str) and raw.startswith("=")) else 0
            helper.cell(helper_row, index, value)
    cats = Reference(helper, min_col=1, min_row=2, max_row=1 + len(body_rows))
    for index, _key in enumerate(value_keys, start=2):
        data = Reference(
            helper,
            min_col=index,
            min_row=1,
            max_row=1 + len(body_rows),
        )
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if len(value_keys) == 1:
        chart.legend = None
    else:
        chart.legend.position = "r"
    if hasattr(chart, "y_axis"):
        chart.y_axis.numFmt = "#,##0"
    position = str(spec.get("position") or "")
    if not re.fullmatch(r"[A-Z]{1,3}[1-9]\d*", position):
        position = f"{get_column_letter(len(column_map) + 2)}3"
    ws.add_chart(chart, position)


def _chart_helper_sheet(wb):
    base = "_图表数据"
    name = base
    counter = 2
    while name in wb.sheetnames:
        name = f"{base}{counter}"
        counter += 1
    ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    return ws


def _chart_record_value(
    record: dict[str, Any],
    key: str,
    columns: list[dict[str, Any]],
    stack: set[str] | None = None,
) -> float | int | str | None:
    value = record.get(key)
    if value not in (None, ""):
        number = _to_number(value)
        return number if number != 0 or str(value).strip() in {"0", "0.0"} else value
    column = next((item for item in columns if item["key"] == key), None)
    if not column or not column.get("formula"):
        return None
    active = set(stack or ())
    if key in active:
        return None
    active.add(key)
    formula = str(column["formula"]).strip().lstrip("=")
    values: dict[str, float] = {}
    for source_key in re.findall(r"\{([A-Za-z_][A-Za-z0-9_]*)\}", formula):
        resolved = _chart_record_value(record, source_key, columns, active)
        if not isinstance(resolved, (int, float)):
            return None
        values[source_key] = float(resolved)
    expression = formula
    for source_key, resolved in values.items():
        expression = expression.replace(f"{{{source_key}}}", str(resolved))
    return _evaluate_chart_expression(expression)


def _evaluate_chart_expression(expression: str) -> float | int | None:
    text = expression.strip()
    upper = text.upper()
    if upper.startswith("IFERROR(") and text.endswith(")"):
        arguments = _split_formula_arguments(text[8:-1])
        return _evaluate_chart_expression(arguments[0]) if arguments else None
    if upper.startswith("ROUND(") and text.endswith(")"):
        arguments = _split_formula_arguments(text[6:-1])
        if not arguments:
            return None
        value = _evaluate_chart_expression(arguments[0])
        digits = int(float(arguments[1])) if len(arguments) > 1 else 0
        return round(float(value), digits) if value is not None else None
    if upper.startswith(("SUM(", "AVERAGE(")) and text.endswith(")"):
        function = upper.split("(", 1)[0]
        values = [
            _evaluate_chart_expression(item)
            for item in _split_formula_arguments(text[text.index("(") + 1 : -1])
        ]
        numeric = [float(item) for item in values if item is not None]
        if not numeric:
            return None
        result = sum(numeric)
        return result / len(numeric) if function == "AVERAGE" else result
    try:
        tree = ast.parse(text, mode="eval")
        return _eval_numeric_ast(tree.body)
    except (SyntaxError, TypeError, ValueError, ZeroDivisionError):
        return None


def _eval_numeric_ast(node: ast.AST) -> float:
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
        return float(node.value)
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
        value = _eval_numeric_ast(node.operand)
        return value if isinstance(node.op, ast.UAdd) else -value
    if isinstance(node, ast.BinOp) and isinstance(
        node.op,
        (ast.Add, ast.Sub, ast.Mult, ast.Div),
    ):
        left = _eval_numeric_ast(node.left)
        right = _eval_numeric_ast(node.right)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        return left / right
    raise ValueError("unsupported chart expression")


def _split_formula_arguments(value: str) -> list[str]:
    result: list[str] = []
    depth = 0
    start = 0
    for index, char in enumerate(value):
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        elif char == "," and depth == 0:
            result.append(value[start:index].strip())
            start = index + 1
    result.append(value[start:].strip())
    return [item for item in result if item]


def _sort_records(
    records: list[dict[str, Any]],
    sort_specs: list[dict[str, Any]],
    columns: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    result = [deepcopy(item) for item in records]
    formula_sources = {
        item["key"]: re.findall(r"\{([A-Za-z_][A-Za-z0-9_]*)\}", item.get("formula", ""))
        for item in columns
        if item.get("formula")
    }
    for spec in reversed(sort_specs):
        if not isinstance(spec, dict):
            continue
        direction = str(spec.get("direction") or "asc").lower()
        reverse = direction == "desc"
        order = spec.get("order")
        if isinstance(order, list) and spec.get("key"):
            order_map = {str(value): index for index, value in enumerate(order)}
            key = str(spec["key"])
            result.sort(
                key=lambda item: order_map.get(str(item.get(key, "")), len(order_map)),
                reverse=reverse,
            )
        elif spec.get("aggregate") == "sum" and isinstance(spec.get("keys"), list):
            keys = []
            for raw_key in spec["keys"]:
                key = str(raw_key)
                keys.extend(formula_sources.get(key) or [key])
            result.sort(
                key=lambda item: sum(_to_number(item.get(key)) for key in keys),
                reverse=reverse,
            )
        elif spec.get("key"):
            key = str(spec["key"])
            result.sort(key=lambda item: _sort_value(item.get(key)), reverse=reverse)
    return result


def _group_records(
    records: list[dict[str, Any]],
    group_key: str,
) -> list[tuple[str, list[dict[str, Any]]]]:
    result: list[tuple[str, list[dict[str, Any]]]] = []
    for record in records:
        value = str(record.get(group_key, ""))
        if not result or result[-1][0] != value:
            result.append((value, [record]))
        else:
            result[-1][1].append(record)
    return result


def _render_formula(template: str, row: int, column_map: dict[str, int]) -> str:
    formula = template
    for key, index in column_map.items():
        formula = formula.replace(f"{{{key}}}", f"{get_column_letter(index)}{row}")
    if "{" in formula or "}" in formula:
        raise ValueError(f"公式模板存在未知列: {template}")
    if not formula.startswith("="):
        formula = f"={formula}"
    return formula


def _sanitize_formula_template(value: Any) -> str:
    formula = str(value or "").strip()
    if not formula:
        return ""
    upper = formula.upper()
    if "[" in formula or any(token in upper for token in PROHIBITED_FORMULA_TOKENS):
        raise ValueError("公式包含禁止的外部调用或引用。")
    return formula


def _normalize_header_groups(value: Any, keys: set[str]) -> list[dict[str, str]]:
    if not isinstance(value, list):
        return []
    result = []
    for item in value:
        if not isinstance(item, dict):
            continue
        start = str(item.get("start_key") or "")
        end = str(item.get("end_key") or "")
        label = str(item.get("label") or "").strip()
        if start in keys and end in keys and label:
            result.append({"label": label, "start_key": start, "end_key": end})
    return result


def _normalize_summary_config(value: Any, keys: set[str]) -> dict[str, Any] | None:
    if not isinstance(value, dict):
        return None
    result = dict(value)
    for field in ("sum_keys", "average_keys", "blank_keys", "merge_label_keys"):
        result[field] = [str(key) for key in value.get(field, []) if str(key) in keys]
    average_from = value.get("average_from")
    result["average_from"] = (
        {
            str(target): str(source)
            for target, source in average_from.items()
            if str(target) in keys and str(source) in keys
        }
        if isinstance(average_from, dict)
        else {}
    )
    value_map = value.get("value_map")
    result["value_map"] = (
        {
            str(key): item
            for key, item in value_map.items()
            if str(key) in keys
            and isinstance(item, (str, int, float, bool))
            and not (isinstance(item, str) and not item.strip())
            and not (isinstance(item, str) and item.startswith("="))
        }
        if isinstance(value_map, dict)
        else {}
    )
    if value.get("group_key") not in keys:
        result["group_key"] = ""
    return result


def _normalize_charts(value: Any, keys: set[str]) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    result = []
    for item in value[:8]:
        if not isinstance(item, dict):
            continue
        chart_type = str(item.get("type") or "column").lower()
        if chart_type not in ALLOWED_CHART_TYPES:
            chart_type = "column"
        category_key = str(item.get("category_key") or "")
        values = [str(key) for key in item.get("value_keys", []) if str(key) in keys]
        if category_key in keys and values:
            normalized = dict(item)
            normalized["type"] = chart_type
            normalized["category_key"] = category_key
            normalized["value_keys"] = values
            result.append(normalized)
    return result


def _default_chart(columns: list[dict[str, Any]]) -> dict[str, Any]:
    text_key = next(
        (item["key"] for item in columns if item["type"] == "text"),
        columns[0]["key"],
    )
    value_keys = [
        item["key"]
        for item in columns
        if item["type"] in {"number", "money", "integer"}
    ][-1:]
    return {
        "type": "column",
        "title": "核心数据对比",
        "category_key": text_key,
        "value_keys": value_keys,
        "position": f"{get_column_letter(len(columns) + 2)}3",
    }


def _contiguous_ranges(rows: list[int]) -> list[tuple[int, int]]:
    result: list[tuple[int, int]] = []
    for row in sorted(rows):
        if not result or row > result[-1][1] + 1:
            result.append((row, row))
        else:
            result[-1] = (result[-1][0], row)
    return result


def _number_format(column_type: str) -> str:
    return {
        "money": '#,##0.00;[Red]-#,##0.00',
        "number": '#,##0.00;[Red]-#,##0.00',
        "integer": '#,##0;[Red]-#,##0',
        "percentage": "0.00%",
        "date": "yyyy-mm-dd",
    }.get(column_type, "")


def _default_width(label: str) -> float:
    return min(max(len(label) * 2 + 3, 10), 24)


def _sort_value(value: Any) -> tuple[int, Any]:
    try:
        return 0, float(value)
    except (TypeError, ValueError):
        return 1, str(value or "")


def _to_number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0
