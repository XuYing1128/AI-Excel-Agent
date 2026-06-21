"""Data cleaning utilities for CSV/XLSX inputs."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .io_utils import ensure_output_path, read_table
from .style_library import (
    apply_header_style,
    apply_print_settings,
    apply_title_style,
    freeze_and_filter,
    set_reasonable_column_widths,
    style_instruction_sheet,
)


def clean_table_file(input_path: str | Path, output: str | Path | None = None) -> Path:
    df = read_table(input_path)
    cleaned, report = clean_dataframe(df)
    output_path = ensure_output_path(output, "cleaned.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = wb.create_sheet("Instructions")
    instructions["A1"] = "数据清洗说明"
    instructions["A3"] = "源文件"
    instructions["B3"] = str(input_path)
    instructions["A4"] = "处理动作"
    instructions["B4"] = "去除空行空列、标准化日期/金额、去重、生成缺失值和异常值报告。"
    style_instruction_sheet(instructions)

    data_ws = wb.create_sheet("CleanedData")
    _write_dataframe(data_ws, cleaned, "清洗后数据")

    report_ws = wb.create_sheet("clean_report")
    _write_report(report_ws, report)

    wb.save(output_path)
    return output_path


def clean_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    original_shape = df.shape
    work = df.copy()
    work = work.dropna(how="all").dropna(axis=1, how="all")
    work.columns = [str(col).strip() if str(col).strip() else f"Column_{idx+1}" for idx, col in enumerate(work.columns)]

    duplicate_count = int(work.duplicated().sum())
    work = work.drop_duplicates().reset_index(drop=True)

    conversions: dict[str, str] = {}
    for col in work.columns:
        name = str(col)
        lower = name.lower()
        if any(key in lower for key in ["date", "日期", "时间"]):
            converted = pd.to_datetime(work[col], errors="coerce")
            if converted.notna().sum() > 0:
                work[col] = converted
                conversions[name] = "date"
        if any(key in lower for key in ["amount", "金额", "价格", "单价", "收入", "成本", "销售额"]):
            work[col] = (
                work[col]
                .astype(str)
                .str.replace(",", "", regex=False)
                .str.replace("¥", "", regex=False)
                .str.replace("￥", "", regex=False)
                .str.strip()
            )
            work[col] = pd.to_numeric(work[col], errors="coerce")
            conversions[name] = "money"

    missing = {str(col): int(work[col].isna().sum()) for col in work.columns}
    outliers: dict[str, int] = {}
    for col in work.select_dtypes(include="number").columns:
        series = work[col].dropna()
        if series.empty:
            continue
        q1 = series.quantile(0.25)
        q3 = series.quantile(0.75)
        iqr = q3 - q1
        if iqr == 0:
            outliers[str(col)] = 0
            continue
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        outliers[str(col)] = int(((series < lower) | (series > upper)).sum())

    report = {
        "original_rows": int(original_shape[0]),
        "original_columns": int(original_shape[1]),
        "cleaned_rows": int(work.shape[0]),
        "cleaned_columns": int(work.shape[1]),
        "duplicate_rows_removed": duplicate_count,
        "conversions": conversions,
        "missing_values": missing,
        "outliers": outliers,
    }
    return work, report


def _write_dataframe(ws, df: pd.DataFrame, title: str) -> None:
    max_col = max(1, len(df.columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = title
    apply_title_style(ws, 1, 1, max_col)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)
    apply_header_style(ws, 3)
    if ws.max_row >= 3 and ws.max_column >= 1:
        freeze_and_filter(ws, "A4", f"A3:{get_column_letter(ws.max_column)}{ws.max_row}")
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _write_report(ws, report: dict[str, Any]) -> None:
    ws["A1"] = "清洗报告"
    apply_title_style(ws, 1, 1, 3)
    ws["A3"] = "项目"
    ws["B3"] = "字段"
    ws["C3"] = "结果"
    apply_header_style(ws, 3)
    row = 4
    for key, value in report.items():
        if isinstance(value, dict):
            for sub_key, sub_value in value.items():
                ws.cell(row, 1, key)
                ws.cell(row, 2, sub_key)
                ws.cell(row, 3, sub_value)
                row += 1
        else:
            ws.cell(row, 1, key)
            ws.cell(row, 3, value)
            row += 1
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)
