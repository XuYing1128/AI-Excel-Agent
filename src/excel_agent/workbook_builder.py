"""Workbook template and report builders."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from . import formula_library as F
from .chart_library import add_bar_chart, add_line_chart, add_pie_chart
from .intent_classifier import normalize_table_type
from .io_utils import ensure_output_path, read_table
from .style_library import (
    apply_calculation_style,
    apply_header_style,
    apply_input_style,
    apply_negative_red,
    apply_number_formats,
    apply_print_settings,
    apply_summary_style,
    apply_title_style,
    freeze_and_filter,
    set_reasonable_column_widths,
    style_instruction_sheet,
)


FormulaFn = Callable[[int], str]
SummaryFn = Callable[[int, int], str]


@dataclass
class TemplateSpec:
    table_type: str
    title: str
    headers: list[str]
    sample_rows: list[dict[str, object]]
    formula_cols: dict[int, FormulaFn]
    summary_rows: list[tuple[str, SummaryFn]]
    input_cols: list[int]
    calc_cols: list[int]
    money_cols: list[int] = field(default_factory=list)
    percent_cols: list[int] = field(default_factory=list)
    date_cols: list[int] = field(default_factory=list)
    integer_cols: list[int] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    chart: str | None = None


FORMULA_ROWS = 30


def _data_range(col: str, start_row: int, end_row: int) -> str:
    return f"Data!${col}${start_row}:${col}${end_row}"


def _guard(row: int, required_col: str, formula: str) -> str:
    return f'=IF({required_col}{row}="","",{F.strip_equals(formula)})'


def _coerce_sample_value(header: str, col: int, value: object, spec: TemplateSpec) -> object:
    if value in (None, ""):
        return value
    if col in spec.date_cols and isinstance(value, str):
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.notna(parsed):
            return parsed.to_pydatetime()
    if "时间" in header and isinstance(value, str) and ":" in value:
        try:
            return datetime.strptime(value, "%H:%M").time()
        except ValueError:
            return value
    return value


def _make_instruction_sheet(wb: Workbook, title: str, table_type: str, notes: list[str]) -> None:
    ws = wb.create_sheet("Instructions", 0)
    ws["A1"] = title
    ws["A3"] = "表格类型"
    ws["B3"] = table_type
    ws["A4"] = "使用方法"
    ws["B4"] = "黄色区域为输入区，绿色区域为公式计算区。请优先替换输入数据，不要直接覆盖公式列。"
    ws["A5"] = "校验要求"
    ws["B5"] = "生成或修改后运行 python -m excel_agent.cli validate --input <文件路径>。"
    ws["A6"] = "风险提示"
    ws["B6"] = "财务、报价、合同、工资、税务、绩效等高风险用途必须人工复核。"
    if notes:
        ws["A8"] = "模板说明"
        ws["B8"] = "\n".join(notes)
    style_instruction_sheet(ws)


def _build_standard_workbook(spec: TemplateSpec, output: str | Path | None) -> Path:
    output_path = ensure_output_path(output, f"{spec.table_type}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    _make_instruction_sheet(wb, spec.title, spec.table_type, spec.notes)

    ws = wb.create_sheet("Data")
    summary = wb.create_sheet("Summary")

    header_row = 3
    start_row = 4
    end_row = start_row + FORMULA_ROWS - 1
    max_col = len(spec.headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.cell(1, 1, spec.title)
    apply_title_style(ws, 1, 1, max_col)

    for col, header in enumerate(spec.headers, start=1):
        ws.cell(header_row, col, header)
    apply_header_style(ws, header_row)

    for idx in range(FORMULA_ROWS):
        row_num = start_row + idx
        sample = spec.sample_rows[idx] if idx < len(spec.sample_rows) else {}
        for col, header in enumerate(spec.headers, start=1):
            if col in spec.formula_cols:
                ws.cell(row_num, col, spec.formula_cols[col](row_num))
            else:
                ws.cell(row_num, col, _coerce_sample_value(header, col, sample.get(header), spec))

    apply_input_style(ws, start_row, end_row, spec.input_cols)
    apply_calculation_style(ws, start_row, end_row, spec.calc_cols)
    apply_number_formats(
        ws,
        money_cols=spec.money_cols,
        percent_cols=spec.percent_cols,
        date_cols=spec.date_cols,
        integer_cols=spec.integer_cols,
        min_row=start_row,
        max_row=end_row,
    )
    for col in spec.money_cols + spec.percent_cols:
        apply_negative_red(ws, f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row}")
    freeze_and_filter(ws, "A4", f"A{header_row}:{get_column_letter(max_col)}{end_row}")
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)

    _write_summary_sheet(summary, spec, start_row, end_row)
    if spec.chart:
        _add_template_chart(summary, spec.chart, len(spec.summary_rows))

    wb.save(output_path)
    return output_path


def _write_summary_sheet(ws, spec: TemplateSpec, data_start: int, data_end: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws["A1"] = f"{spec.title} - 汇总"
    apply_title_style(ws, 1, 1, 4)
    ws["A3"] = "指标"
    ws["B3"] = "结果"
    apply_header_style(ws, 3)
    for offset, (label, formula_fn) in enumerate(spec.summary_rows, start=4):
        ws.cell(offset, 1, label)
        ws.cell(offset, 2, formula_fn(data_start, data_end))
        _format_summary_cell(ws.cell(offset, 2), label)
    validation_row = 4 + len(spec.summary_rows)
    ws.cell(validation_row, 1, "校验点：数据行数")
    ws.cell(validation_row, 2, f'=IF(COUNTA(Data!A{data_start}:A{data_end})>0,"PASS","CHECK")')
    apply_summary_style(ws, 4, validation_row, 1, 2)
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _format_summary_cell(cell, label: str) -> None:
    if any(word in label for word in ["率", "占比", "ROI"]):
        cell.number_format = "0.00%"
    elif any(word in label for word in ["数", "次数", "条数", "订单"]):
        cell.number_format = "0"
    elif any(word in label for word in ["库存", "出库", "周转"]):
        cell.number_format = "0.00"
    else:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'


def _add_template_chart(ws, chart_kind: str, summary_count: int) -> None:
    data_max = 3 + summary_count
    if data_max < 4:
        return
    if chart_kind == "pie":
        add_pie_chart(ws, "汇总占比", data_col=2, data_min_row=4, data_max_row=data_max, cats_col=1, cats_min_row=4, cats_max_row=data_max, anchor="D3")
    elif chart_kind == "line":
        add_line_chart(ws, "汇总趋势", data_min_col=2, data_max_col=2, data_min_row=3, data_max_row=data_max, cats_col=1, cats_min_row=4, cats_max_row=data_max, anchor="D3")
    else:
        add_bar_chart(ws, "核心指标", data_min_col=2, data_max_col=2, data_min_row=3, data_max_row=data_max, cats_col=1, cats_min_row=4, cats_max_row=data_max, anchor="D3")


def _specs() -> dict[str, TemplateSpec]:
    return {
        "personal_budget": TemplateSpec(
            table_type="personal_budget",
            title="个人月度收支预算表",
            headers=["日期", "分类", "项目", "预算金额", "实际金额", "差异", "执行率", "备注"],
            sample_rows=[
                {"日期": "2026-05-01", "分类": "收入", "项目": "工资", "预算金额": 12000, "实际金额": 12000, "备注": ""},
                {"日期": "2026-05-03", "分类": "餐饮", "项目": "日常餐饮", "预算金额": 2000, "实际金额": 1850, "备注": ""},
                {"日期": "2026-05-05", "分类": "交通", "项目": "通勤", "预算金额": 600, "实际金额": 520, "备注": ""},
                {"日期": "2026-05-08", "分类": "娱乐", "项目": "电影/聚餐", "预算金额": 800, "实际金额": 950, "备注": "超预算"},
            ],
            formula_cols={
                6: lambda r: _guard(r, "A", f"=E{r}-D{r}"),
                7: lambda r: _guard(r, "A", F.share(f"E{r}", f"D{r}")),
            },
            summary_rows=[
                ("预算合计", lambda s, e: f"=SUM(Data!D{s}:D{e})"),
                ("实际合计", lambda s, e: f"=SUM(Data!E{s}:E{e})"),
                ("预算差异", lambda s, e: f"=SUM(Data!F{s}:F{e})"),
                ("整体执行率", lambda s, e: f"=IFERROR(B5/B4,0)"),
            ],
            input_cols=[1, 2, 3, 4, 5, 8],
            calc_cols=[6, 7],
            money_cols=[4, 5, 6],
            percent_cols=[7],
            date_cols=[1],
            notes=["差异=实际金额-预算金额；执行率=实际金额/预算金额。"],
            chart="bar",
        ),
        "family_budget": TemplateSpec(
            table_type="family_budget",
            title="家庭年度预算表",
            headers=["月份", "成员", "分类", "项目", "年度预算", "实际支出", "差异", "执行率", "备注"],
            sample_rows=[
                {"月份": "2026-01-01", "成员": "家庭", "分类": "住房", "项目": "房租/房贷", "年度预算": 96000, "实际支出": 8000, "备注": ""},
                {"月份": "2026-01-01", "成员": "家庭", "分类": "餐饮", "项目": "日常餐饮", "年度预算": 60000, "实际支出": 5200, "备注": ""},
                {"月份": "2026-02-01", "成员": "孩子", "分类": "教育", "项目": "培训课程", "年度预算": 30000, "实际支出": 2600, "备注": ""},
                {"月份": "2026-02-01", "成员": "家庭", "分类": "交通", "项目": "油费/公共交通", "年度预算": 18000, "实际支出": 1500, "备注": ""},
            ],
            formula_cols={
                7: lambda r: _guard(r, "A", f"=F{r}-E{r}/12"),
                8: lambda r: _guard(r, "A", f"=IFERROR(F{r}/(E{r}/12),0)"),
            },
            summary_rows=[
                ("年度预算合计", lambda s, e: f"=SUM(Data!E{s}:E{e})"),
                ("实际支出合计", lambda s, e: f"=SUM(Data!F{s}:F{e})"),
                ("月度预算差异", lambda s, e: f"=SUM(Data!G{s}:G{e})"),
                ("平均执行率", lambda s, e: f"=AVERAGE(Data!H{s}:H{e})"),
            ],
            input_cols=[1, 2, 3, 4, 5, 6, 9],
            calc_cols=[7, 8],
            money_cols=[5, 6, 7],
            percent_cols=[8],
            date_cols=[1],
            notes=["差异=实际支出-年度预算/12；执行率=实际支出/(年度预算/12)。"],
            chart="bar",
        ),
        "quotation": TemplateSpec(
            table_type="quotation",
            title="产品报价单",
            headers=["产品", "规格", "数量", "单价", "折扣率", "税率", "未税金额", "税额", "含税总价", "备注"],
            sample_rows=[
                {"产品": "产品A", "规格": "标准版", "数量": 10, "单价": 120, "折扣率": 0.05, "税率": 0.13, "备注": ""},
                {"产品": "产品B", "规格": "高级版", "数量": 5, "单价": 360, "折扣率": 0.1, "税率": 0.13, "备注": ""},
            ],
            formula_cols={
                7: lambda r: _guard(r, "A", f"=C{r}*D{r}*(1-E{r})"),
                8: lambda r: _guard(r, "A", f"=G{r}*F{r}"),
                9: lambda r: _guard(r, "A", f"=G{r}+H{r}"),
            },
            summary_rows=[
                ("未税小计", lambda s, e: f"=SUM(Data!G{s}:G{e})"),
                ("税额合计", lambda s, e: f"=SUM(Data!H{s}:H{e})"),
                ("报价总额", lambda s, e: f"=SUM(Data!I{s}:I{e})"),
            ],
            input_cols=[1, 2, 3, 4, 5, 6, 10],
            calc_cols=[7, 8, 9],
            money_cols=[4, 7, 8, 9],
            percent_cols=[5, 6],
            integer_cols=[3],
            notes=["合同、采购和发票用途需人工复核税率、币种和条款。"],
            chart="bar",
        ),
        "invoice_draft": TemplateSpec(
            table_type="invoice_draft",
            title="发票草稿/收款明细",
            headers=["收款日期", "客户", "发票抬头", "项目", "数量", "单价", "税率", "未税金额", "税额", "含税金额", "收款状态", "备注"],
            sample_rows=[
                {"收款日期": "2026-05-01", "客户": "客户A", "发票抬头": "客户A有限公司", "项目": "服务费", "数量": 1, "单价": 12000, "税率": 0.06, "收款状态": "已收款", "备注": ""},
                {"收款日期": "2026-05-08", "客户": "客户B", "发票抬头": "客户B有限公司", "项目": "软件订阅", "数量": 3, "单价": 3000, "税率": 0.13, "收款状态": "待收款", "备注": ""},
            ],
            formula_cols={
                8: lambda r: _guard(r, "A", f"=E{r}*F{r}"),
                9: lambda r: _guard(r, "A", f"=H{r}*G{r}"),
                10: lambda r: _guard(r, "A", f"=H{r}+I{r}"),
            },
            summary_rows=[
                ("未税金额合计", lambda s, e: f"=SUM(Data!H{s}:H{e})"),
                ("税额合计", lambda s, e: f"=SUM(Data!I{s}:I{e})"),
                ("含税金额合计", lambda s, e: f"=SUM(Data!J{s}:J{e})"),
                ("待收款笔数", lambda s, e: f'=COUNTIF(Data!K{s}:K{e},"待收款")'),
            ],
            input_cols=[1, 2, 3, 4, 5, 6, 7, 11, 12],
            calc_cols=[8, 9, 10],
            money_cols=[6, 8, 9, 10],
            percent_cols=[7],
            date_cols=[1],
            integer_cols=[5],
            notes=["这是发票草稿和收款跟踪，不是正式税务申报文件；开票和税务口径需要人工复核。"],
            chart="bar",
        ),
        "inventory": TemplateSpec(
            table_type="inventory",
            title="小店库存进销存表",
            headers=["SKU", "品名", "期初库存", "入库", "出库", "期末库存", "安全库存", "库存预警", "库存周转率"],
            sample_rows=[
                {"SKU": "SKU-001", "品名": "咖啡豆", "期初库存": 80, "入库": 40, "出库": 95, "安全库存": 30},
                {"SKU": "SKU-002", "品名": "纸杯", "期初库存": 500, "入库": 200, "出库": 360, "安全库存": 150},
                {"SKU": "SKU-003", "品名": "吸管", "期初库存": 400, "入库": 100, "出库": 420, "安全库存": 120},
            ],
            formula_cols={
                6: lambda r: _guard(r, "A", f"=C{r}+D{r}-E{r}"),
                8: lambda r: _guard(r, "A", f'=IF(F{r}<G{r},"补货","正常")'),
                9: lambda r: _guard(r, "A", f"=IFERROR(E{r}/AVERAGE(C{r},F{r}),0)"),
            },
            summary_rows=[
                ("期末库存合计", lambda s, e: f"=SUM(Data!F{s}:F{e})"),
                ("出库合计", lambda s, e: f"=SUM(Data!E{s}:E{e})"),
                ("需补货 SKU 数", lambda s, e: f'=COUNTIF(Data!H{s}:H{e},"补货")'),
                ("平均周转率", lambda s, e: f"=AVERAGE(Data!I{s}:I{e})"),
            ],
            input_cols=[1, 2, 3, 4, 5, 7],
            calc_cols=[6, 8, 9],
            integer_cols=[3, 4, 5, 6, 7],
            percent_cols=[9],
            notes=["库存周转率=出库/平均库存；补货状态按期末库存和安全库存判断。"],
            chart="bar",
        ),
        "sales_report": TemplateSpec(
            table_type="sales_report",
            title="销售月报模板",
            headers=["日期", "区域", "产品", "品类", "数量", "单价", "单位成本", "销售额", "毛利", "毛利率"],
            sample_rows=[
                {"日期": "2026-05-01", "区域": "华东", "产品": "A100", "品类": "饮品", "数量": 30, "单价": 18, "单位成本": 9},
                {"日期": "2026-05-02", "区域": "华南", "产品": "B200", "品类": "零食", "数量": 45, "单价": 12, "单位成本": 6},
                {"日期": "2026-05-03", "区域": "华北", "产品": "C300", "品类": "日用", "数量": 20, "单价": 35, "单位成本": 18},
            ],
            formula_cols={
                8: lambda r: _guard(r, "A", f"=E{r}*F{r}"),
                9: lambda r: _guard(r, "A", f"=E{r}*(F{r}-G{r})"),
                10: lambda r: _guard(r, "A", F.gross_margin_from_profit(f"I{r}", f"H{r}")),
            },
            summary_rows=[
                ("销售额合计", lambda s, e: f"=SUM(Data!H{s}:H{e})"),
                ("毛利合计", lambda s, e: f"=SUM(Data!I{s}:I{e})"),
                ("整体毛利率", lambda s, e: f"=IFERROR(B5/B4,0)"),
                ("订单行数", lambda s, e: f"=COUNTA(Data!A{s}:A{e})"),
            ],
            input_cols=[1, 2, 3, 4, 5, 6, 7],
            calc_cols=[8, 9, 10],
            money_cols=[6, 7, 8, 9],
            percent_cols=[10],
            date_cols=[1],
            integer_cols=[5],
            notes=["如果有实际销售 CSV，请使用 analyze 命令生成按品类和月度趋势汇总的报告。"],
            chart="bar",
        ),
        "ecommerce_analysis": TemplateSpec(
            table_type="ecommerce_analysis",
            title="电商订单 GMV 分析表",
            headers=["日期", "订单号", "SKU", "渠道", "数量", "单价", "单位成本", "退款金额", "GMV", "净销售额", "毛利", "退款率"],
            sample_rows=[
                {"日期": "2026-05-01", "订单号": "E20260501001", "SKU": "SKU-A", "渠道": "天猫", "数量": 2, "单价": 199, "单位成本": 110, "退款金额": 0},
                {"日期": "2026-05-02", "订单号": "E20260502003", "SKU": "SKU-B", "渠道": "抖音", "数量": 1, "单价": 299, "单位成本": 180, "退款金额": 50},
            ],
            formula_cols={
                9: lambda r: _guard(r, "A", F.gmv(f"E{r}", f"F{r}")),
                10: lambda r: _guard(r, "A", f"=I{r}-H{r}"),
                11: lambda r: _guard(r, "A", f"=J{r}-E{r}*G{r}"),
                12: lambda r: _guard(r, "A", f"=IFERROR(H{r}/I{r},0)"),
            },
            summary_rows=[
                ("GMV", lambda s, e: f"=SUM(Data!I{s}:I{e})"),
                ("净销售额", lambda s, e: f"=SUM(Data!J{s}:J{e})"),
                ("毛利", lambda s, e: f"=SUM(Data!K{s}:K{e})"),
                ("退款率", lambda s, e: f"=IFERROR(SUM(Data!H{s}:H{e})/SUM(Data!I{s}:I{e}),0)"),
            ],
            input_cols=[1, 2, 3, 4, 5, 6, 7, 8],
            calc_cols=[9, 10, 11, 12],
            money_cols=[6, 7, 8, 9, 10, 11],
            percent_cols=[12],
            date_cols=[1],
            integer_cols=[5],
            notes=["GMV、净销售额、毛利和退款率均由公式生成。"],
            chart="bar",
        ),
        "project_plan": TemplateSpec(
            table_type="project_plan",
            title="项目计划表",
            headers=["任务", "负责人", "开始日期", "结束日期", "工期天数", "完成率", "状态", "延期标记", "备注"],
            sample_rows=[
                {"任务": "需求确认", "负责人": "张三", "开始日期": "2026-06-01", "结束日期": "2026-06-03", "完成率": 1, "备注": ""},
                {"任务": "原型设计", "负责人": "李四", "开始日期": "2026-06-04", "结束日期": "2026-06-10", "完成率": 0.4, "备注": ""},
                {"任务": "开发实现", "负责人": "王五", "开始日期": "2026-06-11", "结束日期": "2026-06-25", "完成率": 0, "备注": ""},
            ],
            formula_cols={
                5: lambda r: _guard(r, "A", f"=D{r}-C{r}+1"),
                7: lambda r: _guard(r, "A", f'=IF(F{r}>=1,"完成",IF(TODAY()>D{r},"延期","进行中"))'),
                8: lambda r: _guard(r, "A", f'=IF(AND(F{r}<1,TODAY()>D{r}),"延期","正常")'),
            },
            summary_rows=[
                ("任务数", lambda s, e: f"=COUNTA(Data!A{s}:A{e})"),
                ("已完成任务", lambda s, e: f'=COUNTIF(Data!G{s}:G{e},"完成")'),
                ("延期任务", lambda s, e: f'=COUNTIF(Data!H{s}:H{e},"延期")'),
                ("平均完成率", lambda s, e: f"=AVERAGE(Data!F{s}:F{e})"),
            ],
            input_cols=[1, 2, 3, 4, 6, 9],
            calc_cols=[5, 7, 8],
            date_cols=[3, 4],
            integer_cols=[5],
            percent_cols=[6],
            notes=["状态和延期标记根据完成率、结束日期和 TODAY() 自动判断。"],
            chart="bar",
        ),
        "schedule": TemplateSpec(
            table_type="schedule",
            title="课程/排班表",
            headers=["日期", "星期", "时段", "课程/班次", "负责人", "地点", "开始时间", "结束时间", "时长(小时)", "备注"],
            sample_rows=[
                {"日期": "2026-06-01", "星期": "周一", "时段": "上午", "课程/班次": "数学", "负责人": "赵老师", "地点": "A101", "开始时间": "09:00", "结束时间": "10:30"},
                {"日期": "2026-06-01", "星期": "周一", "时段": "下午", "课程/班次": "值班", "负责人": "陈晨", "地点": "前台", "开始时间": "14:00", "结束时间": "18:00"},
            ],
            formula_cols={
                9: lambda r: _guard(r, "A", f"=IFERROR((H{r}-G{r})*24,0)"),
            },
            summary_rows=[
                ("安排条数", lambda s, e: f"=COUNTA(Data!A{s}:A{e})"),
                ("总时长", lambda s, e: f"=SUM(Data!I{s}:I{e})"),
                ("负责人人数", lambda s, e: f"=COUNTA(UNIQUE(Data!E{s}:E{e}))"),
            ],
            input_cols=[1, 2, 3, 4, 5, 6, 7, 8, 10],
            calc_cols=[9],
            date_cols=[1],
            notes=["开始时间和结束时间应使用 Excel 可识别的时间格式。"],
            chart="bar",
        ),
        "attendance": TemplateSpec(
            table_type="attendance",
            title="员工考勤统计表",
            headers=["员工", "日期", "应到小时", "实到小时", "缺勤小时", "迟到分钟", "状态", "出勤率", "备注"],
            sample_rows=[
                {"员工": "张三", "日期": "2026-05-01", "应到小时": 8, "实到小时": 8, "迟到分钟": 0, "备注": ""},
                {"员工": "李四", "日期": "2026-05-01", "应到小时": 8, "实到小时": 7.5, "迟到分钟": 10, "备注": ""},
                {"员工": "王五", "日期": "2026-05-01", "应到小时": 8, "实到小时": 0, "迟到分钟": 0, "备注": "请假"},
            ],
            formula_cols={
                5: lambda r: _guard(r, "A", f"=MAX(C{r}-D{r},0)"),
                7: lambda r: _guard(r, "A", f'=IF(E{r}>0,"缺勤",IF(F{r}>0,"迟到","正常"))'),
                8: lambda r: _guard(r, "A", F.attendance_rate(f"D{r}", f"C{r}")),
            },
            summary_rows=[
                ("记录数", lambda s, e: f"=COUNTA(Data!A{s}:A{e})"),
                ("缺勤小时合计", lambda s, e: f"=SUM(Data!E{s}:E{e})"),
                ("迟到次数", lambda s, e: f'=COUNTIF(Data!G{s}:G{e},"迟到")'),
                ("缺勤次数", lambda s, e: f'=COUNTIF(Data!G{s}:G{e},"缺勤")'),
                ("平均出勤率", lambda s, e: f"=AVERAGE(Data!H{s}:H{e})"),
            ],
            input_cols=[1, 2, 3, 4, 6, 9],
            calc_cols=[5, 7, 8],
            date_cols=[2],
            percent_cols=[8],
            notes=["工资、绩效或处罚用途必须结合制度和人工复核。"],
            chart="bar",
        ),
        "finance_model": TemplateSpec(
            table_type="finance_model",
            title="简单收入成本利润测算模型",
            headers=["月份", "收入", "主营成本", "运营费用", "毛利", "净利润", "毛利率", "净利率", "ROI", "收入增长率"],
            sample_rows=[
                {"月份": "2026-01-01", "收入": 100000, "主营成本": 52000, "运营费用": 18000},
                {"月份": "2026-02-01", "收入": 115000, "主营成本": 59000, "运营费用": 20000},
                {"月份": "2026-03-01", "收入": 128000, "主营成本": 65000, "运营费用": 22000},
            ],
            formula_cols={
                5: lambda r: _guard(r, "A", f"=B{r}-C{r}"),
                6: lambda r: _guard(r, "A", f"=E{r}-D{r}"),
                7: lambda r: _guard(r, "A", F.share(f"E{r}", f"B{r}")),
                8: lambda r: _guard(r, "A", F.share(f"F{r}", f"B{r}")),
                9: lambda r: _guard(r, "A", F.roi(f"F{r}", f"C{r}+D{r}")),
                10: lambda r: _guard(r, "A", f"=IF(ROW()=4,0,IFERROR((B{r}-B{r-1})/B{r-1},0))"),
            },
            summary_rows=[
                ("收入合计", lambda s, e: f"=SUM(Data!B{s}:B{e})"),
                ("净利润合计", lambda s, e: f"=SUM(Data!F{s}:F{e})"),
                ("整体毛利率", lambda s, e: f"=IFERROR(SUM(Data!E{s}:E{e})/SUM(Data!B{s}:B{e}),0)"),
                ("整体净利率", lambda s, e: f"=IFERROR(SUM(Data!F{s}:F{e})/SUM(Data!B{s}:B{e}),0)"),
                ("平均收入增长率", lambda s, e: f"=AVERAGE(Data!J{s}:J{e})"),
            ],
            input_cols=[1, 2, 3, 4],
            calc_cols=[5, 6, 7, 8, 9, 10],
            money_cols=[2, 3, 4, 5, 6],
            percent_cols=[7, 8, 9, 10],
            date_cols=[1],
            notes=["收入、成本、毛利、毛利率、净利润、净利率、ROI 和增长率均由公式生成；不能替代正式财务报表或审计。"],
            chart="line",
        ),
        "generic_table": TemplateSpec(
            table_type="generic_table",
            title="通用表格模板",
            headers=["日期", "类别", "事项", "数量", "单价", "金额", "备注"],
            sample_rows=[
                {"日期": "2026-05-01", "类别": "示例", "事项": "事项A", "数量": 2, "单价": 100, "备注": ""},
                {"日期": "2026-05-02", "类别": "示例", "事项": "事项B", "数量": 3, "单价": 80, "备注": ""},
            ],
            formula_cols={6: lambda r: _guard(r, "A", f"=D{r}*E{r}")},
            summary_rows=[
                ("记录数", lambda s, e: f"=COUNTA(Data!A{s}:A{e})"),
                ("金额合计", lambda s, e: f"=SUM(Data!F{s}:F{e})"),
            ],
            input_cols=[1, 2, 3, 4, 5, 7],
            calc_cols=[6],
            money_cols=[5, 6],
            date_cols=[1],
            integer_cols=[4],
            notes=["可作为未知表格类型的基础模板。"],
            chart="bar",
        ),
    }


def create_workbook(table_type: str, output: str | Path | None = None) -> Path:
    resolved = normalize_table_type(table_type)
    if resolved == "dashboard":
        from .dashboard_builder import build_dashboard_workbook

        return build_dashboard_workbook(output)
    specs = _specs()
    spec = specs.get(resolved, specs["generic_table"])
    return _build_standard_workbook(spec, output)


SALES_ALIASES = {
    "Date": ["date", "日期", "订单日期", "下单日期"],
    "OrderID": ["orderid", "order_id", "订单号", "订单id"],
    "Product": ["product", "产品", "商品", "商品名称"],
    "Category": ["category", "品类", "分类", "类目"],
    "Region": ["region", "区域", "地区", "城市"],
    "Qty": ["qty", "quantity", "数量", "销量", "件数"],
    "UnitPrice": ["unitprice", "unit_price", "单价", "售价"],
    "UnitCost": ["unitcost", "unit_cost", "成本", "单位成本"],
}


def _pick_column(df: pd.DataFrame, aliases: list[str]) -> str | None:
    normalized = {str(col).strip().lower().replace(" ", "").replace("-", "_"): col for col in df.columns}
    for alias in aliases:
        key = alias.lower().replace(" ", "").replace("-", "_")
        if key in normalized:
            return normalized[key]
    return None


def _standardize_sales_df(df: pd.DataFrame) -> pd.DataFrame:
    output = pd.DataFrame()
    for target, aliases in SALES_ALIASES.items():
        col = _pick_column(df, aliases)
        output[target] = df[col] if col is not None else None

    output["Date"] = pd.to_datetime(output["Date"], errors="coerce")
    fallback_ids = pd.Series([f"ORDER-{i+1:04d}" for i in range(len(output))], index=output.index)
    order_text = output["OrderID"].astype(str).str.strip()
    output["OrderID"] = output["OrderID"].where(output["OrderID"].notna() & (order_text != ""), fallback_ids)
    output["Product"] = output["Product"].fillna("未命名产品")
    output["Category"] = output["Category"].fillna("未分类")
    output["Region"] = output["Region"].fillna("未填写")
    for col in ["Qty", "UnitPrice", "UnitCost"]:
        output[col] = pd.to_numeric(output[col], errors="coerce").fillna(0)
    output = output.dropna(subset=["Date"]).reset_index(drop=True)
    if output.empty:
        raise ValueError("销售数据中没有可识别的日期行。")
    return output


def analyze_sales_file(input_path: str | Path, output: str | Path | None = None) -> Path:
    df = _standardize_sales_df(read_table(input_path))
    output_path = ensure_output_path(output, "sales_report.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    _make_instruction_sheet(
        wb,
        "销售月报 + 图表",
        "sales_report",
        ["Data 保存源数据和公式列；Summary 使用 SUMIFS/COUNTIFS 公式按品类和月份汇总。"],
    )
    data = wb.create_sheet("Data")
    summary = wb.create_sheet("Summary")

    headers = ["日期", "订单号", "产品", "品类", "区域", "数量", "单价", "单位成本", "月份", "销售额", "毛利", "毛利率"]
    header_row = 3
    start_row = 4
    end_row = start_row + len(df) - 1

    data.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    data["A1"] = "销售明细数据"
    apply_title_style(data, 1, 1, len(headers))
    for col, header in enumerate(headers, 1):
        data.cell(header_row, col, header)
    apply_header_style(data, header_row)

    for offset, record in enumerate(df.to_dict("records"), start=start_row):
        values = [
            record["Date"].to_pydatetime(),
            record["OrderID"],
            record["Product"],
            record["Category"],
            record["Region"],
            record["Qty"],
            record["UnitPrice"],
            record["UnitCost"],
        ]
        for col, value in enumerate(values, 1):
            data.cell(offset, col, value)
        data.cell(offset, 9, f'=TEXT(A{offset},"yyyy-mm")')
        data.cell(offset, 10, f"=F{offset}*G{offset}")
        data.cell(offset, 11, f"=F{offset}*(G{offset}-H{offset})")
        data.cell(offset, 12, f"=IFERROR(K{offset}/J{offset},0)")

    apply_input_style(data, start_row, end_row, range(1, 9))
    apply_calculation_style(data, start_row, end_row, range(9, 13))
    apply_number_formats(data, money_cols=[7, 8, 10, 11], percent_cols=[12], date_cols=[1], integer_cols=[6], min_row=start_row, max_row=end_row)
    freeze_and_filter(data, "A4", f"A{header_row}:L{end_row}")
    set_reasonable_column_widths(data)
    apply_print_settings(data)

    _write_sales_summary(summary, df, start_row, end_row)
    wb.save(output_path)
    return output_path


def _write_sales_summary(ws, df: pd.DataFrame, data_start: int, data_end: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws["A1"] = "销售月报汇总"
    apply_title_style(ws, 1, 1, 9)

    kpis = [
        ("销售额合计", f"=SUM(Data!J{data_start}:J{data_end})"),
        ("毛利合计", f"=SUM(Data!K{data_start}:K{data_end})"),
        ("整体毛利率", "=IFERROR(B5/B4,0)"),
        ("订单数", f"=COUNTA(Data!B{data_start}:B{data_end})"),
        ("客单价", "=IFERROR(B4/B7,0)"),
    ]
    ws["A3"] = "KPI"
    ws["B3"] = "结果"
    apply_header_style(ws, 3)
    for row, (label, formula) in enumerate(kpis, start=4):
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
    apply_summary_style(ws, 4, 8, 1, 2)
    ws["B4"].number_format = '#,##0.00;[Red]-#,##0.00'
    ws["B5"].number_format = '#,##0.00;[Red]-#,##0.00'
    ws["B6"].number_format = "0.00%"
    ws["B7"].number_format = "0"
    ws["B8"].number_format = '#,##0.00;[Red]-#,##0.00'

    categories = sorted(str(x) for x in df["Category"].dropna().unique())
    ws["A10"] = "品类"
    ws["B10"] = "销售额"
    ws["C10"] = "毛利"
    ws["D10"] = "销售占比"
    apply_header_style(ws, 10)
    for row, category in enumerate(categories, start=11):
        ws.cell(row, 1, category)
        ws.cell(row, 2, f'=SUMIFS(Data!$J${data_start}:$J${data_end},Data!$D${data_start}:$D${data_end},A{row})')
        ws.cell(row, 3, f'=SUMIFS(Data!$K${data_start}:$K${data_end},Data!$D${data_start}:$D${data_end},A{row})')
        ws.cell(row, 4, f"=IFERROR(B{row}/$B$4,0)")
    if categories:
        apply_summary_style(ws, 11, 10 + len(categories), 1, 4)
        apply_number_formats(ws, money_cols=[2, 3], percent_cols=[4], min_row=11, max_row=10 + len(categories))

    months = sorted(pd.to_datetime(df["Date"]).dt.strftime("%Y-%m").unique())
    ws["F10"] = "月份"
    ws["G10"] = "销售额"
    ws["H10"] = "毛利"
    apply_header_style(ws, 10)
    for row, month in enumerate(months, start=11):
        ws.cell(row, 6, month)
        ws.cell(row, 7, f'=SUMIFS(Data!$J${data_start}:$J${data_end},Data!$I${data_start}:$I${data_end},F{row})')
        ws.cell(row, 8, f'=SUMIFS(Data!$K${data_start}:$K${data_end},Data!$I${data_start}:$I${data_end},F{row})')
    if months:
        apply_summary_style(ws, 11, 10 + len(months), 6, 8)
        apply_number_formats(ws, money_cols=[7, 8], min_row=11, max_row=10 + len(months))

    if categories:
        add_bar_chart(ws, "品类销售额", data_min_col=2, data_max_col=2, data_min_row=10, data_max_row=10 + len(categories), cats_col=1, cats_min_row=11, cats_max_row=10 + len(categories), anchor="J3")
    if months:
        add_line_chart(ws, "月度销售趋势", data_min_col=7, data_max_col=7, data_min_row=10, data_max_row=10 + len(months), cats_col=6, cats_min_row=11, cats_max_row=10 + len(months), anchor="J18")
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)
