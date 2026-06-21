"""Helpers for saving structured reports."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook

from .io_utils import ensure_output_path, save_json
from .style_library import apply_header_style, apply_title_style, set_reasonable_column_widths


def write_json_report(report: dict[str, Any], output: str | Path) -> Path:
    return save_json(report, output)


def write_validation_workbook(report: dict[str, Any], output: str | Path | None = None) -> Path:
    output_path = ensure_output_path(output, "validation_report.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "ValidationReport"
    ws["A1"] = "工作簿校验报告"
    apply_title_style(ws, 1, 1, 4)
    ws.append([])
    ws.append(["级别", "检查项", "sheet", "说明"])
    apply_header_style(ws, 3)
    row = 4
    for level in ("errors", "warnings", "checks"):
        for item in report.get(level, []):
            if isinstance(item, dict):
                ws.cell(row, 1, level)
                ws.cell(row, 2, item.get("check", ""))
                ws.cell(row, 3, item.get("sheet", ""))
                ws.cell(row, 4, item.get("message", str(item)))
            else:
                ws.cell(row, 1, level)
                ws.cell(row, 4, str(item))
            row += 1
    set_reasonable_column_widths(ws)
    wb.save(output_path)
    return output_path

