"""Human-facing workbook and report preview helpers."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def workbook_preview(
    path: str | Path,
    *,
    max_rows: int = 80,
    max_columns: int = 24,
) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    sheets = []
    for ws in workbook.worksheets:
        header_row, headers = _find_header(ws)
        data_start = header_row + 1 if header_row else 1
        data_end = ws.max_row
        if header_row and ws.auto_filter.ref:
            try:
                _, filter_min_row, _, filter_max_row = range_boundaries(ws.auto_filter.ref)
                if filter_min_row == header_row:
                    data_end = min(data_end, filter_max_row)
            except ValueError:
                pass
        rows: list[list[Any]] = []
        formula_count = 0
        for row in ws.iter_rows(
            min_row=data_start,
            max_row=min(data_end, data_start + max_rows - 1),
            min_col=1,
            max_col=min(len(headers) or ws.max_column, max_columns),
        ):
            rendered = []
            has_input_value = False
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    value = "自动计算"
                elif value not in (None, ""):
                    has_input_value = True
                rendered.append(value)
            if not has_input_value and any(value == "自动计算" for value in rendered):
                continue
            if not any(value not in (None, "") for value in rendered):
                continue
            rows.append(rendered)
        sheets.append(
            {
                "name": ws.title,
                "rows": rows,
                "title": ws["A1"].value,
                "header_row": header_row,
                "headers": headers,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "formula_count": formula_count,
                "chart_count": len(ws._charts),
            }
        )
    return {"file": str(path), "sheet_names": workbook.sheetnames, "sheets": sheets}


def sheet_preview_dataframe(sheet: dict[str, Any]) -> pd.DataFrame:
    rows = list(sheet.get("rows", []))
    width = max((len(row) for row in rows), default=1)
    normalized: list[list[str]] = [
        [_display_value(value) for value in row] + [""] * (width - len(row))
        for row in rows
    ]
    headers = [str(item) for item in sheet.get("headers", [])][:width]
    if headers:
        headers = _unique_headers(headers)
        return pd.DataFrame(normalized, columns=headers)
    return pd.DataFrame(normalized)


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            if getattr(value, "hour", 0) or getattr(value, "minute", 0):
                return value.strftime("%Y-%m-%d %H:%M")
            return value.strftime("%Y-%m-%d")
        except (ValueError, AttributeError):
            pass
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def _find_header(ws) -> tuple[int | None, list[str]]:
    for row_index in range(1, min(ws.max_row, 15) + 1):
        values = [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]
        nonempty = [value for value in values if value not in (None, "")]
        if len(nonempty) < 2:
            continue
        last = max(
            index
            for index, value in enumerate(values, start=1)
            if value not in (None, "")
        )
        headers = [
            str(value).strip() if value not in (None, "") else f"第{index}列"
            for index, value in enumerate(values[:last], start=1)
        ]
        return row_index, headers
    return None, []


def _unique_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for header in headers:
        counts[header] = counts.get(header, 0) + 1
        result.append(header if counts[header] == 1 else f"{header}_{counts[header]}")
    return result


def combined_revision_prompt(
    subjective: dict[str, Any],
    validation: dict[str, Any],
) -> str:
    lines: list[str] = []
    for review in subjective.get("reviews", []):
        for item in review.get("concerns", []):
            lines.append(f"修正问题：{item}")
        for item in review.get("suggestions", []):
            lines.append(f"采用建议：{item}")
    for item in validation.get("issues", []):
        lines.append(f"修正检查问题：{item.get('message', item)}")
    for item in validation.get("warnings", []):
        lines.append(f"处理检查提醒：{item.get('message', item)}")
    return "\n".join(dict.fromkeys(lines))


def workbook_chart_preview(preview: dict[str, Any]) -> tuple[pd.DataFrame, str] | None:
    """Return a small browser chart dataset when the workbook contains charts."""

    if sum(int(sheet.get("chart_count", 0)) for sheet in preview.get("sheets", [])) == 0:
        return None
    for sheet in preview.get("sheets", []):
        headers = [str(item) for item in sheet.get("headers", [])]
        rows = sheet.get("rows", [])
        if len(headers) < 2 or not rows:
            continue
        frame = pd.DataFrame(rows, columns=headers[: len(rows[0])])
        numeric_columns: list[str] = []
        numeric_data: dict[str, pd.Series] = {}
        for column in frame.columns[1:]:
            converted = pd.to_numeric(frame[column], errors="coerce")
            if converted.notna().sum() >= 2:
                numeric_columns.append(str(column))
                numeric_data[str(column)] = converted
            if len(numeric_columns) >= 3:
                break
        if not numeric_columns:
            continue
        category_name = str(frame.columns[0])
        category = frame.iloc[:, 0].map(_display_value)
        chart_frame = pd.DataFrame(
            {column: numeric_data[column].to_numpy() for column in numeric_columns},
            index=category,
        ).dropna(how="all")
        if chart_frame.empty:
            continue
        kind = "line" if any(word in category_name.lower() for word in ("日期", "月份", "date", "month")) else "bar"
        return chart_frame.head(30), kind
    return None
