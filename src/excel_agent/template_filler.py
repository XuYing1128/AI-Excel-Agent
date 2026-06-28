"""Build a workbook *from* a user-provided template by filling its sheets.

The user often hands us a template that is the only acceptable format (data
import requires it). The right thing is to:

1. Clone the template (preserves every sheet's styles, headers, merges).
2. For each sheet, locate the real header row.
3. Wipe any sample rows the template carried below the header.
4. Fill each sheet by matching column headers to the uploaded data's columns.
5. Compute "derived" sheets like 导出计数_姓名 (group counts) from the data.
6. If the request is an exam-scheduling task, assign rooms/times via
   :mod:`schedule_planner` and fill the matching columns.

This is intentionally narrow and deterministic. The point is not to "design"
anything — just respect the template the user already designed.
"""

from __future__ import annotations

import re
import shutil
import stat
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from .io_utils import convert_legacy_xls, read_table
from .schedule_planner import (
    DEFAULT_SEATS_PER_ROOM,
    detect_exam_schedule_request,
    parse_capacity,
    parse_day_count,
    parse_rooms,
    parse_slots,
    schedule_registrations,
)


HEADER_SCAN_LIMIT = 25


def fill_template(
    template_path: str | Path,
    data_paths: list[str | Path],
    output_path: str | Path,
    *,
    prompt: str = "",
    title_override: str | None = None,
    task_dir: str | Path | None = None,
) -> dict[str, Any]:
    """Return ``{"sheets_filled": [...], "notices": [...]}`` after writing
    a filled copy of ``template_path`` to ``output_path``.

    No model calls. Pure data + structure preservation.
    """

    template = Path(template_path)
    if template.suffix.lower() == ".xls":
        # The basic xls→xlsx converter is enough here: we only need the
        # structure (sheets/headers), not pixel-perfect styles.
        template = convert_legacy_xls(template, Path(task_dir or template.parent) / "converted")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template, output)
    # WeChat / downloaded templates are often read-only. The generated output
    # must be editable so openpyxl can save filled rows and formulas.
    output.chmod(output.stat().st_mode | stat.S_IWRITE | stat.S_IREAD)

    frames = [read_table(path) for path in data_paths]
    combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # Try to infer the session number ("264次") from the prompt, data filenames,
    # or a "考次" column so the title gets rewritten from the template's stale
    # value (the user's template was a previous semester's file).
    session_text = (
        prompt
        + " ".join(Path(p).name for p in data_paths)
        + (" ".join(str(v) for v in combined.get("考次", pd.Series(dtype=str)).head(3))
           if "考次" in combined.columns else "")
    )
    session_match = re.search(r"(\d{2,4})\s*次", session_text)
    inferred_session = session_match.group(1) if session_match else None

    wb = load_workbook(output)
    sheets_filled: list[str] = []
    notices: list[str] = []

    # Set up a schedule context if the template HAS a scheduling sheet (with
    # 考点/考场地址/考试时间 columns). Don't gate on prompt keywords - if the
    # user's template needs those fields filled, fill them.
    schedule_ctx: dict[str, Any] | None = None
    has_schedule_sheet = any(
        _is_schedule_sheet(_find_header(ws)[1]) for ws in wb.worksheets
    )
    if has_schedule_sheet and not combined.empty:
        day_count = parse_day_count(prompt)
        schedule_ctx = {
            "rooms": parse_rooms(prompt),
            "seats_per_room": parse_capacity(prompt),
            "slots": parse_slots(prompt, day_count=day_count),
        }
        notices.append(
            f"已自动编排考试：{len(schedule_ctx['rooms'])} 个考场×"
            f"{schedule_ctx['seats_per_room']} 座×{len(schedule_ctx['slots'])} 场次。"
        )

    for ws in wb.worksheets:
        result = _fill_sheet(
            ws, combined, prompt, schedule_ctx, title_override, inferred_session
        )
        if result.get("filled"):
            sheets_filled.append(ws.title)
        if result.get("notice"):
            notices.append(result["notice"])
    if inferred_session:
        notices.append(f"已根据数据将标题中的考次更新为 {inferred_session} 次。")

    wb.save(output)
    return {"output_file": str(output), "sheets_filled": sheets_filled, "notices": notices}


def _fill_sheet(
    ws,
    data: pd.DataFrame,
    prompt: str,
    schedule_ctx: dict[str, Any] | None,
    title_override: str | None,
    inferred_session: str | None = None,
) -> dict[str, Any]:
    header_row, headers = _find_header(ws)
    if not header_row or not headers:
        return {"filled": False}

    # Build name->column-indices map (1-based) and a normalized lookup. Some
    # teacher templates intentionally repeat a header (for example two
    # "考场地址" columns); fill all matching columns rather than silently leaving
    # the later duplicate blank.
    column_indices: dict[str, list[int]] = {}
    for idx, name in enumerate(headers, start=1):
        if name:
            column_indices.setdefault(name, []).append(idx)
    normalized_index: dict[str, list[int]] = {}
    for name, indices in column_indices.items():
        normalized_index.setdefault(_normalize(name), []).extend(indices)

    # Wipe any sample rows the template carried below the header so the new
    # data is clean.
    _clear_below_header(ws, header_row)

    rows = _rows_for_sheet(ws, data, prompt, schedule_ctx)
    if not rows:
        # Even with no rows, trim ghost rows the template carried so the result
        # doesn't look "messy" with hundreds of blank lines.
        _trim_to_row(ws, header_row)
        return {"filled": False}

    for offset, record in enumerate(rows, start=header_row + 1):
        for header, value in record.items():
            targets = column_indices.get(header)
            if targets is None:
                targets = normalized_index.get(_normalize(header))
            if targets is None:
                continue
            for target in targets:
                cell = ws.cell(offset, target)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = value
    # After writing N rows, drop any leftover blank rows past the new data so
    # max_row reflects reality and Excel doesn't show hundreds of empty lines.
    last_written = header_row + len(rows)
    _trim_to_row(ws, last_written)

    # Update the session number in the title if we inferred a different one
    # (templates from previous semesters carry their old "263次" / etc.).
    if inferred_session:
        for r in range(1, min(3, ws.max_row + 1)):
            for c in range(1, min(ws.max_column, 4) + 1):
                cell = ws.cell(r, c)
                if isinstance(cell, MergedCell) or not isinstance(cell.value, str):
                    continue
                if re.search(r"\d{2,4}\s*次", cell.value) and inferred_session not in cell.value:
                    cell.value = re.sub(r"\d{2,4}\s*次", f"{inferred_session}次", cell.value)
    if title_override:
        title_cell = ws.cell(1, 1)
        if not isinstance(title_cell, MergedCell) and isinstance(title_cell.value, str):
            if any(word in title_cell.value for word in ("名册", "报考", "总信息")):
                title_cell.value = re.sub(r"\d+次", title_override, title_cell.value)

    notice = ""
    if ws.title.startswith("导出计数") or "计数" in ws.title:
        notice = f"已按姓名分组生成 {len(rows)} 条计数。"
    elif _is_schedule_sheet(headers):
        notice = f"已为 {len(rows)} 条记录分配考场和时间。"
    return {"filled": True, "notice": notice}


def _rows_for_sheet(
    ws,
    data: pd.DataFrame,
    prompt: str,
    schedule_ctx: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    if data.empty:
        return []

    _, headers = _find_header(ws)
    header_set = {h for h in headers if h}

    # Derived sheet: 导出计数_姓名 → group counts by 姓名
    if "姓名" in header_set and ("计数" in header_set or "占比" in header_set):
        counts = data["姓名"].value_counts(dropna=True)
        total = int(counts.sum()) or 1
        out = []
        for name, count in counts.items():
            out.append({"姓名": name, "计数": int(count), "占比": float(count) / total})
        return out

    # Standard data sheet: write all rows, mapping by header name
    records: list[dict[str, Any]] = []
    for _, row in data.iterrows():
        record: dict[str, Any] = {}
        for header in header_set:
            if header in data.columns:
                value = row[header]
                if pd.isna(value):
                    value = None
                record[header] = value
        records.append(record)

    # If this is the scheduling sheet and we have a context, assign rooms/times.
    if schedule_ctx and _is_schedule_sheet(headers):
        registrations = data[["准考证号"]].to_dict("records") if "准考证号" in data.columns else [
            {} for _ in records
        ]
        assignments = schedule_registrations(
            registrations,
            rooms=schedule_ctx["rooms"],
            seats_per_room=schedule_ctx["seats_per_room"],
            slots=schedule_ctx["slots"],
        )
        for record, assignment in zip(records, assignments):
            if assignment is None:
                record.setdefault("考点", "—（超出容量）")
                continue
            # Default 考点 to the source school if the prompt or data mentions one,
            # otherwise leave blank so the user can fill it.
            text_pool = prompt + " " + " ".join(
                str(data.iloc[0].get(col, "")) for col in ("院校名称", "助学点名称")
                if col in data.columns
            )
            school = next(
                (
                    candidate
                    for candidate in ("成都文理学院", "成都外国语学院", "川外", "文理学院")
                    if candidate in text_pool
                ),
                "",
            )
            record["考点"] = school
            record["考场"] = assignment.room
            record["考场地址"] = assignment.room
            record["考试日期"] = assignment.date_label
            record["开始时间"] = assignment.start
            record["结束时间"] = assignment.end
            record["考试时间"] = f"{assignment.date_label} {assignment.start}-{assignment.end}"

    return records


def _find_header(ws) -> tuple[int, list[str]]:
    """Return (row, header_labels) — the densest non-trivial text row in the
    top of the sheet. Mirrors :func:`template_adapter.find_template_header`
    but tolerates merged title cells above the headers."""

    best = (0, [])
    best_score = 0
    for row in range(1, min(ws.max_row, HEADER_SCAN_LIMIT) + 1):
        values = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                values.append("")
                continue
            values.append("" if cell.value is None else str(cell.value).strip())
        non_blank = sum(1 for v in values if v)
        if non_blank < 2:
            continue
        # heuristic: prefer rows with mostly short text labels (header-y)
        text_ratio = sum(1 for v in values if v and not _looks_numeric(v))
        score = non_blank + text_ratio
        if score > best_score:
            best_score = score
            best = (row, values)
    return best


def _clear_below_header(ws, header_row: int) -> None:
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _trim_to_row(ws, last_row: int) -> None:
    """Delete rows past ``last_row`` so the sheet's max_row reflects real data.

    openpyxl ``delete_rows`` does honour merged ranges, so unmerge any range
    that extends into the doomed area first.
    """

    if ws.max_row <= last_row:
        return
    to_remove = ws.max_row - last_row
    # Unmerge merges that touch the rows we're about to delete.
    for merged in list(ws.merged_cells.ranges):
        if merged.max_row > last_row:
            ws.unmerge_cells(str(merged))
    ws.delete_rows(last_row + 1, to_remove)


def _is_schedule_sheet(headers: list[str]) -> bool:
    header_set = {h for h in headers if h}
    return bool(header_set & {"考点", "考场地址", "考试时间", "开始时间", "考场"})


def _normalize(value: Any) -> str:
    return re.sub(r"[\s（）()_\-—:：]", "", str(value or "")).lower()


def _looks_numeric(value: str) -> bool:
    try:
        float(str(value).replace(",", ""))
        return True
    except ValueError:
        return False


__all__ = ["fill_template"]
