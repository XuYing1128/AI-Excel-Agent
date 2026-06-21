"""Dashboard workbook builder."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from .chart_library import add_bar_chart, add_line_chart, add_pie_chart
from .io_utils import ensure_output_path
from .style_library import (
    apply_calculation_style,
    apply_header_style,
    apply_input_style,
    apply_number_formats,
    apply_print_settings,
    apply_summary_style,
    apply_title_style,
    freeze_and_filter,
    set_reasonable_column_widths,
    style_instruction_sheet,
)


def build_dashboard_workbook(output: str | Path | None = None) -> Path:
    output_path = ensure_output_path(output, "dashboard.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    instructions = wb.create_sheet("Instructions")
    instructions["A1"] = "综合 Dashboard 使用说明"
    instructions["A3"] = "结构"
    instructions["B3"] = "Data 保存明细数据；Summary 保存汇总公式；Dashboard 保存 KPI、Top N、趋势和图表。"
    instructions["A4"] = "公式"
    instructions["B4"] = "Summary 和 Dashboard 中的 KPI、占比、趋势均通过 Excel 公式引用 Data。"
    instructions["A5"] = "校验"
    instructions["B5"] = "生成后运行 python -m excel_agent.cli validate --input outputs/dashboard.xlsx。"
    style_instruction_sheet(instructions)

    source = wb.create_sheet("Data")
    summary = wb.create_sheet("Summary")
    dash = wb.create_sheet("Dashboard")
    _write_source_data(source)
    _write_dashboard_summary(summary)
    _write_dashboard(dash)

    wb.save(output_path)
    return output_path


def _write_source_data(ws) -> None:
    headers = ["月份", "渠道", "品类", "销售额", "成本", "订单数", "毛利", "毛利率", "客单价"]
    rows = [
        ("2026-01", "线上", "饮品", 120000, 62000, 1800),
        ("2026-01", "门店", "饮品", 90000, 48000, 1200),
        ("2026-02", "线上", "零食", 135000, 71000, 1900),
        ("2026-02", "门店", "日用", 76000, 39000, 860),
        ("2026-03", "线上", "饮品", 158000, 79000, 2200),
        ("2026-03", "门店", "零食", 98000, 53000, 1100),
    ]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = "Dashboard 源数据"
    apply_title_style(ws, 1, 1, len(headers))
    for col, header in enumerate(headers, 1):
        ws.cell(3, col, header)
    apply_header_style(ws, 3)
    for row_idx, row in enumerate(rows, 4):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)
        ws.cell(row_idx, 7, f"=D{row_idx}-E{row_idx}")
        ws.cell(row_idx, 8, f"=IFERROR(G{row_idx}/D{row_idx},0)")
        ws.cell(row_idx, 9, f"=IFERROR(D{row_idx}/F{row_idx},0)")
    apply_input_style(ws, 4, 3 + len(rows), range(1, 7))
    apply_calculation_style(ws, 4, 3 + len(rows), range(7, 10))
    apply_number_formats(ws, money_cols=[4, 5, 7, 9], percent_cols=[8], integer_cols=[6], min_row=4, max_row=3 + len(rows))
    freeze_and_filter(ws, "A4", f"A3:I{3 + len(rows)}")
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _write_dashboard_summary(ws) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"] = "综合经营汇总"
    apply_title_style(ws, 1, 1, 8)

    kpis = [
        ("销售额", "=SUM(Data!D4:D9)"),
        ("毛利", "=SUM(Data!G4:G9)"),
        ("毛利率", "=IFERROR(B4/B3,0)"),
        ("订单数", "=SUM(Data!F4:F9)"),
        ("客单价", "=IFERROR(B3/B6,0)"),
        ("校验点：数据行数", '=IF(COUNTA(Data!A4:A9)>0,"PASS","CHECK")'),
    ]
    ws["A2"] = "KPI"
    ws["B2"] = "结果"
    apply_header_style(ws, 2)
    for row, (label, formula) in enumerate(kpis, 3):
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
    apply_summary_style(ws, 3, 8, 1, 2)
    ws["B3"].number_format = '#,##0.00;[Red]-#,##0.00'
    ws["B4"].number_format = '#,##0.00;[Red]-#,##0.00'
    ws["B5"].number_format = "0.00%"
    ws["B6"].number_format = "0"
    ws["B7"].number_format = '#,##0.00;[Red]-#,##0.00'

    ws["A11"] = "品类"
    ws["B11"] = "销售额"
    ws["C11"] = "占比"
    for row, category in enumerate(["饮品", "零食", "日用"], 12):
        ws.cell(row, 1, category)
        ws.cell(row, 2, f'=SUMIFS(Data!$D$4:$D$9,Data!$C$4:$C$9,A{row})')
        ws.cell(row, 3, f"=IFERROR(B{row}/$B$3,0)")
    apply_header_style(ws, 11)
    apply_summary_style(ws, 12, 14, 1, 3)
    apply_number_formats(ws, money_cols=[2], percent_cols=[3], min_row=12, max_row=14)

    ws["E11"] = "月份"
    ws["F11"] = "销售额"
    ws["G11"] = "毛利"
    for row, month in enumerate(["2026-01", "2026-02", "2026-03"], 12):
        ws.cell(row, 5, month)
        ws.cell(row, 6, f'=SUMIFS(Data!$D$4:$D$9,Data!$A$4:$A$9,E{row})')
        ws.cell(row, 7, f'=SUMIFS(Data!$G$4:$G$9,Data!$A$4:$A$9,E{row})')
    apply_header_style(ws, 11)
    apply_summary_style(ws, 12, 14, 5, 7)
    apply_number_formats(ws, money_cols=[6, 7], min_row=12, max_row=14)
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _write_dashboard(ws) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"] = "综合经营 Dashboard"
    apply_title_style(ws, 1, 1, 8)

    kpis = [
        ("销售额", "=Summary!B3"),
        ("毛利", "=Summary!B4"),
        ("毛利率", "=IFERROR(B4/B3,0)"),
        ("订单数", "=Summary!B6"),
        ("客单价", "=IFERROR(B3/B6,0)"),
    ]
    ws["A2"] = "KPI"
    ws["B2"] = "结果"
    apply_header_style(ws, 2)
    for row, (label, formula) in enumerate(kpis, 3):
        ws.cell(row, 1, label)
        ws.cell(row, 2, formula)
    apply_summary_style(ws, 3, 7, 1, 2)
    ws["B3"].number_format = '#,##0.00;[Red]-#,##0.00'
    ws["B4"].number_format = '#,##0.00;[Red]-#,##0.00'
    ws["B5"].number_format = "0.00%"
    ws["B6"].number_format = "0"
    ws["B7"].number_format = '#,##0.00;[Red]-#,##0.00'

    ws["A10"] = "品类"
    ws["B10"] = "销售额"
    ws["C10"] = "占比"
    for row, category in enumerate(["饮品", "零食", "日用"], 11):
        ws.cell(row, 1, category)
        ws.cell(row, 2, f'=SUMIFS(Data!$D$4:$D$9,Data!$C$4:$C$9,A{row})')
        ws.cell(row, 3, f"=IFERROR(B{row}/$B$3,0)")
    apply_header_style(ws, 10)
    apply_summary_style(ws, 11, 13, 1, 3)
    apply_number_formats(ws, money_cols=[2], percent_cols=[3], min_row=11, max_row=13)

    ws["E10"] = "月份"
    ws["F10"] = "销售额"
    ws["G10"] = "毛利"
    for row, month in enumerate(["2026-01", "2026-02", "2026-03"], 11):
        ws.cell(row, 5, month)
        ws.cell(row, 6, f'=SUMIFS(Data!$D$4:$D$9,Data!$A$4:$A$9,E{row})')
        ws.cell(row, 7, f'=SUMIFS(Data!$G$4:$G$9,Data!$A$4:$A$9,E{row})')
    apply_header_style(ws, 10)
    apply_summary_style(ws, 11, 13, 5, 7)
    apply_number_formats(ws, money_cols=[6, 7], min_row=11, max_row=13)

    add_pie_chart(ws, "分类占比", data_col=2, data_min_row=11, data_max_row=13, cats_col=1, cats_min_row=11, cats_max_row=13, anchor="A16")
    add_line_chart(ws, "月度趋势", data_min_col=6, data_max_col=7, data_min_row=10, data_max_row=13, cats_col=5, cats_min_row=11, cats_max_row=13, anchor="E16")
    add_bar_chart(ws, "Top 品类", data_min_col=2, data_max_col=2, data_min_row=10, data_max_row=13, cats_col=1, cats_min_row=11, cats_max_row=13, anchor="J3")
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)
