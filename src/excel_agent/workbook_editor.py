"""Conservative workbook editing helpers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .io_utils import backup_file, ensure_output_path
from .style_library import style_instruction_sheet


def ensure_instructions_sheet(input_path: str | Path, output: str | Path | None = None, make_backup: bool = True) -> Path:
    source = Path(input_path)
    if make_backup and source.exists():
        backup_file(source)
    wb = load_workbook(source)
    if "Instructions" not in wb.sheetnames:
        ws = wb.create_sheet("Instructions", 0)
        ws["A1"] = "使用说明"
        ws["A3"] = "说明"
        ws["B3"] = "该说明页由 AI-Excel-Agent 自动补充。修改前已尽量保留原始工作簿。"
        style_instruction_sheet(ws)
    output_path = ensure_output_path(output or source, source.name)
    wb.save(output_path)
    return output_path

