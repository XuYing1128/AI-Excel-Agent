"""Static workbook validation with JSON reports."""

from __future__ import annotations

import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)

from .io_utils import save_json


ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"}
INVALID_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\]")
LOCAL_CELL_REF = re.compile(r"(?<![!A-Za-z0-9_])\$?([A-Z]{1,3})\$?(\d+)")
RANGE_REF = re.compile(r"\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)")

INSTRUCTION_NAMES = {"instructions", "readme", "说明", "使用说明"}
DATA_SHEET_NAMES = {"data", "input", "inputs", "sourcedata", "source", "输入", "数据", "明细"}
KEY_SHEET_NAMES = DATA_SHEET_NAMES | {"summary", "dashboard", "汇总", "仪表盘"} | INSTRUCTION_NAMES

KEY_HEADER_WORDS = ("日期", "金额", "数量", "单价", "销售额", "收入", "成本", "sku", "订单", "客户", "员工")
DATE_HEADER_WORDS = ("日期", "date", "下单", "收款日期")
MONTH_HEADER_WORDS = ("月份", "month")
MONEY_HEADER_WORDS = ("金额", "收入", "成本", "利润", "单价", "销售额", "gmv", "报价", "税额", "预算", "支出")
PERCENT_HEADER_WORDS = ("率", "占比", "roi", "执行率")
NUMERIC_HEADER_WORDS = MONEY_HEADER_WORDS + ("数量", "库存", "入库", "出库", "小时", "分钟", "订单数", "工期")
PRIMARY_KEY_WORDS = ("订单号", "orderid", "order_id", "发票号", "invoice_id")
INVENTORY_WORDS = ("期末库存", "库存")


def validate_workbook(path: str | Path, output_json: str | Path | None = None) -> dict[str, Any]:
    workbook_path = Path(path)
    report = _new_report(workbook_path)

    _check_file_level(workbook_path, report)
    if report["errors"]:
        _finalize_report(report, output_json)
        return report

    try:
        wb = load_workbook(workbook_path, data_only=False)
        cached_wb = load_workbook(workbook_path, data_only=True)
    except Exception as exc:  # pragma: no cover - depends on corrupt inputs
        _add_error(report, "open_workbook", f"openpyxl 无法打开文件: {exc}", suggestion="确认文件是真正的 xlsx/xlsm，必要时重新生成。")
        _finalize_report(report, output_json)
        return report

    summary = report["summary"]
    summary["workbook_opened"] = True
    summary["sheet_count"] = len(wb.worksheets)
    summary["visible_sheet_count"] = sum(1 for ws in wb.worksheets if ws.sheet_state == "visible")
    summary["chart_count"] = sum(len(ws._charts) for ws in wb.worksheets)
    summary["sheets"] = [
        {
            "name": ws.title,
            "visible": ws.sheet_state == "visible",
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "auto_filter": ws.auto_filter.ref,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "chart_count": len(ws._charts),
        }
        for ws in wb.worksheets
    ]

    _check_workbook_level(wb, report)
    for ws in wb.worksheets:
        cached_ws = cached_wb[ws.title] if ws.title in cached_wb.sheetnames else None
        _check_sheet_level(ws, report)
        if ws.title.startswith("_图表数据"):
            continue
        header = _header_info(ws)
        _check_headers(ws, header, report)
        _check_formula_cells(ws, cached_ws, header, report)
        _check_formatting(ws, header, report)
        _check_data_quality(ws, cached_ws, header, report)
        _check_merged_filter_area(ws, report)

    _finalize_report(report, output_json)
    return report


def inspect_workbook(path: str | Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False)
    return {
        "file": str(path),
        "sheets": [
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "hidden": ws.sheet_state != "visible",
                "auto_filter": ws.auto_filter.ref,
                "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
                "title": ws["A1"].value,
                "header_row": (
                    _header_info(ws).get("row") if _header_info(ws) is not None else None
                ),
                "headers": (
                    [
                        str(value)
                        for value in _header_info(ws).get("headers", [])
                        if not _is_blank(value)
                    ]
                    if _header_info(ws) is not None
                    else []
                ),
                "formula_columns": _formula_column_names(ws),
                "chart_count": len(ws._charts),
            }
            for ws in wb.worksheets
        ],
    }


def _formula_column_names(ws) -> list[str]:
    header = _header_info(ws)
    if header is None:
        return []
    row = int(header["row"])
    result: list[str] = []
    for column, name in enumerate(header["headers"], start=1):
        if _is_blank(name):
            continue
        if any(
            isinstance(ws.cell(data_row, column).value, str)
            and ws.cell(data_row, column).value.startswith("=")
            for data_row in range(row + 1, min(ws.max_row, row + 20) + 1)
        ):
            result.append(str(name))
    return result


def to_json(report: dict[str, Any]) -> str:
    return json.dumps(report, ensure_ascii=False, indent=2)


def _new_report(path: Path) -> dict[str, Any]:
    return {
        "status": "pass",
        "file": str(path),
        "summary": {
            "file_exists": False,
            "file_size_bytes": 0,
            "workbook_opened": False,
            "sheet_count": 0,
            "visible_sheet_count": 0,
            "blank_sheet_count": 0,
            "formula_cell_count": 0,
            "formula_cached_error_count": 0,
            "data_sheet_present": False,
            "instruction_sheet_present": False,
            "sheets": [],
        },
        "errors": [],
        "warnings": [],
        "suggestions": [],
    }


def _check_file_level(path: Path, report: dict[str, Any]) -> None:
    if not path.exists():
        _add_error(report, "file_exists", "文件不存在", suggestion="检查输入路径，或先生成目标 xlsx。")
        return
    report["summary"]["file_exists"] = True
    size = path.stat().st_size
    report["summary"]["file_size_bytes"] = size
    if size == 0:
        _add_error(report, "file_size", "文件大小为 0，无法作为有效 workbook。", suggestion="重新生成文件。")
    elif size < 1024:
        _add_error(report, "file_size", f"文件过小，可能不是有效 xlsx: {size} bytes", suggestion="确认文件格式并重新导出。")
    elif size < 4096:
        _add_warning(report, "file_size", f"文件大小偏小，请确认内容是否完整: {size} bytes")
    elif size > 50 * 1024 * 1024:
        _add_warning(report, "file_size", f"文件较大，打开和校验可能较慢: {size} bytes")


def _check_workbook_level(wb, report: dict[str, Any]) -> None:
    if len(wb.worksheets) == 0:
        _add_error(report, "empty_workbook", "workbook 不包含任何 sheet。", suggestion="至少创建一个可见 sheet。")
        return
    if report["summary"]["visible_sheet_count"] == 0:
        _add_error(report, "visible_sheet", "workbook 没有任何可见 sheet。", suggestion="至少保留一个可见 sheet。")
    has_content = any(_sheet_has_content(ws) for ws in wb.worksheets)
    if not has_content:
        _add_error(report, "empty_workbook", "workbook 所有 sheet 都为空。", suggestion="写入说明页、数据页或汇总页。")

    instruction_present = any(
        _is_instruction_sheet(ws.title) or _has_instruction_note(ws)
        for ws in wb.worksheets
    )
    data_present = any(
        _is_data_sheet(ws.title) or bool(ws.auto_filter.ref)
        for ws in wb.worksheets
    )
    report["summary"]["instruction_sheet_present"] = instruction_present
    report["summary"]["data_sheet_present"] = data_present
    if not instruction_present:
        _add_warning(report, "instruction_sheet", "缺少 Instructions/README/说明 sheet。", suggestion="添加说明页，描述输入区、公式区和校验方式。")
    if not data_present:
        _add_warning(report, "data_sheet", "未识别到 Data/Input/输入/数据 sheet。", suggestion="至少提供一个数据页或输入区。")


def _has_instruction_note(ws) -> bool:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8)):
        for cell in row:
            value = str(cell.value or "")
            if any(word in value for word in ("使用说明", "录入内容", "自动计算", "填写说明")):
                return True
    return False


def _check_sheet_level(ws, report: dict[str, Any]) -> None:
    title = ws.title
    if not title or not title.strip():
        _add_error(report, "sheet_name", "sheet 名称为空。", sheet=title, suggestion="为 sheet 设置清晰名称。")
    if len(title) > 31 or INVALID_SHEET_CHARS.search(title):
        _add_error(report, "sheet_name", "sheet 名称不符合 Excel 规则。", sheet=title, suggestion="名称不超过 31 字符且不含 []:*?/\\。")
    if not _sheet_has_content(ws):
        report["summary"]["blank_sheet_count"] += 1
        _add_warning(report, "blank_sheet", "存在完全空白 sheet。", sheet=title, suggestion="删除空 sheet 或补充说明。")
    if ws.sheet_state != "visible" and _normalized_name(title) in KEY_SHEET_NAMES:
        _add_warning(report, "hidden_key_sheet", "关键 sheet 被隐藏。", sheet=title, suggestion="将关键数据、汇总或说明 sheet 设为可见。")


def _check_headers(ws, header: dict[str, Any] | None, report: dict[str, Any]) -> None:
    if (
        _is_instruction_sheet(ws.title)
        or ("参数" in ws.title and not ws.auto_filter.ref)
        or header is None
    ):
        return
    row = header["row"]
    headers = header["headers"]
    if not headers:
        _add_warning(report, "headers", "未识别到表头。", sheet=ws.title)
        return

    empty_positions = [
        idx
        for idx, value in enumerate(headers, start=1)
        if _is_blank(value) and not _merged_header_covered(ws, row, idx)
    ]
    if empty_positions:
        _add_warning(report, "empty_header", f"表头存在空白单元格: {empty_positions[:8]}", sheet=ws.title, row=row)

    text_headers = [str(value).strip() for value in headers if not _is_blank(value)]
    lower_counts = _duplicates([item.lower() for item in text_headers])
    if lower_counts:
        _add_warning(report, "duplicate_header", f"存在重复表头: {lower_counts}", sheet=ws.title, row=row)

    unnamed = [item for item in text_headers if item.lower().startswith(("unnamed", "column_", "__empty"))]
    if unnamed:
        _add_warning(report, "unnamed_header", f"存在疑似 unnamed 列: {unnamed[:8]}", sheet=ws.title, row=row)

    nfkc_map: dict[str, list[str]] = {}
    for item in text_headers:
        norm = _normalize_header(item)
        nfkc_map.setdefault(norm, []).append(item)
    mixed = {norm: values for norm, values in nfkc_map.items() if len(set(values)) > 1}
    if mixed:
        _add_warning(report, "full_half_width_duplicate", f"疑似全角/半角或空格差异导致重复列: {mixed}", sheet=ws.title, row=row)

    if _is_data_sheet(ws.title) and not ws.auto_filter.ref:
        _add_warning(report, "auto_filter", "数据页未启用自动筛选。", sheet=ws.title, suggestion="为数据区域设置 auto_filter。")


def _check_formula_cells(ws, cached_ws, header: dict[str, Any] | None, report: dict[str, Any]) -> None:
    formulas_by_col: dict[int, list[int]] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            is_formula = isinstance(value, str) and value.startswith("=")
            if cell.data_type == "f" and not is_formula:
                _add_error(report, "formula_prefix", "公式单元格不是以 = 开头。", sheet=ws.title, cell=cell.coordinate)
            if not is_formula:
                if value in ERROR_VALUES:
                    _add_error(report, "excel_error_value", f"单元格包含 Excel 错误值: {value}", sheet=ws.title, cell=cell.coordinate)
                continue

            report["summary"]["formula_cell_count"] += 1
            formulas_by_col.setdefault(cell.column, []).append(cell.row)
            formula = str(value)
            if any(token in formula for token in ERROR_VALUES):
                _add_error(report, "formula_error_token", f"公式包含明显错误引用或错误值: {formula}", sheet=ws.title, cell=cell.coordinate)
            _check_formula_reference_bounds(ws, cell.coordinate, formula, report)
            _check_formula_range_coverage(ws, cell.coordinate, formula, header, report)
            if _has_division_risk(formula):
                _add_warning(report, "division_by_zero_risk", "公式包含除法且缺少 IFERROR/IF 保护，存在除零风险。", sheet=ws.title, cell=cell.coordinate)

            if cached_ws is not None:
                cached_value = cached_ws[cell.coordinate].value
                if cached_value in ERROR_VALUES:
                    report["summary"]["formula_cached_error_count"] += 1
                    _add_error(report, "formula_cached_error", f"公式结果缓存包含错误值: {cached_value}", sheet=ws.title, cell=cell.coordinate)

    if header and _is_data_sheet(ws.title):
        _check_formula_breaks(ws, header, formulas_by_col, report)


def _check_formula_reference_bounds(ws, coordinate: str, formula: str, report: dict[str, Any]) -> None:
    max_row_limit = max(ws.max_row + 100, 200)
    max_col_limit = max(ws.max_column + 30, 50)
    if "#REF!" in formula:
        _add_error(report, "formula_bad_reference", "公式包含 #REF!。", sheet=ws.title, cell=coordinate)
    for col_letters, row_text in LOCAL_CELL_REF.findall(formula):
        row_num = int(row_text)
        col_num = column_index_from_string(col_letters)
        if row_num > max_row_limit or col_num > max_col_limit:
            _add_warning(
                report,
                "formula_reference_bounds",
                f"公式可能引用越界单元格 {col_letters}{row_num}。",
                sheet=ws.title,
                cell=coordinate,
            )


def _check_formula_range_coverage(ws, coordinate: str, formula: str, header: dict[str, Any] | None, report: dict[str, Any]) -> None:
    if not header or not _is_data_sheet(ws.title):
        return
    data_last_row = header["last_data_row"]
    if data_last_row <= header["row"]:
        return
    for col1, row1, col2, row2 in RANGE_REF.findall(formula):
        start = int(row1)
        end = int(row2)
        if start <= header["row"] + 1 and end < data_last_row:
            _add_warning(
                report,
                "formula_range_short",
                f"公式范围 {col1}{row1}:{col2}{row2} 可能少覆盖数据行，数据末行为 {data_last_row}。",
                sheet=ws.title,
                cell=coordinate,
            )


def _check_formula_breaks(ws, header: dict[str, Any], formulas_by_col: dict[int, list[int]], report: dict[str, Any]) -> None:
    data_rows = header["data_rows"]
    if not data_rows:
        return
    for col, formula_rows in formulas_by_col.items():
        if len(formula_rows) <= 1:
            continue
        expected = [row for row in data_rows if min(formula_rows) <= row <= max(formula_rows)]
        missing = [row for row in expected if row not in formula_rows and not _is_blank_or_header_separator(ws, row)]
        if missing:
            _add_warning(
                report,
                "formula_coverage",
                f"{ws.cell(header['row'], col).value or col} 公式列存在中断行: {missing[:8]}",
                sheet=ws.title,
            )

        last_formula = max(formula_rows)
        trailing = [row for row in data_rows if row > last_formula]
        if trailing:
            _add_warning(
                report,
                "formula_coverage_short",
                f"{ws.cell(header['row'], col).value or col} 公式列可能未覆盖末尾数据行: {trailing[:8]}",
                sheet=ws.title,
            )


def _check_formatting(ws, header: dict[str, Any] | None, report: dict[str, Any]) -> None:
    if _is_instruction_sheet(ws.title) or header is None:
        return
    if _is_data_sheet(ws.title) and not ws.freeze_panes:
        _add_warning(report, "freeze_panes", "数据页未冻结窗格。", sheet=ws.title, suggestion="冻结表头行，例如 A4。")

    title_cell = ws["A1"]
    if not title_cell.value:
        _add_warning(report, "title", "缺少标题或标题区域。", sheet=ws.title)
    elif not (title_cell.font and title_cell.font.bold):
        _add_warning(report, "title_style", "标题未使用明显样式。", sheet=ws.title)

    styled_header_count = 0
    for col, header_value in enumerate(header["headers"], start=1):
        cell = ws.cell(header["row"], col)
        if not _is_blank(header_value) and ((cell.font and cell.font.bold) or (cell.fill and cell.fill.fill_type)):
            styled_header_count += 1
    if styled_header_count < max(1, len([h for h in header["headers"] if not _is_blank(h)]) // 2):
        _add_warning(report, "header_style", "表头样式不明显。", sheet=ws.title)

    for col, header_value in enumerate(header["headers"], start=1):
        header_text = str(header_value).strip().lower() if not _is_blank(header_value) else ""
        if not header_text:
            continue
        sample_cells = [ws.cell(row, col) for row in header["data_rows"][:10]]
        number_format = str(ws.cell(header["row"] + 1, col).number_format or "")
        if _matches(header_text, DATE_HEADER_WORDS):
            if not _is_date_format(number_format) and not _date_column_values_are_parseable(sample_cells):
                _add_warning(report, "date_format", f"日期列 {header_value} 可能未使用日期格式。", sheet=ws.title)
        if _matches(header_text, MONEY_HEADER_WORDS):
            if number_format == "General" and not _numeric_column_values_are_parseable(sample_cells):
                _add_warning(report, "money_format", f"金额/数字列 {header_value} 可能未使用数字格式。", sheet=ws.title)
        if _matches(header_text, PERCENT_HEADER_WORDS):
            if "%" not in number_format:
                _add_warning(report, "percent_format", f"百分比列 {header_value} 未使用百分比格式。", sheet=ws.title)

    _check_column_widths(ws, header, report)


def _check_column_widths(ws, header: dict[str, Any] | None, report: dict[str, Any]) -> None:
    used_cols = range(1, ws.max_column + 1)
    for col in used_cols:
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width
        if width and width < 6:
            _add_warning(report, "column_width", f"{letter} 列宽过窄: {width}", sheet=ws.title)
        if width and width > 60:
            _add_warning(report, "column_width", f"{letter} 列宽过大: {width}", sheet=ws.title)
        if header:
            header_value = ws.cell(header["row"], col).value
            if width and header_value and len(str(header_value)) > max(width * 1.5, 12):
                _add_warning(report, "column_width", f"{letter} 列宽可能截断表头: {header_value}", sheet=ws.title)


def _check_data_quality(ws, cached_ws, header: dict[str, Any] | None, report: dict[str, Any]) -> None:
    if _is_instruction_sheet(ws.title) or header is None:
        return
    headers = header["headers"]
    data_rows = header["non_formula_data_rows"]
    if not data_rows:
        return

    for col, header_value in enumerate(headers, start=1):
        if _is_blank(header_value):
            continue
        header_text = str(header_value).strip().lower()
        raw_cells = [ws.cell(row, col) for row in data_rows]
        formula_cells = [cell for cell in raw_cells if _is_formula(cell.value)]
        values = [_cell_effective_value(cell, cached_ws) for cell in raw_cells]
        formula_col_without_cache = bool(formula_cells) and all(_is_blank(value) for value in values)
        if formula_col_without_cache:
            continue
        nonblank = [value for value in values if not _is_blank(value)]

        if _matches(header_text, KEY_HEADER_WORDS) and values:
            blank_ratio = 1 - len(nonblank) / len(values)
            if blank_ratio >= 0.7:
                _add_warning(report, "critical_column_empty", f"关键列 {header_value} 大面积为空，空值比例 {blank_ratio:.0%}。", sheet=ws.title)

        if _matches(header_text, NUMERIC_HEADER_WORDS) and nonblank:
            text_ratio = _text_ratio_in_numeric_values(nonblank)
            if text_ratio >= 0.3:
                _add_warning(report, "numeric_text_mix", f"数字列 {header_value} 混入大量文本，比例 {text_ratio:.0%}。", sheet=ws.title)

        if _matches(header_text, DATE_HEADER_WORDS) and nonblank:
            bad_ratio = _bad_date_ratio(nonblank)
            if bad_ratio >= 0.2:
                _add_warning(report, "date_parse", f"日期列 {header_value} 存在较多无法解析值，比例 {bad_ratio:.0%}。", sheet=ws.title)

        if _matches(header_text, PRIMARY_KEY_WORDS) and nonblank:
            duplicates = _duplicates([str(value).strip() for value in nonblank])
            if duplicates:
                _add_warning(report, "duplicate_primary_key", f"主键/订单号列 {header_value} 存在重复值: {duplicates[:8]}", sheet=ws.title)

        if _matches(header_text, INVENTORY_WORDS):
            negatives = [value for value in nonblank if _to_number(value) is not None and _to_number(value) < 0]
            if negatives:
                _add_warning(report, "negative_inventory", f"库存列 {header_value} 存在负库存: {negatives[:8]}", sheet=ws.title)


def _cells_in_ref(ws, ref: str) -> set[str]:
    """Return the set of cell coordinates a string range covers.

    ``ws[ref]`` returns a Cell for a single-cell ref ("A1") and tuples-of-rows
    for a real range ("A1:C3") — we have to handle both shapes or it crashes.
    """

    region = ws[ref]
    if hasattr(region, "coordinate"):  # single Cell
        return {region.coordinate}
    cells: set[str] = set()
    for row in region:
        if hasattr(row, "coordinate"):
            cells.add(row.coordinate)
        else:
            cells.update(cell.coordinate for cell in row)
    return cells


def _check_merged_filter_area(ws, report: dict[str, Any]) -> None:
    if not ws.auto_filter.ref:
        return
    header_row = _find_header_row(ws) or 0
    filter_cells = _cells_in_ref(ws, ws.auto_filter.ref)
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= header_row:
            continue
        top_left = ws.cell(merged.min_row, merged.min_col).value
        if merged.min_row == merged.max_row and any(
            word in str(top_left or "") for word in ("小计", "总计", "合计", "汇总")
        ):
            continue
        merged_cells = _cells_in_ref(ws, str(merged))
        if filter_cells & merged_cells:
            _add_warning(report, "merged_filter_area", f"合并单元格 {merged} 与筛选区域重叠。", sheet=ws.title)


def _merged_header_covered(ws, row: int, col: int) -> bool:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col).value not in (None, "")
    return False


def _header_info(ws) -> dict[str, Any] | None:
    header_row = _find_header_row(ws)
    if header_row is None:
        return None
    raw_headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
    last_header_idx = max((idx for idx, value in enumerate(raw_headers, start=1) if not _is_blank(value)), default=0)
    headers = raw_headers[:last_header_idx]
    if not headers:
        return None

    data_rows: list[int] = []
    non_formula_data_rows: list[int] = []
    data_limit = ws.max_row
    if ws.auto_filter.ref:
        try:
            _, filter_min_row, _, filter_max_row = range_boundaries(ws.auto_filter.ref)
            if filter_min_row == header_row:
                data_limit = min(data_limit, filter_max_row)
        except ValueError:
            pass
    for row in range(header_row + 1, data_limit + 1):
        cells = [ws.cell(row, col) for col in range(1, max(last_header_idx, 1) + 1)]
        if any(not _is_blank(cell.value) for cell in cells):
            data_rows.append(row)
        if any(not _is_blank(cell.value) and not _is_formula(cell.value) for cell in cells):
            non_formula_data_rows.append(row)
    last_data_row = max(data_rows, default=header_row)
    return {
        "row": header_row,
        "headers": headers,
        "data_rows": data_rows,
        "non_formula_data_rows": non_formula_data_rows,
        "last_data_row": last_data_row,
    }


def _find_header_row(ws) -> int | None:
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        non_empty = [value for value in values if not _is_blank(value)]
        if len(non_empty) >= 2:
            return row_idx
    return None


def _add_error(report: dict[str, Any], check: str, message: str, **extra: Any) -> None:
    _add_issue(report, "errors", check, message, **extra)


def _add_warning(report: dict[str, Any], check: str, message: str, **extra: Any) -> None:
    _add_issue(report, "warnings", check, message, **extra)


def _add_issue(report: dict[str, Any], bucket: str, check: str, message: str, **extra: Any) -> None:
    suggestion = extra.pop("suggestion", None)
    item = {"check": check, "message": message}
    item.update({key: value for key, value in extra.items() if value is not None})
    report[bucket].append(item)
    if suggestion and suggestion not in report["suggestions"]:
        report["suggestions"].append(suggestion)


def _finalize_report(report: dict[str, Any], output_json: str | Path | None) -> None:
    report["summary"]["error_count"] = len(report["errors"])
    report["summary"]["warning_count"] = len(report["warnings"])
    if report["errors"]:
        report["status"] = "fail"
    elif report["warnings"]:
        report["status"] = "warn"
    else:
        report["status"] = "pass"
    if output_json:
        save_json(report, output_json)


def _is_instruction_sheet(name: str) -> bool:
    normalized = _normalized_name(name)
    return (
        normalized in INSTRUCTION_NAMES
        or "说明" in normalized
        or "readme" in normalized
        or "instruction" in normalized
    )


def _is_data_sheet(name: str) -> bool:
    normalized = _normalized_name(name)
    return normalized in DATA_SHEET_NAMES or any(
        keyword in normalized
        for keyword in (
            "data",
            "输入",
            "数据",
            "明细",
            "记录",
            "录入",
            "名册",
            "清单",
            "原始",
        )
    )


def _normalized_name(value: str) -> str:
    return unicodedata.normalize("NFKC", str(value)).strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def _normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value)).strip().lower()
    return re.sub(r"[\s_　\-]+", "", text)


def _sheet_has_content(ws) -> bool:
    return any(not _is_blank(cell.value) for row in ws.iter_rows() for cell in row)


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _matches(text: str, words: Iterable[str]) -> bool:
    return any(word.lower() in text for word in words)


def _duplicates(values: list[str]) -> list[str]:
    seen: set[str] = set()
    duplicates: list[str] = []
    for value in values:
        if value in seen and value not in duplicates:
            duplicates.append(value)
        seen.add(value)
    return duplicates


def _is_blank_or_header_separator(ws, row: int) -> bool:
    cells = [ws.cell(row, col) for col in range(1, ws.max_column + 1)]
    non_empty = [cell for cell in cells if not _is_blank(cell.value)]
    if not non_empty:
        return True
    return all(cell.font and cell.font.bold for cell in non_empty)


def _is_date_format(number_format: str) -> bool:
    fmt = number_format.lower()
    return any(token in fmt for token in ("yy", "mm", "dd", "yyyy", "m/d", "d/m"))


def _date_column_values_are_parseable(cells: list[Any]) -> bool:
    values = [cell.value for cell in cells if not _is_blank(cell.value) and not _is_formula(cell.value)]
    if not values:
        return True
    return _bad_date_ratio(values) < 0.2


def _numeric_column_values_are_parseable(cells: list[Any]) -> bool:
    values = [cell.value for cell in cells if not _is_blank(cell.value) and not _is_formula(cell.value)]
    if not values:
        return True
    return _text_ratio_in_numeric_values(values) < 0.3


def _bad_date_ratio(values: list[Any]) -> float:
    if not values:
        return 0
    bad = 0
    for value in values:
        if isinstance(value, (datetime, date)):
            continue
        try:
            parsed = datetime.fromisoformat(str(value).replace("/", "-"))
            if parsed:
                continue
        except Exception:
            bad += 1
    return bad / len(values)


def _text_ratio_in_numeric_values(values: list[Any]) -> float:
    if not values:
        return 0
    text_count = 0
    for value in values:
        if _to_number(value) is None:
            text_count += 1
    return text_count / len(values)


def _to_number(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "").replace("%", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cell_effective_value(cell, cached_ws) -> Any:
    if _is_formula(cell.value) and cached_ws is not None:
        return cached_ws[cell.coordinate].value
    return cell.value


def _has_division_risk(formula: str) -> bool:
    formula_without_strings = re.sub(r'"(?:[^"]|"")*"', "", formula)
    upper = formula_without_strings.upper()
    if "/" not in formula_without_strings:
        return False
    if "IFERROR" in upper or "IF(" in upper:
        return False
    # Ignore simple divisions by numeric constants, e.g. /12.
    risky_parts = re.findall(r"/\s*([^,+\-*/\)]+)", formula_without_strings)
    for part in risky_parts:
        try:
            return float(part.strip()) == 0
        except ValueError:
            return True
    return False
