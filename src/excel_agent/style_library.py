"""Reusable openpyxl styling helpers."""

from __future__ import annotations

from typing import Iterable

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


FONT_NAME = "Microsoft YaHei"
COLORS = {
    "title": "1F4E78",
    "header": "D9EAF7",
    "input": "FFF2CC",
    "calc": "E2F0D9",
    "summary": "EADCF8",
    "border": "B7B7B7",
    "white": "FFFFFF",
    "text": "1F1F1F",
    "red": "C00000",
}

MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00'
PERCENT_FORMAT = "0.00%"
DATE_FORMAT = "yyyy-mm-dd"
INTEGER_FORMAT = "0"


thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def apply_title_style(ws: Worksheet, row: int = 1, start_col: int = 1, end_col: int | None = None) -> None:
    end_col = end_col or ws.max_column
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill(COLORS["title"])
        cell.font = Font(name=FONT_NAME, bold=True, color=COLORS["white"], size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26


def apply_header_style(ws: Worksheet, row: int) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = _fill(COLORS["header"])
        cell.font = Font(name=FONT_NAME, bold=True, color=COLORS["text"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def apply_input_style(ws: Worksheet, min_row: int, max_row: int, cols: Iterable[int]) -> None:
    for row in range(min_row, max_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(COLORS["input"])
            cell.border = thin_border


def apply_calculation_style(ws: Worksheet, min_row: int, max_row: int, cols: Iterable[int]) -> None:
    for row in range(min_row, max_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(COLORS["calc"])
            cell.border = thin_border


def apply_summary_style(ws: Worksheet, min_row: int, max_row: int, min_col: int = 1, max_col: int = 2) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = _fill(COLORS["summary"])
            cell.border = thin_border
            if col == min_col:
                cell.font = Font(name=FONT_NAME, bold=True)


def apply_number_formats(
    ws: Worksheet,
    money_cols: Iterable[int] = (),
    percent_cols: Iterable[int] = (),
    date_cols: Iterable[int] = (),
    integer_cols: Iterable[int] = (),
    min_row: int = 1,
    max_row: int | None = None,
) -> None:
    max_row = max_row or ws.max_row
    for cols, fmt in [
        (money_cols, MONEY_FORMAT),
        (percent_cols, PERCENT_FORMAT),
        (date_cols, DATE_FORMAT),
        (integer_cols, INTEGER_FORMAT),
    ]:
        for col in cols:
            for row in range(min_row, max_row + 1):
                ws.cell(row=row, column=col).number_format = fmt


def apply_negative_red(ws: Worksheet, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color=COLORS["red"])),
    )


def freeze_and_filter(ws: Worksheet, freeze_cell: str = "A4", filter_range: str | None = None) -> None:
    ws.freeze_panes = freeze_cell
    if filter_range:
        ws.auto_filter.ref = filter_range


def set_reasonable_column_widths(ws: Worksheet, min_width: int = 10, max_width: int = 32) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            longest = max(longest, min(len(value), max_width))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def apply_print_settings(ws: Worksheet) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.oddHeader.center.text = ws.title
    ws.oddFooter.center.text = "Page &P of &N"


def style_instruction_sheet(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 58
    apply_title_style(ws, 1, 1, 2)
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=1).font = Font(name=FONT_NAME, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (1, 2):
            ws.cell(row=row, column=col).border = thin_border
