"""Inspect and apply uploaded workbook templates without reusing sample data by default."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from .io_utils import convert_legacy_xls


TEMPLATE_MODES = {"reference", "flexible", "strict"}


def normalize_template_mode(value: str | None) -> str:
    mode = str(value or "reference").strip().lower()
    return mode if mode in TEMPLATE_MODES else "reference"


def prepare_template_file(path: str | Path, task_dir: str | Path) -> Path:
    source = Path(path)
    if source.suffix.lower() == ".xls":
        return convert_legacy_xls(source, Path(task_dir) / "converted_templates")
    return source.resolve()


def inspect_template(path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    sheets = []
    for ws in workbook.worksheets:
        header_row, headers = find_template_header(ws)
        sheets.append(
            {
                "name": ws.title,
                "visible": ws.sheet_state == "visible",
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_row": header_row,
                "headers": headers,
                "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
                "freeze_panes": str(ws.freeze_panes or ""),
                "auto_filter": ws.auto_filter.ref,
                "chart_count": len(ws._charts),
            }
        )
    return {
        "file_name": Path(path).name,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "sample_data_included": False,
    }


def apply_template_mode(
    generated_path: str | Path,
    template_path: str | Path,
    *,
    mode: str,
    blueprint: dict[str, Any] | None = None,
    use_template_data: bool = False,
) -> Path:
    resolved_mode = normalize_template_mode(mode)
    if resolved_mode == "strict":
        if not blueprint:
            raise ValueError("严格遵守模板时需要结构化工作簿方案。")
        return _build_strict_from_template(
            generated_path,
            template_path,
            blueprint,
            use_template_data=use_template_data,
        )
    return _apply_visual_template(
        generated_path,
        template_path,
        flexible=resolved_mode == "flexible",
    )


def find_template_header(ws) -> tuple[int | None, list[str]]:
    candidates: list[tuple[int, int, list[str]]] = []
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        nonempty = [value for value in values if value not in (None, "")]
        if len(nonempty) < 2:
            continue
        last = max(
            index
            for index, value in enumerate(values, start=1)
            if value not in (None, "")
        )
        headers = [
            str(value).strip() if value not in (None, "") else ""
            for value in values[:last]
        ]
        candidates.append((len(nonempty), row, headers))
    if not candidates:
        return None, []
    _, row, headers = max(candidates, key=lambda item: (item[0], -item[1]))
    return row, headers


def _apply_visual_template(
    generated_path: str | Path,
    template_path: str | Path,
    *,
    flexible: bool,
) -> Path:
    generated = load_workbook(generated_path)
    template = load_workbook(template_path)
    target = _first_visible_sheet(generated)
    source = _first_visible_sheet(template)
    source_header, _ = find_template_header(source)
    target_header, _ = find_template_header(target)

    _copy_row_style(source, 1, target, 1, target.max_column)
    if source_header and target_header:
        if source_header > 1 and target_header > 1:
            _copy_row_style(
                source,
                source_header - 1,
                target,
                target_header - 1,
                target.max_column,
            )
        _copy_row_style(source, source_header, target, target_header, target.max_column)
        sample_row = min(source_header + 1, source.max_row)
        for row in range(target_header + 1, target.max_row + 1):
            _copy_row_style(source, sample_row, target, row, target.max_column)

    for index in range(1, min(source.max_column, target.max_column) + 1):
        source_width = source.column_dimensions[get_column_letter(index)].width
        if source_width:
            target.column_dimensions[get_column_letter(index)].width = source_width

    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.page_setup = copy(source.page_setup)
    target.page_margins = copy(source.page_margins)
    target.sheet_properties.pageSetUpPr = copy(source.sheet_properties.pageSetUpPr)
    if flexible and source.freeze_panes:
        target.freeze_panes = source.freeze_panes
    generated.save(generated_path)
    return Path(generated_path)


def _build_strict_from_template(
    generated_path: str | Path,
    template_path: str | Path,
    blueprint: dict[str, Any],
    *,
    use_template_data: bool,
) -> Path:
    workbook = load_workbook(template_path)
    ws = _first_visible_sheet(workbook)
    header_row, headers = find_template_header(ws)
    if not header_row or not headers:
        raise ValueError("无法识别模板表头，严格模式不能安全写入。")

    columns = list(blueprint.get("columns") or [])
    labels = [str(item.get("label", "")).strip() for item in columns]
    template_labels = [item for item in headers if item]
    if labels != template_labels:
        raise ValueError(
            "严格模板模式要求字段名称和顺序完全一致。"
            f"模板：{'、'.join(template_labels)}；需求：{'、'.join(labels)}。"
        )

    column_map = {item["key"]: index for index, item in enumerate(columns, start=1)}
    data_start = header_row + 1
    existing_end = max(data_start, ws.max_row)
    for row in range(data_start, existing_end + 1):
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None
    write_start = data_start

    records = list(blueprint.get("records") or [])
    style_row = data_start if data_start <= ws.max_row else header_row
    for offset, record in enumerate(records, start=write_start):
        if offset > ws.max_row:
            _copy_row_style(ws, style_row, ws, offset, len(columns))
        for col, column in enumerate(columns, start=1):
            formula = str(column.get("formula") or "")
            value = (
                _render_formula(formula, offset, column_map)
                if formula
                else record.get(column["key"], "")
            )
            ws.cell(offset, col, value)

    final_row = max(write_start + len(records) - 1, header_row)
    if ws.auto_filter.ref:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(columns))}{final_row}"
        )
    workbook.save(generated_path)
    return Path(generated_path)


def _copy_row_style(source, source_row: int, target, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source_cell = source.cell(source_row, min(col, source.max_column))
        target_cell = target.cell(target_row, col)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
    if source.row_dimensions[source_row].height:
        target.row_dimensions[target_row].height = source.row_dimensions[source_row].height


def _render_formula(template: str, row: int, column_map: dict[str, int]) -> str:
    formula = template
    for key, index in column_map.items():
        formula = formula.replace(
            f"{{{key}}}",
            f"{get_column_letter(index)}{row}",
        )
    if "{" in formula or "}" in formula:
        raise ValueError(f"公式模板存在未知列：{template}")
    return formula if formula.startswith("=") else f"={formula}"


def _first_visible_sheet(workbook):
    return next(
        (sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"),
        workbook.worksheets[0],
    )
