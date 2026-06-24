"""Deterministic builders for complex business workbooks with known semantics."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    RadarChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .io_utils import ensure_output_path
from .style_library import apply_print_settings, set_reasonable_column_widths


THIN = Side(style="thin", color="B7C2D0")
HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="B4C6E7")
REGION_ORDER = ["北美", "欧洲", "亚洲", "其他"]
PRODUCT_ORDER = ["电子产品", "家居用品", "服装"]
MONTHS = [f"{month}月" for month in range(1, 7)]
MAJOR_ORDER = ["计算机科学", "软件工程", "信息安全"]
COURSE_ORDER = ["CS101", "CS102", "CS103", "CS104"]


def can_build_student_grade_analysis(
    prompt: str,
    plan: dict[str, Any],
) -> bool:
    """Recognize the 6-sheet student grade analysis workbook task."""

    text = str(prompt or "")
    tables = list(plan.get("inline_tables") or [])
    header_sets = [
        {str(item) for item in table.get("columns", [])}
        for table in tables
        if isinstance(table, dict)
    ]
    return (
        "学生成绩" in text
        and "成绩录入" in text
        and "学期总评" in text
        and "专业汇总" in text
        and "课程统计" in text
        and any({"课程编号", "课程名称", "学分", "课程类别"}.issubset(headers) for headers in header_sets)
        and any({"学号", "姓名", "专业", "班级"}.issubset(headers) for headers in header_sets)
        and len(_extract_student_grade_scores(text)) >= 80
    )


def build_student_grade_analysis_workbook(
    plan: dict[str, Any],
    prompt: str,
    output: str | Path,
) -> Path:
    """Build the requested student grade workbook with formulas and formatting."""

    courses = _extract_grade_courses(plan)
    students = _extract_grade_students(plan)
    score_map = _extract_student_grade_scores(prompt)
    if len(courses) != 4:
        raise ValueError(f"课程参数应为 4 行，实际识别到 {len(courses)} 行。")
    if len(students) != 20:
        raise ValueError(f"学生信息应为 20 行，实际识别到 {len(students)} 行。")
    if len(score_map) != 80:
        raise ValueError(f"成绩明细应为 80 行，实际识别到 {len(score_map)} 行。")

    output_path = ensure_output_path(output, "2026春季学期成绩分析.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    _write_course_parameter_sheet(wb.create_sheet("课程参数"), courses)
    _write_student_info_sheet(wb.create_sheet("学生信息"), students)
    grade_rows = _write_grade_entry_sheet(
        wb.create_sheet("成绩录入"),
        students,
        courses,
        score_map,
    )
    _write_term_summary_sheet(wb.create_sheet("学期总评"), students, grade_rows)
    _write_major_summary_sheet(wb.create_sheet("专业汇总"))
    _write_course_statistics_sheet(wb.create_sheet("课程统计"), courses)
    wb.save(output_path)
    return output_path


def _extract_grade_courses(plan: dict[str, Any]) -> list[dict[str, Any]]:
    table = _find_table(
        [dict(item) for item in plan.get("inline_tables", [])],
        {"课程编号", "课程名称", "学分", "课程类别"},
    )
    records = []
    for record in (table or {}).get("records", []):
        records.append(
            {
                "课程编号": str(record.get("课程编号", "")).strip(),
                "课程名称": str(record.get("课程名称", "")).strip(),
                "学分": int(_number(record.get("学分"))),
                "课程类别": str(record.get("课程类别", "")).strip(),
            }
        )
    return sorted(
        [item for item in records if item["课程编号"]],
        key=lambda item: COURSE_ORDER.index(item["课程编号"]) if item["课程编号"] in COURSE_ORDER else 99,
    )


def _extract_grade_students(plan: dict[str, Any]) -> list[dict[str, Any]]:
    table = _find_table(
        [dict(item) for item in plan.get("inline_tables", [])],
        {"学号", "姓名", "专业", "班级"},
    )
    records = []
    for record in (table or {}).get("records", []):
        records.append(
            {
                "学号": str(int(_number(record.get("学号")))) if _number(record.get("学号")) else str(record.get("学号", "")).strip(),
                "姓名": str(record.get("姓名", "")).strip(),
                "专业": str(record.get("专业", "")).strip(),
                "班级": str(record.get("班级", "")).strip(),
            }
        )
    return sorted(records, key=lambda item: (_major_sort_key(item["专业"]), item["学号"]))


def _extract_student_grade_scores(prompt: str) -> dict[tuple[str, str], tuple[float, float]]:
    text = str(prompt or "")
    line_re = re.compile(
        r"(\d{8})\s*[：:]\s*((?:CS\d{3}\s*\(\s*\d+(?:\.\d+)?\s*,\s*\d+(?:\.\d+)?\s*\)\s*,?\s*)+)"
    )
    pair_re = re.compile(r"(CS\d{3})\s*\(\s*(\d+(?:\.\d+)?)\s*,\s*(\d+(?:\.\d+)?)\s*\)")
    result: dict[tuple[str, str], tuple[float, float]] = {}
    for student_id, pairs_text in line_re.findall(text):
        for course_id, regular, final in pair_re.findall(pairs_text):
            result[(student_id, course_id)] = (float(regular), float(final))
    return result


def _write_course_parameter_sheet(ws, courses: list[dict[str, Any]]) -> None:
    headers = ["课程编号", "课程名称", "学分", "课程类别"]
    _setup_grade_title(ws, "课程学分及类别", len(headers))
    ws["A2"] = "使用说明：本页为课程基准数据；后续工作表通过公式引用课程编号、课程名称和学分。"
    ws.merge_cells("A2:D2")
    ws["A2"].font = Font(name="Microsoft YaHei", italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    _write_header_row(ws, 3, headers)
    for row, record in enumerate(courses, start=4):
        values = [record["课程编号"], record["课程名称"], record["学分"], record["课程类别"]]
        _write_values(ws, row, values, fill=INPUT_FILL)
        ws.cell(row, 3).number_format = "0"
    _finish_grade_sheet(ws, len(headers), 7, freeze="A4", filter_ref="A3:D7")


def _write_student_info_sheet(ws, students: list[dict[str, Any]]) -> None:
    headers = ["学号", "姓名", "专业", "班级"]
    _setup_grade_title(ws, "学生基本信息", len(headers))
    _write_header_row(ws, 3, headers)
    for row, record in enumerate(students, start=4):
        _write_values(
            ws,
            row,
            [record["学号"], record["姓名"], record["专业"], record["班级"]],
            fill=INPUT_FILL,
        )
    _merge_contiguous_labels(ws, 4, 3 + len(students), 3)
    _finish_grade_sheet(ws, len(headers), 3 + len(students), freeze="A4", filter_ref=f"A3:D{3 + len(students)}")


def _write_grade_entry_sheet(
    ws,
    students: list[dict[str, Any]],
    courses: list[dict[str, Any]],
    score_map: dict[tuple[str, str], tuple[float, float]],
) -> dict[str, Any]:
    course_credits = {item["课程编号"]: _number(item["学分"]) for item in courses}
    estimated_gpa: dict[str, float] = {}
    headers = [
        "学号",
        "姓名",
        "专业",
        "课程编号",
        "课程名称",
        "平时成绩",
        "期末成绩",
        "总评成绩",
        "学分",
        "绩点",
        "备注",
    ]
    helper_col = 12
    _setup_grade_title(ws, "课程成绩录入（平时30% + 期末70%）", len(headers))
    _write_header_row(ws, 3, headers)
    _style_header(ws.cell(3, helper_col, "专业_完整"))
    row = 4
    for student in students:
        weighted_gpa = 0.0
        total_credit = 0.0
        for course_id in COURSE_ORDER:
            regular, final = score_map[(student["学号"], course_id)]
            total_score = regular * 0.3 + final * 0.7
            credit = course_credits.get(course_id, 0)
            weighted_gpa += _score_to_gpa(total_score) * credit
            total_credit += credit
            ws.cell(row, 1, student["学号"])
            ws.cell(row, 2, f'=IFERROR(INDEX(\'学生信息\'!$B:$B,MATCH(A{row},\'学生信息\'!$A:$A,0)),"")')
            ws.cell(row, 3, f'=IFERROR(INDEX(\'学生信息\'!$C:$C,MATCH(A{row},\'学生信息\'!$A:$A,0)),"")')
            ws.cell(row, 4, course_id)
            ws.cell(row, 5, f'=IFERROR(INDEX(\'课程参数\'!$B:$B,MATCH(D{row},\'课程参数\'!$A:$A,0)),"")')
            ws.cell(row, 6, regular)
            ws.cell(row, 7, final)
            ws.cell(row, 8, f"=IFERROR(F{row}*30%+G{row}*70%,0)")
            ws.cell(row, 9, f'=IFERROR(INDEX(\'课程参数\'!$C:$C,MATCH(D{row},\'课程参数\'!$A:$A,0)),0)')
            ws.cell(row, 10, _gpa_formula(f"H{row}"))
            ws.cell(row, 11, f'=IF(H{row}<60,"⚠️不及格","")')
            ws.cell(row, helper_col, f'=IFERROR(INDEX(\'学生信息\'!$C:$C,MATCH(A{row},\'学生信息\'!$A:$A,0)),"")')
            for col in range(1, helper_col + 1):
                fill = FORMULA_FILL if col in {2, 3, 5, 8, 9, 10, 11, 12} else INPUT_FILL
                _style_body_cell(ws.cell(row, col), fill)
            row += 1
        estimated_gpa[student["学号"]] = weighted_gpa / total_credit if total_credit else 0
    end_row = row - 1
    _merge_major_groups_by_students(ws, students, 4, rows_per_student=len(COURSE_ORDER), col=3)
    _add_grade_entry_rules(ws, 4, end_row)
    validation = DataValidation(type="list", formula1='"CS101,CS102,CS103,CS104"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(f"D4:D{end_row}")
    ws.column_dimensions[get_column_letter(helper_col)].hidden = True
    for row_index in range(4, end_row + 1):
        for col in (6, 7):
            ws.cell(row_index, col).number_format = "0"
        ws.cell(row_index, 8).number_format = "0.0"
        ws.cell(row_index, 9).number_format = "0"
        ws.cell(row_index, 10).number_format = "0.0"
    _finish_grade_sheet(ws, helper_col, end_row, freeze="A4", filter_ref=f"A3:K{end_row}")
    return {"start": 4, "end": end_row, "estimated_gpa": estimated_gpa}


def _write_term_summary_sheet(ws, students: list[dict[str, Any]], grade_rows: dict[str, Any]) -> None:
    headers = ["学号", "姓名", "专业", "班级", "总修学分", "加权平均分", "平均绩点(GPA)", "专业排名", "等级"]
    helper_col = 10
    sorted_students = sorted(students, key=lambda item: (_major_sort_key(item["专业"]), -_estimated_student_gpa(item["学号"], grade_rows)))
    _setup_grade_title(ws, "2026春季学期总评成绩", len(headers))
    _write_header_row(ws, 3, headers)
    _style_header(ws.cell(3, helper_col, "专业_完整"))
    start_row = 4
    end_row = start_row + len(sorted_students) - 1
    grade_start, grade_end = int(grade_rows["start"]), int(grade_rows["end"])
    for row, student in enumerate(sorted_students, start=start_row):
        ws.cell(row, 1, student["学号"])
        ws.cell(row, 2, f'=IFERROR(INDEX(\'学生信息\'!$B:$B,MATCH(A{row},\'学生信息\'!$A:$A,0)),"")')
        ws.cell(row, 3, f'=IFERROR(INDEX(\'学生信息\'!$C:$C,MATCH(A{row},\'学生信息\'!$A:$A,0)),"")')
        ws.cell(row, 4, f'=IFERROR(INDEX(\'学生信息\'!$D:$D,MATCH(A{row},\'学生信息\'!$A:$A,0)),"")')
        ws.cell(row, 5, f'=SUMIF(\'成绩录入\'!$A$4:$A${grade_end},A{row},\'成绩录入\'!$I$4:$I${grade_end})')
        ws.cell(row, 6, f'=IFERROR(SUMPRODUCT((\'成绩录入\'!$A$4:$A${grade_end}=A{row})*(\'成绩录入\'!$I$4:$I${grade_end})*(\'成绩录入\'!$H$4:$H${grade_end}))/E{row},0)')
        ws.cell(row, 7, f'=IFERROR(SUMPRODUCT((\'成绩录入\'!$A$4:$A${grade_end}=A{row})*(\'成绩录入\'!$I$4:$I${grade_end})*(\'成绩录入\'!$J$4:$J${grade_end}))/E{row},0)')
        ws.cell(row, 10, f'=IFERROR(INDEX(\'学生信息\'!$C:$C,MATCH(A{row},\'学生信息\'!$A:$A,0)),"")')
        for col in range(1, helper_col + 1):
            fill = FORMULA_FILL if col in {2, 3, 4, 5, 6, 7, 8, 9, 10} else INPUT_FILL
            _style_body_cell(ws.cell(row, col), fill)
    for row in range(start_row, end_row + 1):
        ws.cell(row, 8, f'=1+SUMPRODUCT(($J$4:$J${end_row}=J{row})*($G$4:$G${end_row}>G{row}))')
        ws.cell(row, 9, f'=IF(G{row}>=3.7,"优秀",IF(G{row}>=3,"良好",IF(G{row}>=2,"中等",IF(G{row}>=1,"及格","不及格"))))')
        ws.cell(row, 5).number_format = "0"
        ws.cell(row, 6).number_format = "0.00"
        ws.cell(row, 7).number_format = "0.00"
        ws.cell(row, 8).number_format = "0"
    _merge_major_groups_by_students(ws, sorted_students, start_row, rows_per_student=1, col=3)
    _add_term_summary_rules(ws, start_row, end_row)
    ws.column_dimensions[get_column_letter(helper_col)].hidden = True
    _finish_grade_sheet(ws, helper_col, end_row, freeze="A4", filter_ref=f"A3:I{end_row}")


def _write_major_summary_sheet(ws) -> None:
    headers = ["专业", "人数", "总修学分", "加权平均分均值", "GPA均值", "最高GPA", "最低GPA", "不及格人次"]
    _setup_grade_title(ws, "各专业成绩汇总统计", len(headers))
    _write_header_row(ws, 3, headers)
    start_row = 4
    for row, major in enumerate(MAJOR_ORDER, start=start_row):
        ws.cell(row, 1, major)
        ws.cell(row, 2, f'=COUNTIF(\'学期总评\'!$J:$J,A{row})')
        ws.cell(row, 3, f'=SUMIF(\'学期总评\'!$J:$J,A{row},\'学期总评\'!$E:$E)')
        ws.cell(row, 4, f'=IFERROR(AVERAGEIF(\'学期总评\'!$J:$J,A{row},\'学期总评\'!$F:$F),0)')
        ws.cell(row, 5, f'=IFERROR(AVERAGEIF(\'学期总评\'!$J:$J,A{row},\'学期总评\'!$G:$G),0)')
        ws.cell(row, 6, f'=IFERROR(MAXIFS(\'学期总评\'!$G:$G,\'学期总评\'!$J:$J,A{row}),0)')
        ws.cell(row, 7, f'=IFERROR(MINIFS(\'学期总评\'!$G:$G,\'学期总评\'!$J:$J,A{row}),0)')
        ws.cell(row, 8, f'=COUNTIFS(\'成绩录入\'!$L:$L,A{row},\'成绩录入\'!$H:$H,"<60")')
        for col in range(1, len(headers) + 1):
            _style_body_cell(ws.cell(row, col), FORMULA_FILL if col > 1 else INPUT_FILL)
    total_row = start_row + len(MAJOR_ORDER)
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 2, f"=SUM(B{start_row}:B{total_row - 1})")
    ws.cell(total_row, 3, f"=SUM(C{start_row}:C{total_row - 1})")
    ws.cell(total_row, 4, "=IFERROR(AVERAGE('学期总评'!F4:F23),0)")
    ws.cell(total_row, 5, "=IFERROR(AVERAGE('学期总评'!G4:G23),0)")
    ws.cell(total_row, 6, "=MAX('学期总评'!G4:G23)")
    ws.cell(total_row, 7, "=MIN('学期总评'!G4:G23)")
    ws.cell(total_row, 8, f"=SUM(H{start_row}:H{total_row - 1})")
    _style_sales_total_row(ws, total_row, len(headers))
    for row in range(start_row, total_row + 1):
        for col in (4, 5, 6, 7):
            ws.cell(row, col).number_format = "0.00"
    ws.conditional_formatting.add(
        f"H{start_row}:H{total_row - 1}",
        FormulaRule(formula=[f"H{start_row}>3"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    _finish_grade_sheet(ws, len(headers), total_row, freeze="A4", filter_ref=f"A3:H{total_row}")


def _write_course_statistics_sheet(ws, courses: list[dict[str, Any]]) -> None:
    _setup_grade_title(ws, "各课程成绩分布统计", 7)
    ws.merge_cells("A3:A4")
    ws.merge_cells("B3:B4")
    ws.merge_cells("C3:G3")
    for cell, value in (("A3", "课程编号"), ("B3", "课程名称"), ("C3", "成绩统计")):
        ws[cell] = value
        _style_header(ws[cell])
    for col, value in enumerate(["平均分", "最高分", "最低分", "及格率（≥60）", "优秀率（≥90）"], start=3):
        ws.cell(4, col, value)
        _style_header(ws.cell(4, col))
    start_row = 5
    for row, course in enumerate(courses, start=start_row):
        course_id = course["课程编号"]
        ws.cell(row, 1, course_id)
        ws.cell(row, 2, f'=IFERROR(INDEX(\'课程参数\'!$B:$B,MATCH(A{row},\'课程参数\'!$A:$A,0)),"")')
        ws.cell(row, 3, f'=IFERROR(AVERAGEIF(\'成绩录入\'!$D:$D,A{row},\'成绩录入\'!$H:$H),0)')
        ws.cell(row, 4, f'=IFERROR(MAXIFS(\'成绩录入\'!$H:$H,\'成绩录入\'!$D:$D,A{row}),0)')
        ws.cell(row, 5, f'=IFERROR(MINIFS(\'成绩录入\'!$H:$H,\'成绩录入\'!$D:$D,A{row}),0)')
        ws.cell(row, 6, f'=IFERROR(COUNTIFS(\'成绩录入\'!$D:$D,A{row},\'成绩录入\'!$H:$H,">=60")/COUNTIF(\'成绩录入\'!$D:$D,A{row}),0)')
        ws.cell(row, 7, f'=IFERROR(COUNTIFS(\'成绩录入\'!$D:$D,A{row},\'成绩录入\'!$H:$H,">=90")/COUNTIF(\'成绩录入\'!$D:$D,A{row}),0)')
        for col in range(1, 8):
            _style_body_cell(ws.cell(row, col), FORMULA_FILL if col > 1 else INPUT_FILL)
    total_row = start_row + len(courses)
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 2, "")
    ws.cell(total_row, 3, "=IFERROR(AVERAGE('成绩录入'!H4:H83),0)")
    ws.cell(total_row, 4, "=MAX('成绩录入'!H4:H83)")
    ws.cell(total_row, 5, "=MIN('成绩录入'!H4:H83)")
    ws.cell(total_row, 6, '=IFERROR(COUNTIF(\'成绩录入\'!H4:H83,">=60")/COUNT(\'成绩录入\'!H4:H83),0)')
    ws.cell(total_row, 7, '=IFERROR(COUNTIF(\'成绩录入\'!H4:H83,">=90")/COUNT(\'成绩录入\'!H4:H83),0)')
    _style_sales_total_row(ws, total_row, 7)
    for row in range(start_row, total_row + 1):
        for col in (3, 4, 5):
            ws.cell(row, col).number_format = "0.0"
        for col in (6, 7):
            ws.cell(row, col).number_format = "0.0%"
    ws.conditional_formatting.add(
        f"F{start_row}:F{total_row - 1}",
        FormulaRule(formula=[f"F{start_row}<0.7"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"G{start_row}:G{total_row - 1}",
        FormulaRule(formula=[f"G{start_row}>0.3"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    _finish_grade_sheet(ws, 7, total_row, freeze="A5", filter_ref=f"A4:G{total_row}")


def _setup_grade_title(ws, title: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.sheet_view.showGridLines = False


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        ws.cell(row, col, header)
        _style_header(ws.cell(row, col))


def _write_values(ws, row: int, values: list[Any], fill=INPUT_FILL) -> None:
    for col, value in enumerate(values, start=1):
        ws.cell(row, col, value)
        _style_body_cell(ws.cell(row, col), fill)


def _finish_grade_sheet(ws, max_col: int, max_row: int, *, freeze: str, filter_ref: str | None) -> None:
    widths = {
        "A": 14,
        "B": 14,
        "C": 16,
        "D": 14,
        "E": 18,
        "F": 14,
        "G": 14,
        "H": 16,
        "I": 16,
        "J": 12,
        "K": 16,
        "L": 14,
    }
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 14)
    ws.freeze_panes = freeze
    if filter_ref:
        ws.auto_filter.ref = filter_ref
    ws.print_area = f"A1:{get_column_letter(min(max_col, 11))}{max_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    apply_print_settings(ws)


def _merge_contiguous_labels(ws, start_row: int, end_row: int, col: int) -> None:
    group_start = start_row
    previous = ws.cell(start_row, col).value
    for row in range(start_row + 1, end_row + 2):
        value = ws.cell(row, col).value if row <= end_row else object()
        if value != previous:
            if row - group_start > 1:
                ws.merge_cells(start_row=group_start, start_column=col, end_row=row - 1, end_column=col)
                ws.cell(group_start, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            group_start = row
            previous = value


def _merge_major_groups_by_students(
    ws,
    students: list[dict[str, Any]],
    start_row: int,
    *,
    rows_per_student: int,
    col: int,
) -> None:
    if not students:
        return
    group_start = start_row
    current_major = students[0]["专业"]
    for index, student in enumerate([*students, {"专业": object()}]):
        if index == 0:
            continue
        row = start_row + index * rows_per_student
        if student["专业"] != current_major:
            group_end = row - 1
            if group_end > group_start:
                ws.merge_cells(start_row=group_start, start_column=col, end_row=group_end, end_column=col)
                ws.cell(group_start, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            group_start = row
            current_major = student["专业"]


def _add_grade_entry_rules(ws, start_row: int, end_row: int) -> None:
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        FormulaRule(formula=[f"H{start_row}>=90"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        FormulaRule(
            formula=[f"H{start_row}<60"],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(color="C00000", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        f"J{start_row}:J{end_row}",
        FormulaRule(
            formula=[f"J{start_row}=0"],
            fill=PatternFill("solid", fgColor="FFC7CE"),
            font=Font(bold=True),
        ),
    )
    ws.conditional_formatting.add(
        f"K{start_row}:K{end_row}",
        FormulaRule(
            formula=[f'K{start_row}<>""'],
            font=Font(color="C00000", bold=True),
        ),
    )


def _add_term_summary_rules(ws, start_row: int, end_row: int) -> None:
    ws.conditional_formatting.add(
        f"I{start_row}:I{end_row}",
        FormulaRule(formula=[f'I{start_row}="优秀"'], fill=PatternFill("solid", fgColor="FFD700"), font=Font(bold=True)),
    )
    ws.conditional_formatting.add(
        f"I{start_row}:I{end_row}",
        FormulaRule(formula=[f'I{start_row}="不及格"'], fill=PatternFill("solid", fgColor="FF0000"), font=Font(color="FFFFFF", bold=True)),
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        FormulaRule(formula=[f"F{start_row}>=90"], font=Font(color="008000", bold=True)),
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        FormulaRule(formula=[f"F{start_row}<60"], font=Font(color="C00000", bold=True)),
    )


def _gpa_formula(score_ref: str) -> str:
    return (
        f"=IFERROR(LOOKUP({score_ref},"
        "{0,60,64,68,72,75,78,82,85,90},"
        "{0,1,1.5,2,2.3,2.7,3,3.3,3.7,4}),0)"
    )


def _major_sort_key(major: str) -> int:
    return MAJOR_ORDER.index(major) if major in MAJOR_ORDER else len(MAJOR_ORDER)


def _estimated_student_gpa(student_id: str, grade_rows: dict[str, Any]) -> float:
    values = grade_rows.get("estimated_gpa", {}) if isinstance(grade_rows, dict) else {}
    try:
        return float(values.get(student_id, 0))
    except (TypeError, ValueError):
        return 0.0


def _score_to_gpa(score: float) -> float:
    if score >= 90:
        return 4.0
    if score >= 85:
        return 3.7
    if score >= 82:
        return 3.3
    if score >= 78:
        return 3.0
    if score >= 75:
        return 2.7
    if score >= 72:
        return 2.3
    if score >= 68:
        return 2.0
    if score >= 64:
        return 1.5
    if score >= 60:
        return 1.0
    return 0.0


def can_build_performance_compensation(
    prompt: str,
    plan: dict[str, Any],
) -> bool:
    text = str(prompt or "")
    tables = list(plan.get("inline_tables") or [])
    required = ("绩效评估", "薪酬调整")
    header_sets = [
        {str(item) for item in table.get("columns", [])}
        for table in tables
        if isinstance(table, dict)
    ]
    return (
        all(word in text for word in required)
        and any({"指标", "权重"}.issubset(headers) for headers in header_sets)
        and any({"等级", "最低分", "最高分"}.issubset(headers) for headers in header_sets)
        and any("员工编号" in headers and "当前月薪(元)" in headers for headers in header_sets)
    )


def build_performance_compensation_workbook(
    plan: dict[str, Any],
    prompt: str,
    output: str | Path,
) -> Path:
    """Build a parameter-driven performance and salary adjustment workbook."""

    tables = [dict(item) for item in plan.get("inline_tables", [])]
    weights = _find_table(tables, {"指标", "权重"})
    grades = _find_table(tables, {"等级", "最低分", "最高分"})
    adjustments = _find_table(tables, {"等级", "调整比例"})
    deductions = _find_table(tables, {"扣分值"})
    employees = max(tables, key=lambda item: len(item.get("columns", [])))
    if not all((weights, grades, adjustments, deductions, employees)):
        raise ValueError("绩效薪酬模型缺少必要参数表或员工基础数据。")

    output_path = ensure_output_path(output, "员工绩效评估及薪酬调整表.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    _write_instructions(wb.create_sheet("说明"), plan)
    parameter_refs = _write_parameter_sheet(
        wb.create_sheet("参数表"),
        weights,
        grades,
        adjustments,
        deductions,
    )
    _write_performance_detail(
        wb.create_sheet("明细表"),
        employees,
        parameter_refs,
        str(plan.get("title") or "员工绩效评估及薪酬调整明细表"),
        prompt,
    )
    wb.save(output_path)
    return output_path


def can_build_global_sales_analysis(prompt: str, plan: dict[str, Any]) -> bool:
    """Recognize the 5-sheet global sales analysis workbook task."""

    text = str(prompt or "")
    tables = list(plan.get("inline_tables") or [])
    has_target_table = any(
        {"地区", "产品", "月度销售目标(元)"}.issubset(
            {str(item) for item in table.get("columns", [])}
        )
        for table in tables
        if isinstance(table, dict)
    )
    return (
        has_target_table
        and "销售明细" in text
        and "地区汇总" in text
        and "产品汇总" in text
        and "交叉汇总" in text
        and "基础数据" in text
    )


def build_global_sales_analysis_workbook(
    plan: dict[str, Any],
    prompt: str,
    output: str | Path,
    *,
    include_charts: bool = False,
    chart_types: list[str] | None = None,
) -> Path:
    """Build the requested 5-sheet sales workbook with formulas and styling.

    This intentionally bypasses generic inline-table preservation: the prompt
    contains enough business semantics to compile a real workbook with formulas,
    summaries and a pivot-style cross table.
    """

    targets = _extract_sales_targets(plan, prompt)
    details = _extract_sales_detail_records(prompt)
    if len(targets) != 12:
        raise ValueError(f"销售目标参数应为 12 行，实际识别到 {len(targets)} 行。")
    if len(details) != 72:
        raise ValueError(f"销售明细应为 72 行，实际识别到 {len(details)} 行。")

    output_path = ensure_output_path(output, "2026上半年全球销售分析.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    _write_sales_target_sheet(wb.create_sheet("参数表"), targets)
    detail_bounds = _write_sales_detail_sheet(wb.create_sheet("明细"), details)
    _write_sales_region_summary(wb.create_sheet("地区汇总"), details, detail_bounds)
    _write_sales_product_summary(wb.create_sheet("产品汇总"), details, detail_bounds)
    cross_ws = wb.create_sheet("交叉汇总")
    _write_sales_cross_summary(cross_ws, detail_bounds)
    if include_charts:
        _add_global_sales_charts(cross_ws, details, chart_types or ["column"])
    wb.save(output_path)
    return output_path


def _extract_sales_targets(plan: dict[str, Any], prompt: str) -> list[dict[str, Any]]:
    tables = [dict(item) for item in plan.get("inline_tables", [])]
    target_table = _find_table(tables, {"地区", "产品", "月度销售目标(元)"})
    if target_table:
        records = []
        for record in target_table.get("records", []):
            records.append(
                {
                    "地区": str(record.get("地区", "")).strip(),
                    "产品": str(record.get("产品", "")).strip(),
                    "月度销售目标(元)": _number(record.get("月度销售目标(元)")),
                }
            )
        if records:
            return sorted(
                records,
                key=lambda item: (
                    REGION_ORDER.index(item["地区"]) if item["地区"] in REGION_ORDER else 99,
                    PRODUCT_ORDER.index(item["产品"]) if item["产品"] in PRODUCT_ORDER else 99,
                ),
            )

    records: list[dict[str, Any]] = []
    pattern = re.compile(
        r"^\s*(北美|欧洲|亚洲|其他)\s+"
        r"(电子产品|家居用品|服装)\s+([\d,]+)\s*$",
        re.MULTILINE,
    )
    for region, product, target in pattern.findall(str(prompt or "")):
        records.append(
            {
                "地区": region,
                "产品": product,
                "月度销售目标(元)": _number(target),
            }
        )
    return records


def _extract_sales_detail_records(prompt: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    text = str(prompt or "")
    line_re = re.compile(
        r"(北美|欧洲|亚洲|其他)\s*-\s*(电子产品|家居用品|服装)\s*[：:]"
        r"\s*((?:\(\s*[\d,]+\s*,\s*[\d,]+\s*\)\s*)+)"
    )
    pair_re = re.compile(r"\(\s*([\d,]+)\s*,\s*([\d,]+)\s*\)")
    matched: dict[tuple[str, str], list[tuple[float, float]]] = {}
    for region, product, pairs_text in line_re.findall(text):
        pairs = [(_number(sales), _number(profit)) for sales, profit in pair_re.findall(pairs_text)]
        matched[(region, product)] = pairs[:6]

    for region in REGION_ORDER:
        for product in PRODUCT_ORDER:
            pairs = matched.get((region, product), [])
            for month, values in zip(MONTHS, pairs):
                sales, profit = values
                records.append(
                    {
                        "地区": region,
                        "产品": product,
                        "月份": month,
                        "销售额(元)": sales,
                        "利润(元)": profit,
                    }
                )
    return records


def _write_sales_target_sheet(ws, targets: list[dict[str, Any]]) -> None:
    headers = ["地区", "产品", "月度销售目标(元)"]
    _setup_title(ws, "各区域产品月度销售目标（元）", 1, len(headers))
    ws["A2"] = "使用说明：参数表为输入基准；明细、汇总和交叉汇总通过 Excel 公式联动。"
    ws.merge_cells("A2:C2")
    ws["A2"].font = Font(name="Microsoft YaHei", italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(3, col, header))
    _style_header(ws.cell(3, 4, "匹配键"))
    for row, record in enumerate(targets, start=4):
        ws.cell(row, 1, record["地区"])
        ws.cell(row, 2, record["产品"])
        ws.cell(row, 3, record["月度销售目标(元)"])
        ws.cell(row, 4, f"=A{row}&B{row}")
        for col in range(1, 5):
            _style_body_cell(ws.cell(row, col), INPUT_FILL if col < 4 else FORMULA_FILL)
        ws.cell(row, 3).number_format = "#,##0"
    ws.column_dimensions["D"].hidden = True
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:C{3 + len(targets)}"
    _finish_sales_sheet(ws, 4)


def _write_sales_detail_sheet(
    ws,
    details: list[dict[str, Any]],
) -> dict[str, int]:
    headers = [
        "地区",
        "产品",
        "月份",
        "销售额(元)",
        "利润(元)",
        "利润率",
        "目标销售额(元)",
        "达成率",
    ]
    _setup_title(ws, "2026年1-6月销售明细", 1, len(headers))
    ws["A2"] = "数据验证：地区列可选“北美,欧洲,亚洲,其他”；产品列可选“电子产品,家居用品,服装”。"
    ws.merge_cells("A2:H2")
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(3, col, header))
    for col, header in enumerate(["地区_完整", "产品_完整", "匹配键"], start=9):
        _style_header(ws.cell(3, col, header))

    start_row = 4
    for row, record in enumerate(details, start=start_row):
        ws.cell(row, 1, record["地区"])
        ws.cell(row, 2, record["产品"])
        ws.cell(row, 3, record["月份"])
        ws.cell(row, 4, record["销售额(元)"])
        ws.cell(row, 5, record["利润(元)"])
        ws.cell(row, 6, f"=IFERROR(E{row}/D{row},0)")
        ws.cell(row, 9, record["地区"])
        ws.cell(row, 10, record["产品"])
        ws.cell(row, 11, f"=I{row}&J{row}")
        ws.cell(
            row,
            7,
            f'=IFERROR(INDEX(\'参数表\'!$C:$C,MATCH(K{row},\'参数表\'!$D:$D,0)),0)',
        )
        ws.cell(row, 8, f"=IFERROR(D{row}/G{row}-1,0)")
        for col in range(1, 12):
            fill = FORMULA_FILL if col in {6, 7, 8, 11} else INPUT_FILL
            _style_body_cell(ws.cell(row, col), fill)
        for col in (4, 5, 7):
            ws.cell(row, col).number_format = "#,##0"
        for col in (6, 8):
            ws.cell(row, col).number_format = "0.0%"

    end_row = start_row + len(details) - 1
    _merge_sales_detail_labels(ws, details, start_row)
    _add_sales_validations(ws, start_row, end_row)
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        FormulaRule(formula=[f"F{start_row}<0.05"], font=Font(color="C00000", bold=True)),
    )
    ws.conditional_formatting.add(
        f"F{start_row}:F{end_row}",
        FormulaRule(formula=[f"F{start_row}>0.2"], font=Font(color="008000", bold=True)),
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        FormulaRule(formula=[f"H{start_row}<0"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"H{start_row}:H{end_row}",
        FormulaRule(formula=[f"H{start_row}>0.1"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    for col in ("I", "J", "K"):
        ws.column_dimensions[col].hidden = True
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{end_row}"
    _finish_sales_sheet(ws, 11)
    return {"start": start_row, "end": end_row}


def _write_sales_region_summary(
    ws,
    details: list[dict[str, Any]],
    bounds: dict[str, int],
) -> None:
    del bounds
    headers = ["地区", "总销售额(元)", "总利润(元)", "平均利润率", "总销售额占比", "平均达成率"]
    _setup_title(ws, "按地区汇总（上半年）", 1, len(headers))
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(3, col, header))
    regions = sorted(REGION_ORDER, key=lambda item: -_sum_details(details, "地区", item, "销售额(元)"))
    start_row = 4
    for row, region in enumerate(regions, start=start_row):
        ws.cell(row, 1, region)
        ws.cell(row, 2, f'=SUMIF(\'明细\'!$I:$I,A{row},\'明细\'!$D:$D)')
        ws.cell(row, 3, f'=SUMIF(\'明细\'!$I:$I,A{row},\'明细\'!$E:$E)')
        ws.cell(row, 4, f"=IFERROR(C{row}/B{row},0)")
        ws.cell(row, 5, f"=IFERROR(B{row}/$B${start_row + len(regions)},0)")
        ws.cell(row, 6, f'=AVERAGEIF(\'明细\'!$I:$I,A{row},\'明细\'!$H:$H)')
        _style_summary_body_row(ws, row, len(headers))
    total_row = start_row + len(regions)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=1)
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 2, f"=SUM(B{start_row}:B{total_row - 1})")
    ws.cell(total_row, 3, f"=SUM(C{start_row}:C{total_row - 1})")
    ws.cell(total_row, 4, f"=IFERROR(C{total_row}/B{total_row},0)")
    ws.cell(total_row, 5, 1)
    ws.cell(total_row, 6, f"=AVERAGE(F{start_row}:F{total_row - 1})")
    _style_sales_total_row(ws, total_row, len(headers))
    ws.conditional_formatting.add(
        f"E{start_row}:E{total_row - 1}",
        FormulaRule(formula=[f"E{start_row}>0.3"], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    _finish_summary_sheet(ws, total_row, len(headers))


def _write_sales_product_summary(
    ws,
    details: list[dict[str, Any]],
    bounds: dict[str, int],
) -> None:
    del bounds
    headers = ["产品", "总销售额(元)", "总利润(元)", "平均利润率", "总销售额占比", "平均达成率"]
    _setup_title(ws, "按产品汇总（上半年）", 1, len(headers))
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(3, col, header))
    products = sorted(PRODUCT_ORDER, key=lambda item: -_sum_details(details, "产品", item, "销售额(元)"))
    start_row = 4
    for row, product in enumerate(products, start=start_row):
        ws.cell(row, 1, product)
        ws.cell(row, 2, f'=SUMIF(\'明细\'!$J:$J,A{row},\'明细\'!$D:$D)')
        ws.cell(row, 3, f'=SUMIF(\'明细\'!$J:$J,A{row},\'明细\'!$E:$E)')
        ws.cell(row, 4, f"=IFERROR(C{row}/B{row},0)")
        ws.cell(row, 5, f"=IFERROR(B{row}/$B${start_row + len(products)},0)")
        ws.cell(row, 6, f'=AVERAGEIF(\'明细\'!$J:$J,A{row},\'明细\'!$H:$H)')
        _style_summary_body_row(ws, row, len(headers))
    total_row = start_row + len(products)
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 2, f"=SUM(B{start_row}:B{total_row - 1})")
    ws.cell(total_row, 3, f"=SUM(C{start_row}:C{total_row - 1})")
    ws.cell(total_row, 4, f"=IFERROR(C{total_row}/B{total_row},0)")
    ws.cell(total_row, 5, 1)
    ws.cell(total_row, 6, f"=AVERAGE(F{start_row}:F{total_row - 1})")
    _style_sales_total_row(ws, total_row, len(headers))
    _finish_summary_sheet(ws, total_row, len(headers))


def _write_sales_cross_summary(ws, bounds: dict[str, int]) -> None:
    del bounds
    _setup_title(ws, "交叉汇总表（模拟数据透视表）", 1, 7)
    ws.merge_cells("A3:A4")
    _style_header(ws["A3"])
    ws["A3"] = "地区 / 产品"
    for idx, product in enumerate(PRODUCT_ORDER):
        start_col = 2 + idx * 2
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + 1)
        _style_header(ws.cell(3, start_col, product))
        _style_header(ws.cell(4, start_col, "总销售额"))
        _style_header(ws.cell(4, start_col + 1, "总利润"))
    start_row = 5
    for row, region in enumerate([*REGION_ORDER, "合计"], start=start_row):
        ws.cell(row, 1, region)
        for idx, product in enumerate(PRODUCT_ORDER):
            sales_col = 2 + idx * 2
            profit_col = sales_col + 1
            if region == "合计":
                ws.cell(row, sales_col, f"=SUM({get_column_letter(sales_col)}{start_row}:{get_column_letter(sales_col)}{row - 1})")
                ws.cell(row, profit_col, f"=SUM({get_column_letter(profit_col)}{start_row}:{get_column_letter(profit_col)}{row - 1})")
            else:
                ws.cell(
                    row,
                    sales_col,
                    f'=SUMIFS(\'明细\'!$D:$D,\'明细\'!$I:$I,$A{row},\'明细\'!$J:$J,{get_column_letter(sales_col)}$3)',
                )
                ws.cell(
                    row,
                    profit_col,
                    f'=SUMIFS(\'明细\'!$E:$E,\'明细\'!$I:$I,$A{row},\'明细\'!$J:$J,{get_column_letter(sales_col)}$3)',
                )
            ws.cell(row, sales_col).number_format = "#,##0"
            ws.cell(row, profit_col).number_format = "#,##0"
        fill = TOTAL_FILL if region == "合计" else INPUT_FILL
        for col in range(1, 8):
            _style_body_cell(ws.cell(row, col), fill)
            if region == "合计":
                ws.cell(row, col).font = Font(name="Microsoft YaHei", bold=True)
    ws.freeze_panes = "B5"
    _finish_sales_sheet(ws, 7)


def _add_global_sales_charts(
    ws,
    details: list[dict[str, Any]],
    chart_types: list[str],
) -> None:
    """Add stable, non-empty charts based on helper values.

    The helper data is intentionally written as static values in the same sheet:
    WPS and web Office sometimes show blank charts when the chart only points to
    formulas that have not been recalculated yet.
    """

    normalized_types = [
        item
        for item in dict.fromkeys(str(chart_type or "column").lower() for chart_type in chart_types)
        if item
    ] or ["column"]
    helper = _write_global_sales_chart_sources(ws, details)
    anchors = ["N3", "N20", "N37"]
    for index, chart_type in enumerate(normalized_types[:3]):
        anchor = anchors[index]
        if chart_type in {"line", "area"}:
            chart = _build_monthly_sales_chart(ws, helper, chart_type)
        elif chart_type in {"pie", "doughnut"}:
            chart = _build_region_share_chart(ws, helper, chart_type)
        elif chart_type == "scatter":
            chart = _build_sales_profit_scatter(ws, helper)
        elif chart_type == "radar":
            chart = _build_region_radar_chart(ws, helper)
        elif chart_type == "combo":
            chart = _build_sales_combo_chart(ws, helper)
        else:
            chart = _build_region_comparison_chart(ws, helper, chart_type)
        ws.add_chart(chart, anchor)


def _write_global_sales_chart_sources(
    ws,
    details: list[dict[str, Any]],
) -> dict[str, int]:
    start_col = 8
    region_header_row = 3
    region_start = region_header_row + 1
    ws.cell(2, start_col, "图表数据源（稳定数值，避免未重算导致空图）")
    ws.merge_cells(
        start_row=2,
        start_column=start_col,
        end_row=2,
        end_column=start_col + 4,
    )
    _style_section_title(ws.cell(2, start_col))
    for offset, header in enumerate(["地区", "总销售额(元)", "总利润(元)", "销售占比", "平均利润率"]):
        _style_header(ws.cell(region_header_row, start_col + offset, header))
    total_sales = sum(_number(item.get("销售额(元)")) for item in details) or 1
    for row_offset, region in enumerate(REGION_ORDER):
        row = region_start + row_offset
        sales = _sum_details(details, "地区", region, "销售额(元)")
        profit = _sum_details(details, "地区", region, "利润(元)")
        values = [
            region,
            sales,
            profit,
            sales / total_sales if total_sales else 0,
            profit / sales if sales else 0,
        ]
        for offset, value in enumerate(values):
            cell = ws.cell(row, start_col + offset, value)
            _style_body_cell(cell, FORMULA_FILL if offset else INPUT_FILL)
        for col in (start_col + 1, start_col + 2):
            ws.cell(row, col).number_format = "#,##0"
        for col in (start_col + 3, start_col + 4):
            ws.cell(row, col).number_format = "0.0%"

    month_header_row = region_start + len(REGION_ORDER) + 3
    month_start = month_header_row + 1
    ws.cell(month_header_row - 1, start_col, "月度趋势数据源")
    ws.merge_cells(
        start_row=month_header_row - 1,
        start_column=start_col,
        end_row=month_header_row - 1,
        end_column=start_col + 2,
    )
    _style_section_title(ws.cell(month_header_row - 1, start_col))
    for offset, header in enumerate(["月份", "销售额(元)", "利润(元)"]):
        _style_header(ws.cell(month_header_row, start_col + offset, header))
    for row_offset, month in enumerate(MONTHS):
        row = month_start + row_offset
        sales = sum(
            _number(item.get("销售额(元)")) for item in details if item.get("月份") == month
        )
        profit = sum(
            _number(item.get("利润(元)")) for item in details if item.get("月份") == month
        )
        for offset, value in enumerate([month, sales, profit]):
            cell = ws.cell(row, start_col + offset, value)
            _style_body_cell(cell, FORMULA_FILL if offset else INPUT_FILL)
        ws.cell(row, start_col + 1).number_format = "#,##0"
        ws.cell(row, start_col + 2).number_format = "#,##0"

    for col in range(start_col, start_col + 5):
        ws.column_dimensions[get_column_letter(col)].width = 16
    return {
        "region_header_row": region_header_row,
        "region_start": region_start,
        "region_end": region_start + len(REGION_ORDER) - 1,
        "month_header_row": month_header_row,
        "month_start": month_start,
        "month_end": month_start + len(MONTHS) - 1,
        "start_col": start_col,
    }


def _build_region_comparison_chart(ws, helper: dict[str, int], chart_type: str):
    start_col = helper["start_col"]
    chart = BarChart()
    chart.type = "bar" if chart_type == "bar" else "col"
    chart.grouping = "clustered"
    chart.overlap = -10
    chart.gapWidth = 80
    chart.style = 10
    chart.title = "各地区销售额与利润对比"
    chart.height = 8
    chart.width = 16
    chart.add_data(
        Reference(
            ws,
            min_col=start_col + 1,
            max_col=start_col + 2,
            min_row=helper["region_header_row"],
            max_row=helper["region_end"],
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=start_col,
            min_row=helper["region_start"],
            max_row=helper["region_end"],
        )
    )
    chart.y_axis.title = "金额（元）"
    chart.x_axis.title = "地区"
    chart.y_axis.numFmt = "#,##0"
    return chart


def _build_monthly_sales_chart(ws, helper: dict[str, int], chart_type: str):
    start_col = helper["start_col"]
    chart = AreaChart() if chart_type == "area" else LineChart()
    chart.title = "月度销售额与利润趋势"
    chart.style = 13
    chart.height = 8
    chart.width = 16
    chart.add_data(
        Reference(
            ws,
            min_col=start_col + 1,
            max_col=start_col + 2,
            min_row=helper["month_header_row"],
            max_row=helper["month_end"],
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=start_col,
            min_row=helper["month_start"],
            max_row=helper["month_end"],
        )
    )
    chart.y_axis.title = "金额（元）"
    chart.x_axis.title = "月份"
    chart.y_axis.numFmt = "#,##0"
    return chart


def _build_region_share_chart(ws, helper: dict[str, int], chart_type: str):
    start_col = helper["start_col"]
    chart = DoughnutChart() if chart_type == "doughnut" else PieChart()
    chart.title = "各地区销售额占比"
    chart.height = 8
    chart.width = 12
    chart.add_data(
        Reference(
            ws,
            min_col=start_col + 3,
            min_row=helper["region_header_row"],
            max_row=helper["region_end"],
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=start_col,
            min_row=helper["region_start"],
            max_row=helper["region_end"],
        )
    )
    return chart


def _build_sales_profit_scatter(ws, helper: dict[str, int]):
    start_col = helper["start_col"]
    chart = ScatterChart()
    chart.title = "销售额与利润相关性"
    chart.height = 8
    chart.width = 14
    chart.x_axis.title = "销售额（元）"
    chart.y_axis.title = "利润（元）"
    xvalues = Reference(
        ws,
        min_col=start_col + 1,
        min_row=helper["region_start"],
        max_row=helper["region_end"],
    )
    yvalues = Reference(
        ws,
        min_col=start_col + 2,
        min_row=helper["region_start"],
        max_row=helper["region_end"],
    )
    chart.series.append(Series(yvalues, xvalues, title="地区"))
    return chart


def _build_region_radar_chart(ws, helper: dict[str, int]):
    start_col = helper["start_col"]
    chart = RadarChart()
    chart.title = "地区销售与利润雷达图"
    chart.height = 8
    chart.width = 14
    chart.add_data(
        Reference(
            ws,
            min_col=start_col + 1,
            max_col=start_col + 2,
            min_row=helper["region_header_row"],
            max_row=helper["region_end"],
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=start_col,
            min_row=helper["region_start"],
            max_row=helper["region_end"],
        )
    )
    return chart


def _build_sales_combo_chart(ws, helper: dict[str, int]):
    start_col = helper["start_col"]
    chart = BarChart()
    chart.type = "col"
    chart.title = "销售额柱线组合图"
    chart.style = 10
    chart.height = 8
    chart.width = 16
    chart.add_data(
        Reference(
            ws,
            min_col=start_col + 1,
            min_row=helper["month_header_row"],
            max_row=helper["month_end"],
        ),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(
            ws,
            min_col=start_col,
            min_row=helper["month_start"],
            max_row=helper["month_end"],
        )
    )
    line = LineChart()
    line.add_data(
        Reference(
            ws,
            min_col=start_col + 2,
            min_row=helper["month_header_row"],
            max_row=helper["month_end"],
        ),
        titles_from_data=True,
    )
    chart += line
    chart.y_axis.title = "金额（元）"
    return chart


def _setup_title(ws, title: str, start_col: int, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(1, start_col, title)
    cell.font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.sheet_view.showGridLines = False


def _style_body_cell(cell, fill=INPUT_FILL) -> None:
    cell.fill = fill
    cell.font = Font(name="Microsoft YaHei", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_summary_body_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        _style_body_cell(ws.cell(row, col), FORMULA_FILL if col > 1 else INPUT_FILL)
    for col in (2, 3):
        ws.cell(row, col).number_format = "#,##0"
    for col in (4, 5, 6):
        ws.cell(row, col).number_format = "0.0%"


def _style_sales_total_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        _style_body_cell(ws.cell(row, col), TOTAL_FILL)
        ws.cell(row, col).font = Font(name="Microsoft YaHei", bold=True, color="17324D")
    for col in (2, 3):
        ws.cell(row, col).number_format = "#,##0"
    for col in (4, 5, 6):
        ws.cell(row, col).number_format = "0.0%"


def _finish_sales_sheet(ws, max_col: int) -> None:
    widths = {
        1: 14,
        2: 16,
        3: 14,
        4: 16,
        5: 16,
        6: 14,
        7: 18,
        8: 14,
    }
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 15)
    apply_print_settings(ws)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _finish_summary_sheet(ws, total_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col > 1 else 14
    for row in range(4, total_row + 1):
        for col in (2, 3):
            ws.cell(row, col).number_format = "#,##0"
        for col in (4, 5, 6):
            ws.cell(row, col).number_format = "0.0%"
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{total_row}"
    _finish_sales_sheet(ws, max_col)


def _merge_sales_detail_labels(
    ws,
    details: list[dict[str, Any]],
    start_row: int,
) -> None:
    region_start = start_row
    product_start = start_row
    previous_region = details[0]["地区"]
    previous_product = details[0]["产品"]
    for offset, record in enumerate(details[1:], start=1):
        row = start_row + offset
        if record["产品"] != previous_product or record["地区"] != previous_region:
            if row - 1 > product_start:
                ws.merge_cells(start_row=product_start, start_column=2, end_row=row - 1, end_column=2)
            product_start = row
            previous_product = record["产品"]
        if record["地区"] != previous_region:
            if row - 1 > region_start:
                ws.merge_cells(start_row=region_start, start_column=1, end_row=row - 1, end_column=1)
            region_start = row
            previous_region = record["地区"]
    end_row = start_row + len(details) - 1
    if end_row > product_start:
        ws.merge_cells(start_row=product_start, start_column=2, end_row=end_row, end_column=2)
    if end_row > region_start:
        ws.merge_cells(start_row=region_start, start_column=1, end_row=end_row, end_column=1)
    for merged in ws.merged_cells.ranges:
        ws.cell(merged.min_row, merged.min_col).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _add_sales_validations(ws, start_row: int, end_row: int) -> None:
    region_validation = DataValidation(
        type="list",
        formula1='"北美,欧洲,亚洲,其他"',
        allow_blank=False,
    )
    product_validation = DataValidation(
        type="list",
        formula1='"电子产品,家居用品,服装"',
        allow_blank=False,
    )
    ws.add_data_validation(region_validation)
    ws.add_data_validation(product_validation)
    region_validation.add(f"A{start_row}:A{end_row}")
    product_validation.add(f"B{start_row}:B{end_row}")


def _sum_details(
    details: list[dict[str, Any]],
    key: str,
    value: str,
    measure: str,
) -> float:
    return sum(_number(item.get(measure)) for item in details if item.get(key) == value)


def _write_instructions(ws, plan: dict[str, Any]) -> None:
    ws["A1"] = str(plan.get("title") or "绩效薪酬模型使用说明")
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="24588A")
    ws.merge_cells("A1:D1")
    rows = [
        ("工作簿结构", "参数表保存权重、等级、调薪和缺勤扣分参数；明细表通过公式引用参数表。"),
        ("输入区域", "明细表浅黄色单元格为原始输入；隐藏辅助列保存缺勤天数和未合并部门。"),
        ("计算区域", "浅绿色单元格为 Excel 公式，包括考勤得分、加权总分、等级、排名和调整后薪资。"),
        ("维护方式", "修改参数表后，使用 Excel、WPS 或 LibreOffice 打开并重算即可更新结果。"),
        ("风险提醒", "绩效、工资和薪酬调整结果需要人工复核。"),
    ]
    for row, (label, value) in enumerate(rows, start=3):
        ws.cell(row, 1, label).font = Font(name="Microsoft YaHei", bold=True)
        ws.cell(row, 2, value)
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)


def _write_parameter_sheet(
    ws,
    weights: dict[str, Any],
    grades: dict[str, Any],
    adjustments: dict[str, Any],
    deductions: dict[str, Any],
) -> dict[str, Any]:
    ws["A1"] = "绩效评估参数表"
    ws.merge_cells("A1:M1")
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="24588A")
    ws["A1"].alignment = Alignment(horizontal="center")

    weight_rows = _write_block(ws, 3, 1, "指标权重", weights)
    grade_records = sorted(
        grades.get("records", []),
        key=lambda item: _number(item.get("最低分")),
    )
    grade_table = {**grades, "records": grade_records}
    grade_rows = _write_block(ws, 3, 4, "绩效等级判定标准", grade_table)
    adjustment_rows = _write_block(ws, 3, 8, "薪酬调整系数", adjustments)

    ws.cell(3, 11, "缺勤扣分对照表")
    ws.merge_cells(start_row=3, start_column=11, end_row=3, end_column=13)
    _style_section_title(ws.cell(3, 11))
    for col, value in enumerate(["起始缺勤天数", "缺勤天数范围（天）", "扣分值"], 11):
        ws.cell(4, col, value)
        _style_header(ws.cell(4, col))
    deduction_rows = []
    for row, record in enumerate(deductions.get("records", []), start=5):
        range_value = next(
            (
                value
                for key, value in record.items()
                if "缺勤天数" in str(key) or "范围" in str(key)
            ),
            "",
        )
        lower = _range_lower_bound(range_value)
        deduction = next(
            (value for key, value in record.items() if "扣分" in str(key)),
            0,
        )
        ws.cell(row, 11, lower)
        ws.cell(row, 12, range_value)
        ws.cell(row, 13, _number(deduction))
        deduction_rows.append(row)
    _style_block_body(ws, 5, 4 + len(deduction_rows), 11, 13)

    for row in weight_rows:
        ws.cell(row, 2).number_format = "0%"
    for row in adjustment_rows:
        ws.cell(row, 9).number_format = "0%"
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    set_reasonable_column_widths(ws)
    apply_print_settings(ws)

    weight_map = {
        str(ws.cell(row, 1).value): f"'参数表'!$B${row}"
        for row in weight_rows
    }
    weight_values = {
        str(ws.cell(row, 1).value): _number(ws.cell(row, 2).value)
        for row in weight_rows
    }
    return {
        "weights": weight_map,
        "weight_values": weight_values,
        "grade_range": (
            f"'参数表'!$E${min(grade_rows)}:$E${max(grade_rows)}",
            f"'参数表'!$D${min(grade_rows)}:$D${max(grade_rows)}",
        ),
        "adjustment_range": (
            f"'参数表'!$H${min(adjustment_rows)}:$H${max(adjustment_rows)}",
            f"'参数表'!$I${min(adjustment_rows)}:$I${max(adjustment_rows)}",
        ),
        "deduction_range": (
            f"'参数表'!$K${min(deduction_rows)}:$K${max(deduction_rows)}",
            f"'参数表'!$M${min(deduction_rows)}:$M${max(deduction_rows)}",
        ),
        "deduction_rules": [
            (_number(ws.cell(row, 11).value), _number(ws.cell(row, 13).value))
            for row in deduction_rows
        ],
    }


def _write_block(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    table: dict[str, Any],
) -> list[int]:
    columns = [str(item) for item in table.get("columns", [])]
    records = [dict(item) for item in table.get("records", [])]
    end_col = start_col + len(columns) - 1
    ws.cell(start_row, start_col, title)
    if end_col > start_col:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=end_col,
        )
    _style_section_title(ws.cell(start_row, start_col))
    for offset, name in enumerate(columns):
        _style_header(ws.cell(start_row + 1, start_col + offset, name))
    rows: list[int] = []
    for row, record in enumerate(records, start=start_row + 2):
        rows.append(row)
        for offset, name in enumerate(columns):
            ws.cell(row, start_col + offset, record.get(name, ""))
    _style_block_body(ws, start_row + 2, start_row + 1 + len(records), start_col, end_col)
    return rows


def _write_performance_detail(
    ws,
    employees: dict[str, Any],
    refs: dict[str, Any],
    title: str,
    prompt: str,
) -> None:
    del prompt
    source_records = [dict(item) for item in employees.get("records", [])]
    records = sorted(
        source_records,
        key=lambda item: (
            _department_order(str(item.get("部门", ""))),
            -_estimated_score(item, refs),
        ),
    )
    headers = [
        "部门",
        "员工编号",
        "姓名",
        "当前月薪(元)",
        "工作质量",
        "工作效率",
        "团队协作",
        "考勤",
        "加权总分",
        "绩效等级",
        "部门内排名",
        "薪酬调整比例",
        "调整后薪资(元)",
        "备注",
    ]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="24588A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"] = "需要人工复核：绩效等级、排名和薪酬调整均由公式生成。"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    grouped_headers = [
        ("部门", 1, 1),
        ("员工编号", 2, 2),
        ("姓名", 3, 3),
        ("当前月薪(元)", 4, 4),
        ("指标评分（百分制）", 5, 8),
        ("加权总分", 9, 9),
        ("绩效等级", 10, 10),
        ("部门内排名", 11, 11),
        ("薪酬调整比例", 12, 12),
        ("调整后薪资(元)", 13, 13),
        ("备注", 14, 14),
    ]
    for label, start, end in grouped_headers:
        if start == end and start not in range(5, 9):
            ws.merge_cells(start_row=3, start_column=start, end_row=4, end_column=end)
        elif end > start:
            ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        cell = ws.cell(3, start, label)
        _style_header(cell)
    for col, label in enumerate(["工作质量(40%)", "工作效率(30%)", "团队协作(20%)", "考勤(10%)"], start=5):
        _style_header(ws.cell(4, col, label))
    ws.cell(4, 15, "缺勤天数")
    ws.cell(4, 16, "原部门")
    for row in range(3, 5):
        for col in range(1, len(headers) + 1):
            _style_header(ws.cell(row, col))

    data_rows: list[int] = []
    current_row = 5
    for department, group in _group_records(records):
        rows: list[int] = []
        for record in group:
            row = current_row
            rows.append(row)
            data_rows.append(row)
            values = [
                department,
                record.get("员工编号", ""),
                record.get("姓名", ""),
                _number(record.get("当前月薪(元)")),
                _number(record.get("工作质量")),
                _number(record.get("工作效率")),
                _number(record.get("团队协作")),
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row, col, value)
            ws.cell(row, 15, _number(record.get("缺勤天数")))
            ws.cell(row, 16, department)
            deduction_keys, deduction_values = refs["deduction_range"]
            ws.cell(
                row,
                8,
                f"=IFERROR(10-LOOKUP(O{row},{deduction_keys},{deduction_values}),0)",
            )
            weights = refs["weights"]
            ws.cell(
                row,
                9,
                "=IFERROR("
                f"E{row}*{_weight_ref(weights, '工作质量')}+"
                f"F{row}*{_weight_ref(weights, '工作效率')}+"
                f"G{row}*{_weight_ref(weights, '团队协作')}+"
                f"H{row}*{_weight_ref(weights, '考勤')}"
                ",0)",
            )
            grade_scores, grade_labels = refs["grade_range"]
            ws.cell(
                row,
                10,
                f'=IFERROR(LOOKUP(I{row},{grade_scores},{grade_labels}),"")',
            )
            adjustment_labels, adjustment_values = refs["adjustment_range"]
            ws.cell(
                row,
                12,
                f'=IFERROR(INDEX({adjustment_values},MATCH(J{row},{adjustment_labels},0)),0)',
            )
            ws.cell(row, 13, f"=IFERROR(D{row}*(1+L{row}),0)")
            ws.cell(row, 14, f'=IF(I{row}<60,"⚠️警告：需改进","")')
            current_row += 1

        subtotal_row = current_row
        ws.merge_cells(
            start_row=subtotal_row,
            start_column=1,
            end_row=subtotal_row,
            end_column=3,
        )
        ws.cell(subtotal_row, 1, f"{department}汇总")
        ws.cell(subtotal_row, 4, f"=SUM(D{rows[0]}:D{rows[-1]})")
        ws.cell(
            subtotal_row,
            9,
            f'="平均："&TEXT(AVERAGE(I{rows[0]}:I{rows[-1]}),"0.0")'
            f'&" / 最高："&TEXT(MAX(I{rows[0]}:I{rows[-1]}),"0.0")'
            f'&" / 最低："&TEXT(MIN(I{rows[0]}:I{rows[-1]}),"0.0")',
        )
        ws.cell(subtotal_row, 13, f"=SUM(M{rows[0]}:M{rows[-1]})")
        _style_summary_row(ws, subtotal_row, 14, SUBTOTAL_FILL)
        if len(rows) > 1:
            ws.merge_cells(
                start_row=rows[0],
                start_column=1,
                end_row=rows[-1],
                end_column=1,
            )
        current_row += 1

    last_data_bound = current_row - 1
    for row in data_rows:
        ws.cell(
            row,
            11,
            f"=1+SUMPRODUCT(($P$5:$P${last_data_bound}=P{row})*"
            f"($I$5:$I${last_data_bound}>I{row}))",
        )
    total_row = current_row
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.cell(total_row, 1, "总计")
    salary_ranges = ",".join(f"D{row}" for row in data_rows)
    score_ranges = ",".join(f"I{row}" for row in data_rows)
    adjusted_ranges = ",".join(f"M{row}" for row in data_rows)
    ws.cell(total_row, 4, f"=SUM({salary_ranges})")
    ws.cell(
        total_row,
        9,
        f'="平均："&TEXT(AVERAGE({score_ranges}),"0.0")'
        f'&" / 最高："&TEXT(MAX({score_ranges}),"0.0")'
        f'&" / 最低："&TEXT(MIN({score_ranges}),"0.0")',
    )
    ws.cell(total_row, 13, f"=SUM({adjusted_ranges})")
    _style_summary_row(ws, total_row, 14, TOTAL_FILL)

    formula_columns = {8, 9, 10, 11, 12, 13, 14}
    for row in data_rows:
        for col in range(1, 15):
            cell = ws.cell(row, col)
            cell.fill = FORMULA_FILL if col in formula_columns else INPUT_FILL
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Microsoft YaHei", size=10)
        ws.cell(row, 4).number_format = '¥#,##0;[Red]-¥#,##0'
        ws.cell(row, 9).number_format = "0.0"
        ws.cell(row, 12).number_format = "0%"
        ws.cell(row, 13).number_format = '¥#,##0;[Red]-¥#,##0'
    first, last = min(data_rows), max(data_rows)
    ws.conditional_formatting.add(
        f"J{first}:J{last}",
        FormulaRule(
            formula=[f'J{first}="A"'],
            fill=PatternFill("solid", fgColor="C6EFCE"),
        ),
    )
    ws.conditional_formatting.add(
        f"J{first}:J{last}",
        FormulaRule(
            formula=[f'OR(J{first}="D",J{first}="E")'],
            fill=PatternFill("solid", fgColor="FFC7CE"),
        ),
    )
    ws.conditional_formatting.add(
        f"N{first}:N{last}",
        FormulaRule(
            formula=[f'N{first}<>""'],
            font=Font(color="C00000", bold=True),
        ),
    )

    ws.column_dimensions["O"].hidden = True
    ws.column_dimensions["P"].hidden = True
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = (
            18 if col in {9, 13, 14} else 14
        )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"B4:N{total_row - 1}"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:N{total_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    apply_print_settings(ws)

    # A comparison chart of real input values (names vs current salary) so the
    # workbook satisfies the user's chart request even on this fallback path.
    if data_rows:
        first_data, last_data = min(data_rows), max(data_rows)
        chart = BarChart()
        chart.type = "col"
        chart.title = "各员工当前月薪对比"
        chart.style = 10
        chart.height = 8
        chart.width = 18
        chart.gapWidth = 80
        chart.add_data(
            Reference(ws, min_col=4, min_row=first_data, max_row=last_data),
            titles_from_data=False,
        )
        chart.set_categories(
            Reference(ws, min_col=3, min_row=first_data, max_row=last_data)
        )
        chart.legend = None
        chart.y_axis.numFmt = "#,##0"
        ws.add_chart(chart, f"B{total_row + 2}")


def _find_table(
    tables: list[dict[str, Any]],
    required_headers: set[str],
) -> dict[str, Any] | None:
    for table in tables:
        headers = {str(item) for item in table.get("columns", [])}
        if required_headers.issubset(headers):
            return table
    return None


def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_section_title(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.font = Font(name="Microsoft YaHei", bold=True, color="17324D")
    cell.alignment = Alignment(horizontal="center")


def _style_block_body(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_summary_row(ws, row: int, max_col: int, fill) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = Font(name="Microsoft YaHei", bold=True, color="17324D")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 4).number_format = '¥#,##0;[Red]-¥#,##0'
    ws.cell(row, 13).number_format = '¥#,##0;[Red]-¥#,##0'


def _range_lower_bound(value: Any) -> int:
    text = str(value)
    match = re.search(r"\d+", text)
    return int(match.group()) if match else 0


def _number(value: Any) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value or "").strip().replace(",", "").replace("¥", "").replace("￥", "")
    if text.endswith("%"):
        text = text[:-1]
        try:
            return float(text.lstrip("+")) / 100
        except ValueError:
            return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _department_order(value: str) -> tuple[int, str]:
    order = {"技术部": 0, "市场部": 1, "行政部": 2}
    return order.get(value, len(order)), value


def _estimated_score(record: dict[str, Any], refs: dict[str, Any]) -> float:
    absence = _number(record.get("缺勤天数"))
    deduction = 0.0
    for lower_bound, value in sorted(refs.get("deduction_rules", [])):
        if absence >= lower_bound:
            deduction = value
    attendance = 10 - deduction
    weights = refs.get("weight_values", {})
    return (
        _number(record.get("工作质量")) * _weight_value(weights, "工作质量")
        + _number(record.get("工作效率")) * _weight_value(weights, "工作效率")
        + _number(record.get("团队协作")) * _weight_value(weights, "团队协作")
        + attendance * _weight_value(weights, "考勤")
    )


def _group_records(records: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    groups: list[tuple[str, list[dict[str, Any]]]] = []
    for record in records:
        department = str(record.get("部门", ""))
        if not groups or groups[-1][0] != department:
            groups.append((department, [record]))
        else:
            groups[-1][1].append(record)
    return groups


def _weight_ref(weights: dict[str, str], name: str) -> str:
    return next(
        (reference for label, reference in weights.items() if name in label),
        "0",
    )


def _weight_value(weights: dict[str, float], name: str) -> float:
    return next(
        (float(value) for label, value in weights.items() if name in label),
        0.0,
    )
