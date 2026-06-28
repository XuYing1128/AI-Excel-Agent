# -*- coding: utf-8 -*-
"""对比看 output 目录里两个 xlsx：豆包自己存的 vs 兜底重造的。"""
import glob

from openpyxl import load_workbook

for p in sorted(glob.glob("outputs/_filltable_verify/**/output/*.xlsx", recursive=True)):
    print("=" * 60)
    print(p)
    wb = load_workbook(p)
    print("工作表:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  -- {sn} ({ws.max_row} 行 x {ws.max_column} 列)")
        for row in ws.iter_rows(values_only=True):
            print("    ", row)
