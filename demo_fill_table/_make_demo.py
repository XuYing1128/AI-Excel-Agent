# -*- coding: utf-8 -*-
"""生成“填表验证”demo：表A（工资数据源）+ 表B（待填花名册，含需保留的格式与其它工作表）。

运行：.venv/Scripts/python.exe demo_fill_table/_make_demo.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

here = Path(__file__).resolve().parent

# 表A：数据源（只有姓名 + 月工资）
a = Workbook()
ws = a.active
ws.title = "工资数据"
ws.append(["姓名", "月工资"])
for name, salary in [("张三", 8000), ("李四", 9000), ("王五", 7500)]:
    ws.append([name, salary])
a.save(here / "表A_工资数据.xlsx")

# 表B：目标花名册（有自己的表头样式、列顺序、备注列，以及第二个工作表；月工资留空）
b = Workbook()
ws = b.active
ws.title = "花名册"
headers = ["姓名", "部门", "月工资", "备注"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
for name, dept, note in [("张三", "销售部", "组长"), ("李四", "研发部", ""), ("王五", "市场部", "新人")]:
    ws.append([name, dept, None, note])
note_sheet = b.create_sheet("填表说明")
note_sheet["A1"] = "本工作表与花名册的格式、列顺序、备注列都请保持不变，只把月工资填上即可。"
b.save(here / "表B_花名册.xlsx")

print("OK:", here)
