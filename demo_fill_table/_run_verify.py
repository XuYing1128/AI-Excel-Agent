# -*- coding: utf-8 -*-
"""用真实豆包跑“填表 demo”，并自动核对就地编辑是否成功。"""
import sys
import time
from pathlib import Path

sys.path.insert(0, "src")

from openpyxl import load_workbook

from excel_agent.task_spec import TaskSpec
from excel_agent.task_paths import create_task_paths
from excel_agent.services.generation_service import generate_from_task_spec

demo = Path("demo_fill_table").resolve()
spec = TaskSpec(
    task_type="generic_table",
    user_goal="把表A里每个人的月工资，按姓名填到表B花名册的“月工资”列，表B其它内容都不要动。",
    input_files=[str(demo / "表A_工资数据.xlsx"), str(demo / "表B_花名册.xlsx")],
)
paths = create_task_paths("generic_table", Path("outputs") / "_filltable_verify")


def prog(stage, msg):
    print(f"  [{stage}] {msg}", flush=True)


t0 = time.time()
print(">>> 调用豆包跑填表任务……", flush=True)
result = generate_from_task_spec(spec, paths, progress=prog)
print(f">>> 完成，用时 {time.time() - t0:.0f}s", flush=True)
print("success:", result.success, "| mode:", result.mode)
print("output:", result.output_file)
for n in (result.notices or []):
    print("notice:", n)
if result.error:
    print("error:", result.error)

checks = []
if result.output_file and Path(result.output_file).exists():
    wb = load_workbook(result.output_file)
    sheets = wb.sheetnames
    print("--- 结果工作表:", sheets)
    checks.append(("两个工作表都在", {"花名册", "填表说明"}.issubset(set(sheets))))
    if "花名册" in sheets:
        rows = list(wb["花名册"].iter_rows(values_only=True))
        for r in rows:
            print("   花名册:", r)
        flat = [str(c) for r in rows for c in r if c is not None]
        nums = [c for r in rows for c in r if isinstance(c, (int, float))]
        checks.append(("月工资已填(8000/9000/7500)", {8000, 9000, 7500}.issubset(set(nums))))
        checks.append(("部门保留", all(d in flat for d in ["销售部", "研发部", "市场部"])))
        checks.append(("备注保留", all(d in flat for d in ["组长", "新人"])))
    if "填表说明" in sheets:
        a1 = str(wb["填表说明"]["A1"].value or "")
        print("   填表说明A1:", a1)
        checks.append(("填表说明工作表保留", "保持不变" in a1))

print("=== 判定 ===")
for name, passed in checks:
    print(("[OK]" if passed else "[FAIL]"), name)
print("ALL_PASS:", bool(checks) and all(p for _, p in checks))
