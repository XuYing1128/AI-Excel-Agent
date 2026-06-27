# -*- coding: utf-8 -*-
"""端到端测根因诊断：mock builder 跑出『循环引用→修复』真实轨迹，再用真实 reviewer 归因。"""
import shutil
import sys
import types
from pathlib import Path

sys.path.insert(0, "src")
from excel_agent.api_settings import ApiSettings
from excel_agent.services.agent import orchestrator
from excel_agent.services.agent.orchestrator import run_agent
from excel_agent.services.diagnostic_report import generate_diagnostic_report
from excel_agent.services.recalc import recalc_workbook
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec import TaskSpec
from excel_agent.validators import inspect_workbook, validate_workbook


class _Call:
    def __init__(self, name, args, cid):
        self.id = cid
        self.name = name
        self.arguments = args


class _Resp:
    def __init__(self, tool_calls):
        self.success = True
        self.error = None
        self.message = None
        self.tool_calls = tool_calls
        self.content = ""


# step1：模型写了自引用的平均分公式（B2 引用含自己的范围）→ 真算会报 #VALUE!
WRITE_BAD = (
    "from openpyxl import Workbook\n"
    "wb = Workbook(); ws = wb.active; ws.title = '科目统计'\n"
    "ws['A1'] = '科目'; ws['A2'] = '语文'; ws['A3'] = '数学'\n"
    "ws['B1'] = '平均分'\n"
    "ws['B2'] = '=AVERAGE(B2:B3)'\n"
    "ws['B3'] = '=AVERAGE(B3:B4)'\n"
    "wb.save(OUTPUT_FILE)\n"
)
# step2：补样本列、改成引用真实数值
FIX = (
    "from openpyxl import load_workbook\n"
    "wb = load_workbook(OUTPUT_FILE); ws = wb['科目统计']\n"
    "ws['C1'] = '样本'; ws['C2'] = 80; ws['C3'] = 90\n"
    "ws['B2'] = '=AVERAGE(C2:C2)'\n"
    "ws['B3'] = '=AVERAGE(C3:C3)'\n"
    "wb.save(OUTPUT_FILE)\n"
)

out_dir = Path("outputs/_test_diagnosis")
if out_dir.exists():
    shutil.rmtree(out_dir)

responses = [
    _Resp([_Call("run_python", {"code": WRITE_BAD}, "c1"), _Call("finish_task", {"summary": "做完了"}, "f1")]),
    _Resp([_Call("run_python", {"code": FIX}, "c2"), _Call("finish_task", {"summary": "已修"}, "f2")]),
]


def fake_chat(*a, **k):
    return responses.pop(0) if responses else _Resp([])


# 只把 builder 的实际模型调用换成 mock；load_model_settings / get_provider / reviewer 都用真实配置，
# 这样 run_agent 能正常起步(真实豆包 provider)，诊断也能拿到真实 reviewer(deepseek)。
orchestrator.model_registry.chat_with_tools = fake_chat

spec = TaskSpec(task_type="generic_table", user_goal="做一个各科平均分统计表")
paths = create_task_paths("generic_table", out_dir)
res = run_agent(spec, paths, max_steps=6)

rec = recalc_workbook(res.output_file) if res.output_file else {"available": False, "ok": True, "error_cells": []}
rep = validate_workbook(res.output_file) if res.output_file else {"status": "error", "warnings": [], "errors": []}
wbsum = inspect_workbook(res.output_file) if res.output_file else {"sheets": []}
wbsum["sheet_count"] = len(wbsum.get("sheets", []))

print(">>> 调用真实 reviewer 做根因诊断……", flush=True)
report = generate_diagnostic_report(
    spec, paths,
    validation_report=rep, recalc_result=rec, workbook_summary=wbsum,
    generation_summary=res.to_dict(), api_settings=ApiSettings(),
)

lines = [
    f"builder success={res.success}  steps={res.steps}",
    f"诊断 enabled={report.get('enabled')}  model={report.get('model')}  error={report.get('error')}",
    f"总评：{report.get('overall')}",
    "",
]
for i, p in enumerate(report.get("problems", []), 1):
    lines += [
        f"[{i}] {p['title']}  ({p['layer']}层, 把握{p['confidence']})",
        f"    证据: {p['evidence']}",
        f"    原因: {p['likely_cause']}",
        f"    改进: {p['suggestion']}",
        "",
    ]
Path("demo_teaching/_test_diagnosis.txt").write_text("\n".join(lines), encoding="utf-8")
print("done; problems:", len(report.get("problems", [])))
