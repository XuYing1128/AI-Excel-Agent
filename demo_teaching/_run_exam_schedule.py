# -*- coding: utf-8 -*-
"""排考场景压测：验证改进后 AI 排考的稳定性（公式防 #VALUE!、validation 结果）。

Codex 报告里排考是唯一“不合格”场景（冲突检查表出现 #VALUE!）。这里看自检清单
（公式用 IFERROR、注意类型匹配）和稳定性改动是否改善。
"""
import sys
import time
from pathlib import Path

sys.path.insert(0, "src")

from openpyxl import Workbook, load_workbook

from excel_agent.api_settings import ApiSettings
from excel_agent.task_spec import TaskSpec
from excel_agent.task_paths import create_task_paths
from excel_agent.services.generation_service import generate_from_task_spec
from excel_agent.validators import validate_workbook

here = Path("demo_teaching").resolve()
here.mkdir(exist_ok=True)
src = here / "期末考试报名名单.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "报名名单"
ws.append(["学号", "姓名", "报考科目"])
names = ["张伟", "王芳", "李娜", "刘洋", "陈静", "杨杰", "赵敏", "黄磊",
         "周强", "吴婷", "孙莉", "马涛", "胡军", "郭敏", "林峰"]
subjects = ["高等数学", "大学英语", "线性代数", "大学物理"]
for i, name in enumerate(names):
    ws.append([f"2024{301 + i}", name, subjects[i % len(subjects)]])
wb.save(src)

prompt = (
    "这是期末考试报名名单。请安排考场并生成排考表：\n"
    "1. 共有 3 个考场（每个考场 30 个座位），分 2 天、每天上午下午共 4 个场次；\n"
    "2. 给每位考生分配：考场、场次（第几天第几时段）、座位号；\n"
    "3. 单独一个『冲突检查』工作表，用公式检查是否有同一考场同一场次座位号重复，"
    "或同一考生被排进多个场次；\n"
    "4. 一个『考场容量』工作表，统计每个考场每场次已分配人数，不超过 30。"
)

spec = TaskSpec(task_type="generic_table", user_goal=prompt, input_files=[str(src)])
paths = create_task_paths("generic_table", Path("outputs") / "_exam_schedule")

print(">>> 排考任务，调用豆包……", flush=True)
t0 = time.time()
gen = generate_from_task_spec(
    spec, paths, api_settings=ApiSettings(), progress=lambda s, m: print(f"  [{s}] {m}", flush=True)
)
print(f">>> 用时 {time.time() - t0:.0f}s | success={gen.success} | mode={gen.mode}", flush=True)
print("output:", gen.output_file)
for n in (gen.notices or []):
    print("notice:", n)

if gen.output_file and Path(gen.output_file).exists():
    report = validate_workbook(gen.output_file)
    print(
        "validation:", report.get("status"),
        "| warnings:", len(report.get("warnings", [])),
        "| errors:", len(report.get("errors", [])),
    )
    wb2 = load_workbook(gen.output_file)
    print("工作表:", wb2.sheetnames)
    total_f = 0
    iferror_f = 0
    for sn in wb2.sheetnames:
        for row in wb2[sn].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    total_f += 1
                    if "IFERROR" in c.value.upper():
                        iferror_f += 1
    print(f"公式总数: {total_f} | 含 IFERROR: {iferror_f}")
    print("ALL_DONE: True")
else:
    print("ALL_DONE: False（没有生成文件）")
