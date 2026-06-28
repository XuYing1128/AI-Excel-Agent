# -*- coding: utf-8 -*-
"""测 多模型校验(planner→builder→reviewer) + 真实场景多图表(柱/折线/饼)，并收集诊断信号。
结果写 UTF-8 文件，避免控制台乱码。"""
import json
import sys
import time
from pathlib import Path

sys.path.insert(0, "src")
from openpyxl import load_workbook

from excel_agent.api_settings import ApiSettings
from excel_agent.services.api_task_planner import enhance_task_spec_draft
from excel_agent.services.generation_service import generate_from_task_spec
from excel_agent.services.recalc import recalc_workbook
from excel_agent.services.subjective_review_service import run_subjective_review
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec_builder import build_task_spec_draft
from excel_agent.validators import inspect_workbook, validate_workbook

PROMPT = (
    "这是某公司2025年各产品线四个季度销售额（万元），请做一份销售分析工作簿：\n"
    "1. 明细表保留原始数据，并用公式算每个产品线的全年合计；\n"
    "2. 画一张各产品线全年合计的柱状图；\n"
    "3. 画一张四个季度总销售额的折线图（看趋势）；\n"
    "4. 画一张各产品线全年合计占比的饼图。\n"
    "产品线\t一季度\t二季度\t三季度\t四季度\n"
    "智能手机\t1200\t1350\t1500\t1680\n"
    "笔记本\t800\t820\t790\t910\n"
    "平板\t450\t480\t520\t560\n"
    "配件\t300\t340\t380\t420"
)

out = []
def log(*a):
    out.append(" ".join(str(x) for x in a))


def dump():
    Path("demo_teaching/_test_charts_review.txt").write_text("\n".join(out), encoding="utf-8")


# ---- planner ----
draft = build_task_spec_draft(PROMPT, [])
plan = enhance_task_spec_draft(draft, user_prompt=PROMPT, input_file_names=[], settings=ApiSettings())
spec = plan.draft.task_spec
log("=== planner ===")
log("used_api:", plan.used_api, "| task_type:", spec.task_type, "| include_charts:", spec.include_charts)
log("chart_types:", spec.options.get("chart_types"), "| chart_req:", spec.options.get("chart_requirements"))
log("planner_msg:", (plan.message or "")[:200])

paths = create_task_paths(spec.task_type or "generic_table", Path("outputs") / "_charts_review")

# ---- builder ----
log("\n=== builder ===")
t0 = time.time()
gen = generate_from_task_spec(spec, paths, api_settings=ApiSettings())
log(f"time={time.time() - t0:.0f}s success={gen.success} mode={gen.mode}")
for n in (gen.notices or []):
    log("  notice:", n)
if gen.error:
    log("  error:", gen.error)

if not (gen.output_file and Path(gen.output_file).exists()):
    log("RESULT: NO FILE")
    dump()
    print("done (no file)")
    sys.exit()

# ---- 真算（诊断信号：具体错误单元格）----
rec = recalc_workbook(gen.output_file)
log("\n=== 真算 ===")
log("available:", rec.get("available"), "ok:", rec.get("ok"), "error_cells:", len(rec.get("error_cells", [])))
for e in rec.get("error_cells", [])[:10]:
    log("   ", e)

# ---- 图表细节 ----
log("\n=== 图表 ===")
wb = load_workbook(gen.output_file)
log("sheets:", wb.sheetnames)
chart_types = []
for ws in wb.worksheets:
    for ch in ws._charts:
        cname = type(ch).__name__
        chart_types.append(cname)
        refs = []
        for s in ch.series:
            try:
                refs.append(s.val.numRef.f if (s.val and s.val.numRef) else None)
            except Exception:
                refs.append("?")
        log(f"  [{ws.title}] {cname} series={len(ch.series)} refs={refs[:5]}")
log("chart_types:", chart_types)
log("覆盖 柱/折线/饼:",
    any("Bar" in c for c in chart_types),
    any("Line" in c for c in chart_types),
    any("Pie" in c or "Doughnut" in c for c in chart_types))

# ---- validation + inspect ----
rep = validate_workbook(gen.output_file)
wbsum = inspect_workbook(gen.output_file)
wbsum["sheet_count"] = len(wbsum.get("sheets", []))
log("\n=== validation ===", rep.get("status"), "warn:", len(rep.get("warnings", [])), "err:", len(rep.get("errors", [])))
for w in rep.get("warnings", [])[:8]:
    log("  WARN:", w.get("check"), w.get("message"))

# ---- reviewer 多模型校验 ----
log("\n=== reviewer 多模型校验 ===")
rev = run_subjective_review(
    spec,
    {"status": rep.get("status"), "error_count": len(rep.get("errors", [])), "warning_count": len(rep.get("warnings", []))},
    wbsum,
    gen.to_dict(),
    paths,
    ApiSettings(),
)
log("enabled:", rev.get("enabled"), "| agreement:", rev.get("agreement"))
log("user_notice:", rev.get("user_notice"))
for r in rev.get("reviews", []):
    log(f"  model={r.get('model')} status={r.get('status')} fit={r.get('fit_to_user_goal')} risk={r.get('over_design_risk')}")
    for c in r.get("concerns", []):
        log("    concern:", c)
    for s in r.get("suggestions", []):
        log("    suggestion:", s)

# ---- run_log 事件（诊断信号：轨迹）----
log("\n=== run_log 事件序列 ===")
rl = paths.task_dir / "run_log.json"
if rl.exists():
    data = json.loads(rl.read_text(encoding="utf-8"))
    events = data if isinstance(data, list) else data.get("events", [])
    for e in events:
        if isinstance(e, dict):
            extra = ""
            if e.get("event") == "agent_tool_called":
                extra = f" tool={e.get('details',{}).get('tool')}"
            log(f"  {e.get('event')} [{e.get('status')}]{extra}")

log("\nALL_DONE")
dump()
print("done")
