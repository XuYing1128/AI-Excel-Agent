# -*- coding: utf-8 -*-
"""第2步回归：成绩 / 排考 / 填表 / 内联表 4 个真实任务。
断言：① 都走 agent_orchestrator（内联表不再被本地抢跑）；② LibreOffice 重算 0 错误值；
③ 各自业务检查（保真、公式、图表）。结果写 UTF-8 文件，避免控制台乱码。
"""
import shutil
import subprocess
import sys
import time
from pathlib import Path

sys.path.insert(0, "src")
from openpyxl import load_workbook
from excel_agent.api_settings import ApiSettings
from excel_agent.task_spec_builder import build_task_spec_draft
from excel_agent.task_paths import create_task_paths
from excel_agent.services.generation_service import generate_from_task_spec
from excel_agent.validators import ERROR_VALUES, validate_workbook

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
PROFILE = (Path("temp/_lo")).resolve().as_uri()
OUT: list[str] = []


def log(*a):
    OUT.append(" ".join(str(x) for x in a))


def recalc_errors(xlsx, slug):
    tmp = Path("temp/_rc") / slug
    if tmp.exists():
        shutil.rmtree(tmp)
    tmp.mkdir(parents=True)
    try:
        subprocess.run(
            [SOFFICE, f"-env:UserInstallation={PROFILE}", "--headless", "--calc",
             "--convert-to", "xlsx", "--outdir", str(tmp), str(xlsx)],
            capture_output=True, text=True, timeout=180,
        )
    except subprocess.TimeoutExpired:
        return None
    files = list(tmp.glob("*.xlsx"))
    if not files:
        return None
    cwb = load_workbook(files[0], data_only=True)
    return [f"{sn}!{c.coordinate}={c.value}" for sn in cwb.sheetnames
            for row in cwb[sn].iter_rows() for c in row if c.value in ERROR_VALUES]


GRADE_PROMPT = (
    "这是高一3班期中考试成绩（语文、数学、英语三科）。请做一份成绩统计分析工作簿：\n"
    "1. 在成绩表中为每个学生新增『总分』『平均分』『班级排名』三列，用 Excel 公式计算"
    "（排名按总分从高到低）；\n"
    "2. 新增一个『科目统计』工作表，按语文/数学/英语三科分别统计：平均分、最高分、最低分、"
    "及格率（≥60）、优秀率（≥85），都用公式；\n"
    "3. 画一张各科平均分的柱状图。\n"
    "保留每个学生的原始三科成绩，不要改动。"
)
EXAM_PROMPT = (
    "这是期末考试报名名单。请安排考场并生成排考表：\n"
    "1. 共有 3 个考场（每个考场 30 个座位），分 2 天、每天上午下午共 4 个场次；\n"
    "2. 给每位考生分配：考场、场次（第几天第几时段）、座位号；\n"
    "3. 单独一个『冲突检查』工作表，用公式检查是否有同一考场同一场次座位号重复，"
    "或同一考生被排进多个场次；\n"
    "4. 一个『考场容量』工作表，统计每个考场每场次已分配人数，不超过 30。"
)
FILL_PROMPT = "把表A里每个人的月工资，按姓名填到表B花名册的“月工资”列，表B其它内容都不要动。"
INLINE_PROMPT = (
    "下面是某店三个月销售额，请做成Excel：用公式计算每月环比增长率，并画一张销售额柱状图。\n"
    "月份\t销售额\n1月\t12000\n2月\t15000\n3月\t13500"
)


def grade_checks(wb):
    first = wb[wb.sheetnames[0]]
    has_formula = any(isinstance(c.value, str) and c.value.startswith("=")
                      for row in first.iter_rows() for c in row)
    flat = [str(c.value) for row in first.iter_rows() for c in row if c.value is not None]
    charts = sum(len(getattr(wb[sn], "_charts", [])) for sn in wb.sheetnames)
    return [
        ("成绩表含公式", has_formula),
        ("原始分保留(张伟88)", "88" in flat or 88 in [c.value for row in first.iter_rows() for c in row]),
        ("有柱状图", charts >= 1),
    ]


def exam_checks(wb):
    return [("工作表≥3", len(wb.sheetnames) >= 3)]


def fill_checks(wb):
    # 月工资可能是静态值，也可能是 VLOOKUP 公式（真算 0 错误即代表 VLOOKUP 匹配成功），
    # 两者都算“已填”；不再只找静态数字（那是上一版脚本的误判）。
    res = [("花名册在", "花名册" in wb.sheetnames), ("填表说明保留", "填表说明" in wb.sheetnames)]
    if "花名册" in wb.sheetnames:
        ws = wb["花名册"]
        flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        header = [str(c.value or "") for c in ws[1]]
        salary_filled = False
        if "月工资" in header:
            col = header.index("月工资") + 1
            filled = [ws.cell(r, col).value for r in range(2, ws.max_row + 1)
                      if ws.cell(r, col).value not in (None, "")]
            salary_filled = len(filled) >= 3
        res.append(("月工资已填(3行非空)", salary_filled))
        res.append(("部门保留", all(d in flat for d in ["销售部", "研发部", "市场部"])))
        res.append(("备注保留", all(d in flat for d in ["组长", "新人"])))
    return res


def inline_checks(wb):
    has_formula = any(isinstance(c.value, str) and c.value.startswith("=")
                      for sn in wb.sheetnames for row in wb[sn].iter_rows() for c in row)
    charts = sum(len(getattr(wb[sn], "_charts", [])) for sn in wb.sheetnames)
    flat = [str(c.value) for sn in wb.sheetnames for row in wb[sn].iter_rows() for c in row if c.value is not None]
    return [
        ("含公式(环比)", has_formula),
        ("销售额保留(12000)", any("12000" in s for s in flat) or 12000 in
         [c.value for sn in wb.sheetnames for row in wb[sn].iter_rows() for c in row]),
        ("有柱状图", charts >= 1),
    ]


dt = Path("demo_teaching")
ft = Path("demo_fill_table")
CASES = [
    ("成绩统计", "_rg_grade", GRADE_PROMPT, [(dt / "高一3班期中成绩.xlsx")], grade_checks),
    ("排考", "_rg_exam", EXAM_PROMPT, [(dt / "期末考试报名名单.xlsx")], exam_checks),
    ("填表(就地编辑)", "_rg_fill", FILL_PROMPT,
     [(ft / "表A_工资数据.xlsx"), (ft / "表B_花名册.xlsx")], fill_checks),
    ("内联数据表(无文件)", "_rg_inline", INLINE_PROMPT, [], inline_checks),
]

summary = []
for name, slug, prompt, files, checks in CASES:
    log("\n" + "=" * 54)
    log(f"=== {name} ===")
    afiles = [str(p.resolve()) for p in files]
    for p in files:
        if not p.exists():
            log(f"!! 缺少输入文件: {p}")
    draft = build_task_spec_draft(prompt, [p.name for p in files])
    spec = draft.task_spec
    spec.input_files = afiles
    paths = create_task_paths(spec.task_type or "generic_table", Path("outputs") / slug)
    t0 = time.time()
    try:
        gen = generate_from_task_spec(spec, paths, api_settings=ApiSettings())
    except Exception as e:
        log(f"EXCEPTION: {type(e).__name__}: {e}")
        summary.append((name, False, "exception"))
        continue
    used = time.time() - t0
    log(f"success={gen.success} mode={gen.mode} time={used:.0f}s")
    if gen.error:
        log(f"error={gen.error}")
    log(f"output={gen.output_file}")
    if not (gen.success and gen.output_file and Path(gen.output_file).exists()):
        log("RESULT: NO FILE / FAIL")
        summary.append((name, False, gen.mode))
        continue
    rep = validate_workbook(gen.output_file)
    log(f"validation={rep['status']} warn={len(rep['warnings'])} err={len(rep['errors'])}")
    errs = recalc_errors(gen.output_file, slug)
    recalc_ok = errs is not None and len(errs) == 0
    log(f"RECALC error cells={'<recalc failed>' if errs is None else len(errs)} {(errs or [])[:8]}")
    wb = load_workbook(gen.output_file)
    log(f"sheets={wb.sheetnames}")
    mode_ok = gen.mode == "agent_orchestrator"
    log(f"mode_is_agent={mode_ok}")
    case_ok = recalc_ok and mode_ok
    for cname, passed in checks(wb):
        log(f"  check[{cname}]={'OK' if passed else 'FAIL'}")
        case_ok = case_ok and passed
    log(f"RESULT: {'PASS' if case_ok else 'FAIL'}")
    summary.append((name, case_ok, gen.mode))

log("\n" + "=" * 54)
log("=== 汇总 ===")
for name, ok, mode in summary:
    log(f"  [{'PASS' if ok else 'FAIL'}] {name} (mode={mode})")
log(f"ALL_PASS: {all(ok for _, ok, _ in summary)}")

Path("demo_teaching/_regress_step2.txt").write_text("\n".join(OUT), encoding="utf-8")
print("regression done; ALL_PASS:", all(ok for _, ok, _ in summary))
