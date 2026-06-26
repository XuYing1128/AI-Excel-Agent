# -*- coding: utf-8 -*-
"""教务处场景端到端测试：学生成绩统计分析。

串起 planner(理解需求) → builder(做表) → reviewer(审查) 三个模型，
既测主模型做复杂表（公式 + 多工作表 + 图表）的能力，也测多模型协作。
传入空 ApiSettings()，让每个环节各自回退到对应角色（planner/builder/reviewer）。
"""
import sys
import time
from pathlib import Path

sys.path.insert(0, "src")

from openpyxl import Workbook, load_workbook

from excel_agent.api_settings import ApiSettings
from excel_agent.task_spec_builder import build_task_spec_draft
from excel_agent.services.api_task_planner import enhance_task_spec_draft
from excel_agent.services.generation_service import generate_from_task_spec
from excel_agent.services.subjective_review_service import run_subjective_review
from excel_agent.validators import validate_workbook, inspect_workbook
from excel_agent.task_paths import create_task_paths

# ---- 0. 造一份真实的班级成绩 input ----
here = Path("demo_teaching").resolve()
here.mkdir(exist_ok=True)
src_xlsx = here / "高一3班期中成绩.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "成绩"
ws.append(["学号", "姓名", "语文", "数学", "英语"])
for row in [
    ["2024301", "张伟", 88, 92, 79],
    ["2024302", "王芳", 76, 65, 88],
    ["2024303", "李娜", 91, 84, 95],
    ["2024304", "刘洋", 59, 73, 62],
    ["2024305", "陈静", 84, 90, 81],
    ["2024306", "杨杰", 67, 55, 70],
    ["2024307", "赵敏", 95, 88, 90],
    ["2024308", "黄磊", 72, 68, 58],
    ["2024309", "周强", 80, 77, 85],
    ["2024310", "吴婷", 63, 49, 71],
    ["2024311", "孙莉", 78, 82, 76],
    ["2024312", "马涛", 55, 60, 64],
]:
    ws.append(row)
wb.save(src_xlsx)

prompt = (
    "这是高一3班期中考试成绩（语文、数学、英语三科）。请做一份成绩统计分析工作簿：\n"
    "1. 在成绩表中为每个学生新增『总分』『平均分』『班级排名』三列，用 Excel 公式计算"
    "（排名按总分从高到低）；\n"
    "2. 新增一个『科目统计』工作表，按语文/数学/英语三科分别统计：平均分、最高分、最低分、"
    "及格率（≥60）、优秀率（≥85），都用公式；\n"
    "3. 画一张各科平均分的柱状图。\n"
    "保留每个学生的原始三科成绩，不要改动。"
)


def banner(text):
    print("\n" + "=" * 60)
    print(text, flush=True)


banner("阶段 1/3：planner 理解需求")
draft = build_task_spec_draft(prompt, [src_xlsx.name])
plan = enhance_task_spec_draft(
    draft, user_prompt=prompt, input_file_names=[src_xlsx.name], settings=ApiSettings()
)
spec = plan.draft.task_spec
print("  planner 是否真的调用了模型(used_api):", plan.used_api)
print("  理解出的 task_type:", spec.task_type, "| 需要图表:", spec.include_charts)
print("  planner 说明:", (plan.message or "")[:200])

spec.input_files = [str(src_xlsx)]
paths = create_task_paths(spec.task_type or "generic_table", Path("outputs") / "_teaching_grade")

banner("阶段 2/3：builder 做表")
t0 = time.time()
gen = generate_from_task_spec(
    spec, paths, api_settings=ApiSettings(), progress=lambda s, m: print(f"  [{s}] {m}", flush=True)
)
print(f"  用时 {time.time() - t0:.0f}s | success: {gen.success} | mode: {gen.mode}")
print("  output:", gen.output_file)
for n in (gen.notices or []):
    print("  notice:", n)

report = validate_workbook(paths.output_file) if paths.output_file.exists() else {"status": "error"}
wbsum = inspect_workbook(paths.output_file) if paths.output_file.exists() else {}
wbsum["sheet_count"] = len(wbsum.get("sheets", []))

banner("阶段 3/3：reviewer 审查")
rev = run_subjective_review(
    spec,
    {
        "status": report.get("status"),
        "error_count": len(report.get("errors", [])),
        "warning_count": len(report.get("warnings", [])),
    },
    wbsum,
    gen.to_dict(),
    paths,
    ApiSettings(),
)
print("  reviewer 是否启用:", rev.get("enabled"))
for r in rev.get("reviews", []):
    print("  - 审查模型:", r.get("model"), "| 结论:", r.get("status"))
    if r.get("concerns"):
        print("    关注:", r.get("concerns"))
    if r.get("suggestions"):
        print("    建议:", r.get("suggestions"))

banner("产物检查")
if paths.output_file.exists():
    out = load_workbook(paths.output_file)
    print("  工作表:", out.sheetnames)
    for sn in out.sheetnames:
        sheet = out[sn]
        print(f"  -- {sn}: {sheet.max_row} 行 x {sheet.max_column} 列")
    chart_total = sum(len(getattr(out[sn], "_charts", [])) for sn in out.sheetnames)
    print("  图表总数:", chart_total)
    print("  validation status:", report.get("status"), "| warnings:", len(report.get("warnings", [])))
    for sn in out.sheetnames:
        sheet = out[sn]
        print(f"  [{sn}] freeze={sheet.freeze_panes} filter={sheet.auto_filter.ref}")
    # 抽查成绩表是否含公式、原始分是否保留
    first = out[out.sheetnames[0]]
    header = [c.value for c in first[1]]
    print("  首表表头:", header)
    formula_cells = [
        c.value
        for row in first.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    print("  首表公式样例:", formula_cells[:3])
    print("ALL_DONE: True")
else:
    print("ALL_DONE: False（没有生成文件）")
