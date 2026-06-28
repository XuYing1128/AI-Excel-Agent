# -*- coding: utf-8 -*-
"""Agent 测试 harness：一道题(prompt[+输入文件][+期望]) → 生成 + 真算 + 校验 + (失败)诊断 → 结构化判定。

给“测试智能体”批量自动出题用。智能体只管出题，调 run_one_case() 拿**客观判定**——尤其
LibreOffice 真算金标准(抓运行时 #VALUE!/循环引用，validate 抓不到)。失败案例自动跑根因诊断、
把问题归因到 模型/工具/代码/网络 层，便于统计与改进。结果逐行写 outputs/_agent_test/results.jsonl。

命令行：
  python scripts/agent_test_harness.py --prompt "做一张..." --expect '{"min_sheets":2,"min_charts":1}'
  python scripts/agent_test_harness.py --prompt "把表A工资填到表B" --input demo/A.xlsx demo/B.xlsx --expect '{"keep_sheets_from_input":true}'
  → stdout 打印关键判定(完整记录在 results.jsonl)

import：
  import sys; sys.path.insert(0, "scripts")
  from agent_test_harness import run_one_case
  r = run_one_case(prompt, input_files=[...], expect={...}, slug="case_001")
"""
from __future__ import annotations

import argparse
import json
import sys
import time
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import load_workbook  # noqa: E402

from excel_agent.api_settings import ApiSettings  # noqa: E402
from excel_agent.services.api_task_planner import enhance_task_spec_draft  # noqa: E402
from excel_agent.services.diagnostic_report import generate_diagnostic_report  # noqa: E402
from excel_agent.services.generation_service import generate_from_task_spec  # noqa: E402
from excel_agent.services.recalc import recalc_workbook  # noqa: E402
from excel_agent.task_paths import create_task_paths  # noqa: E402
from excel_agent.task_spec_builder import build_task_spec_draft  # noqa: E402
from excel_agent.validators import inspect_workbook, validate_workbook  # noqa: E402

RESULTS_DIR = ROOT / "outputs" / "_agent_test"
RESULTS_FILE = RESULTS_DIR / "results.jsonl"


def run_one_case(
    prompt: str,
    *,
    input_files: list[str] | None = None,
    expect: dict | None = None,
    slug: str | None = None,
    use_planner: bool = True,
    diagnose_on_fail: bool = True,
) -> dict:
    """跑一道题，返回结构化判定 dict（并追加到 results.jsonl）。

    expect 支持的硬指标(都可选)：
      min_sheets:int、sheet_names_include:[str]、min_charts:int、chart_types:[str(类名子串,如 bar/line/pie)]、
      min_formula_count:int、must_contain:[str]、keep_sheets_from_input:bool(就地编辑必须保留输入的所有工作表)、
      max_time_s:int(生成耗时上限)。
    判定 passed = 生成成功 且 真算无错(LibreOffice 可用时) 且 所有 expect 硬指标通过。
    """

    expect = expect or {}
    files = [str(p) for p in (input_files or [])]
    names = [Path(p).name for p in files]
    slug = slug or f"case_{int(time.time() * 1000)}"
    record: dict = {"slug": slug, "ts": time.strftime("%Y-%m-%d %H:%M:%S"), "prompt": prompt, "input_files": names}

    try:
        draft = build_task_spec_draft(prompt, names)
        spec = draft.task_spec
        if use_planner:
            try:
                plan = enhance_task_spec_draft(draft, user_prompt=prompt, input_file_names=names, settings=ApiSettings())
                spec = plan.draft.task_spec
            except Exception:
                pass  # planner 失败不挡路，用本地草案
        spec.input_files = files
        paths = create_task_paths(spec.task_type or "generic_table", RESULTS_DIR / slug)

        t0 = time.time()
        gen = generate_from_task_spec(spec, paths, api_settings=ApiSettings())
        gen_time = round(time.time() - t0, 1)
        record["generation"] = {"success": gen.success, "mode": gen.mode, "time_s": gen_time, "error": gen.error}

        out = Path(gen.output_file) if gen.output_file else None
        if not (gen.success and out and out.exists()):
            record["passed"] = False
            record["fail_reason"] = "未产出文件 / 生成失败"
            record["task_dir"] = str(paths.task_dir)
            _maybe_diagnose(record, spec, paths, diagnose_on_fail)
            return _finish(record)

        rec = recalc_workbook(out)
        record["recalc"] = {
            "available": rec.get("available"),
            "ok": rec.get("ok"),
            "error_cells": rec.get("error_cells", [])[:20],
        }
        rep = validate_workbook(out)
        record["validation"] = {
            "status": rep.get("status"),
            "warnings": [w.get("message") for w in rep.get("warnings", [])][:10],
            "errors": [e.get("message") for e in rep.get("errors", [])][:10],
        }

        wb = load_workbook(out)
        sheets = wb.sheetnames
        chart_types = [type(ch).__name__ for ws in wb.worksheets for ch in ws._charts]
        formula_count = sum(
            1
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str) and c.value.startswith("=")
        )
        all_text = [str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value is not None]
        record["structure"] = {"sheets": sheets, "chart_types": chart_types, "formula_count": formula_count}

        checks = _run_checks(expect, sheets, chart_types, formula_count, all_text, files, gen_time)
        record["checks"] = checks

        recalc_ok = (not rec.get("available")) or bool(rec.get("ok"))
        passed = bool(gen.success and recalc_ok and all(c["passed"] for c in checks))
        record["passed"] = passed
        record["task_dir"] = str(paths.task_dir)
        if not passed:
            reasons = []
            if rec.get("available") and not rec.get("ok"):
                reasons.append(f"真算{len(rec.get('error_cells', []))}处报错")
            reasons += [c["name"] for c in checks if not c["passed"]]
            record["fail_reason"] = "；".join(reasons) or "未通过"
            _maybe_diagnose(record, spec, paths, diagnose_on_fail)
        return _finish(record)

    except Exception as exc:
        record["passed"] = False
        record["fail_reason"] = f"harness异常：{type(exc).__name__}: {exc}"
        record["traceback"] = traceback.format_exc()[-1500:]
        return _finish(record)


def _run_checks(expect, sheets, chart_types, formula_count, all_text, files, gen_time) -> list[dict]:
    checks: list[dict] = []

    def add(name: str, ok: bool) -> None:
        checks.append({"name": name, "passed": bool(ok)})

    if "min_sheets" in expect:
        add(f"工作表≥{expect['min_sheets']}", len(sheets) >= expect["min_sheets"])
    for nm in expect.get("sheet_names_include", []):
        add(f"含工作表[{nm}]", any(nm in s for s in sheets))
    if "min_charts" in expect:
        add(f"图表≥{expect['min_charts']}", len(chart_types) >= expect["min_charts"])
    for ct in expect.get("chart_types", []):
        add(f"含图表类型[{ct}]", any(ct.lower() in t.lower() for t in chart_types))
    if "min_formula_count" in expect:
        add(f"公式≥{expect['min_formula_count']}", formula_count >= expect["min_formula_count"])
    for kw in expect.get("must_contain", []):
        add(f"含内容[{kw}]", any(kw in t for t in all_text))
    if "max_time_s" in expect:
        add(f"耗时≤{expect['max_time_s']}s", gen_time <= expect["max_time_s"])
    if expect.get("keep_sheets_from_input") and files:
        kept = True
        try:
            for f in files:
                if Path(f).suffix.lower() in {".xlsx", ".xlsm"}:
                    for s in load_workbook(f).sheetnames:
                        if s not in sheets:
                            kept = False
        except Exception:
            kept = False
        add("保留输入文件全部工作表", kept)
    return checks


def _maybe_diagnose(record, spec, paths, diagnose_on_fail) -> None:
    if not diagnose_on_fail:
        return
    try:
        out = paths.output_file
        exists = out.exists()
        validation_report = validate_workbook(out) if exists else {"status": "error", "warnings": [], "errors": []}
        rec = recalc_workbook(out) if exists else {"available": False, "ok": True, "error_cells": []}
        wbsum = inspect_workbook(out) if exists else {"sheets": []}
        wbsum["sheet_count"] = len(wbsum.get("sheets", []))
        report = generate_diagnostic_report(
            spec, paths,
            validation_report=validation_report, recalc_result=rec, workbook_summary=wbsum,
            api_settings=ApiSettings(),
        )
        record["diagnosis"] = {
            "overall": report.get("overall"),
            "problems": [
                {"layer": p.get("layer"), "title": p.get("title"), "confidence": p.get("confidence")}
                for p in report.get("problems", [])
            ],
            "report_md": str(Path(paths.task_dir) / "diagnostic_report.md"),
        }
    except Exception as exc:
        record["diagnosis"] = {"error": str(exc)}


def _finish(record: dict) -> dict:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    with RESULTS_FILE.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record, ensure_ascii=False) + "\n")
    return record


def main() -> None:
    ap = argparse.ArgumentParser(description="Agent 测试 harness：跑一道题并给客观判定。")
    ap.add_argument("--prompt", required=True, help="题目（自然语言需求）")
    ap.add_argument("--input", nargs="*", default=[], help="可选输入文件路径")
    ap.add_argument("--expect", default="{}", help="期望硬指标 JSON，如 '{\"min_sheets\":2,\"min_charts\":1}'")
    ap.add_argument("--slug", default=None, help="用例标识（用于任务目录名）")
    ap.add_argument("--no-planner", action="store_true", help="跳过 planner 增强")
    ap.add_argument("--no-diagnose", action="store_true", help="失败时不跑诊断（更快）")
    args = ap.parse_args()

    result = run_one_case(
        args.prompt,
        input_files=args.input,
        expect=json.loads(args.expect),
        slug=args.slug,
        use_planner=not args.no_planner,
        diagnose_on_fail=not args.no_diagnose,
    )
    brief = {k: result.get(k) for k in ("slug", "passed", "fail_reason", "generation", "recalc", "checks", "diagnosis")}
    print(json.dumps(brief, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
