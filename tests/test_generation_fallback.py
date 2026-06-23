"""Regression tests for stable generation: the model never blocks a workbook.

These lock in the core fix - when the configured model cannot drive generation
(no tool calls, no JSON), the user still gets a usable table, and when the model
returns a plain JSON plan it is built without needing tool-calling support.
"""

import json

from openpyxl import load_workbook

from excel_agent.api_settings import ApiSettings
from excel_agent.services import generation_service
from excel_agent.services import llm_workbook_agent as agent
from excel_agent.services.custom_api_service import ApiCallResult, ToolChatResult
from excel_agent.rich_workbook_builder import build_rich_workbook
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec import TaskSpec


def _configured_settings() -> ApiSettings:
    return ApiSettings(
        enabled=True,
        base_url="https://example.com/v1",
        api_key="secret",
        model="model",
        use_for_generation=True,
    )


def _explicit_plan_spec() -> TaskSpec:
    spec = TaskSpec(
        task_type="generic_table",
        user_goal="做一个项目进度表，列为：任务、负责人、开始日期、结束日期、状态",
        output_name="项目进度表.xlsx",
        include_charts=False,
    )
    spec.options["content_plan"] = {
        "title": "项目进度表",
        "sheet_name": "进度",
        "columns": [
            {"name": "任务", "kind": "text", "role": "input"},
            {"name": "负责人", "kind": "text", "role": "input"},
            {"name": "开始日期", "kind": "date", "role": "input"},
            {"name": "结束日期", "kind": "date", "role": "input"},
            {"name": "状态", "kind": "text", "role": "input"},
        ],
        "records": [],
        "formula_rules": [],
        "summary_rules": [],
        "explicit_structure": True,
        "layout": "single_sheet",
    }
    return spec


def test_model_failure_falls_back_to_local_and_produces_file(tmp_path, monkeypatch):
    # Model cannot tool-call and returns no usable JSON in either path.
    monkeypatch.setattr(
        agent,
        "chat_completion_with_tools",
        lambda *a, **k: ToolChatResult(
            True, "我先解释一下需求……", [], None, 200, 10, {"role": "assistant", "content": "x"}
        ),
    )
    monkeypatch.setattr(
        agent,
        "chat_completion",
        lambda *a, **k: ApiCallResult(True, "这是一些解释文字，并不是 JSON。", None, 200, 10),
    )
    paths = create_task_paths("generic_table", tmp_path / "tasks")
    result = generation_service.generate_from_task_spec(
        _explicit_plan_spec(), paths, api_settings=_configured_settings()
    )

    assert result.success is True
    assert paths.output_file.exists()
    assert result.mode.startswith("local_fallback")
    assert any("本地规则" in notice for notice in result.notices)


def test_json_mode_blueprint_builds_workbook(tmp_path, monkeypatch):
    blueprint = {
        "title": "区域销售对比",
        "sheet_name": "对比",
        "columns": [
            {"key": "region", "label": "区域", "type": "text"},
            {"key": "jan", "label": "1月", "type": "money"},
            {"key": "feb", "label": "2月", "type": "money"},
            {"key": "total", "label": "合计", "type": "money", "formula": "=SUM({jan},{feb})"},
        ],
        "records": [
            {"region": "华东", "jan": 120, "feb": 130},
            {"region": "华南", "jan": 90, "feb": 95},
        ],
        "charts": [
            {"type": "column", "title": "对比", "category_key": "region", "value_keys": ["jan", "feb"]}
        ],
    }
    # Tool-calling unsupported, but plain JSON works.
    monkeypatch.setattr(
        agent,
        "chat_completion_with_tools",
        lambda *a, **k: ToolChatResult(
            True, "无法工具调用", [], None, 200, 10, {"role": "assistant", "content": "x"}
        ),
    )
    monkeypatch.setattr(
        agent,
        "chat_completion",
        lambda *a, **k: ApiCallResult(
            True, json.dumps({"blueprint": blueprint}, ensure_ascii=False), None, 200, 10
        ),
    )
    spec = TaskSpec(
        task_type="sales_report",
        user_goal="按区域对比一季度销售并做柱状对比图",
        output_name="区域销售对比.xlsx",
        include_charts=True,
    )
    paths = create_task_paths("sales_report", tmp_path / "tasks")
    result = generation_service.generate_from_task_spec(
        spec, paths, api_settings=_configured_settings()
    )

    assert result.success is True
    assert result.mode == "llm_tool_agent"
    assert paths.output_file.exists()
    wb = load_workbook(paths.output_file)
    charts = [chart for ws in wb.worksheets for chart in ws._charts]
    assert charts, "JSON-mode workbook should contain the requested chart"


def test_truncated_reasoning_reply_returns_clear_hint(monkeypatch):
    """A length-truncated, empty reply (typical of an over-budget reasoning model)
    must surface a clear, actionable message instead of a generic parse error."""
    from excel_agent.services import custom_api_service as svc

    class _Resp:
        ok = True
        status_code = 200

        class elapsed:  # noqa: N801 - mimic requests' .elapsed.total_seconds()
            @staticmethod
            def total_seconds():
                return 0.1

        @staticmethod
        def json():
            return {"choices": [{"finish_reason": "length", "message": {"content": ""}}]}

    monkeypatch.setattr(svc.requests, "post", lambda *a, **k: _Resp())
    settings = ApiSettings(
        enabled=True, base_url="https://e/v1", api_key="k", model="m", use_for_generation=True
    )
    result = svc.chat_completion(settings, system_prompt="s", user_prompt="u")
    assert result.success is False
    assert "截断" in (result.error or "")
    assert result.finish_reason == "length"


def test_default_chart_is_multi_series(tmp_path):
    blueprint = {
        "title": "对比",
        "sheet_name": "对比",
        "columns": [
            {"key": "region", "label": "区域", "type": "text"},
            {"key": "jan", "label": "1月", "type": "money"},
            {"key": "feb", "label": "2月", "type": "money"},
            {"key": "mar", "label": "3月", "type": "money"},
        ],
        "records": [
            {"region": "华东", "jan": 1, "feb": 2, "mar": 3},
            {"region": "华南", "jan": 4, "feb": 5, "mar": 6},
        ],
    }
    out = build_rich_workbook(blueprint, tmp_path / "chart.xlsx", require_charts=True)
    wb = load_workbook(out)
    chart = wb["对比"]._charts[0]
    assert len(chart.series) >= 2, "comparison chart must plot multiple series"
    assert getattr(chart, "type", None) == "col"


def test_malformed_model_json_is_repaired():
    from excel_agent.services.custom_api_service import parse_json_object

    repaired = parse_json_object(
        '{"title":"测试","columns":[{"key":"name" "label":"姓名","type":"text"}]}'
    )
    assert repaired["title"] == "测试"
    assert repaired["columns"][0]["label"] == "姓名"
