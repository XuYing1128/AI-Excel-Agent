from openpyxl import load_workbook

from excel_agent.validators import validate_workbook
from excel_agent.workbook_builder import create_workbook


def test_can_create_budget_xlsx(tmp_path):
    output = create_workbook("budget", tmp_path / "budget.xlsx")
    assert output.exists()
    wb = load_workbook(output, data_only=False)
    assert "Data" in wb.sheetnames
    assert wb["Data"]["F4"].value.startswith("=")


def test_all_core_templates_generate(tmp_path):
    table_types = [
        "personal_budget",
        "family_budget",
        "quotation",
        "invoice_draft",
        "inventory",
        "sales_report",
        "ecommerce_analysis",
        "project_plan",
        "schedule",
        "attendance",
        "finance_model",
        "dashboard",
    ]
    for table_type in table_types:
        output = create_workbook(table_type, tmp_path / f"{table_type}.xlsx")
        report = validate_workbook(output)
        assert output.exists()
        assert report["status"] in {"pass", "warn"}
