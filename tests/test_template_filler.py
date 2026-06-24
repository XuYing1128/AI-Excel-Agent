import stat

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_agent.template_filler import fill_template


def test_fill_template_clears_readonly_output_attribute(tmp_path):
    template = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["姓名", "科目"])
    wb.save(template)
    template.chmod(template.stat().st_mode & ~stat.S_IWRITE)

    data = tmp_path / "data.csv"
    pd.DataFrame([{"姓名": "张三", "科目": "语文"}]).to_csv(data, index=False)
    output = tmp_path / "out.xlsx"

    fill_template(template, [data], output, prompt="填充模板")

    # If the copied output stayed read-only, this save would fail on Windows.
    result = load_workbook(output)
    result["Data"]["C1"] = "可编辑"
    result.save(output)
    assert load_workbook(output)["Data"]["A2"].value == "张三"


def test_exam_template_fills_room_aliases_and_duplicate_headers(tmp_path):
    template = tmp_path / "exam_template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "总信息"
    ws.append(["准考证号", "姓名", "考场", "考试时间", "开始时间", "结束时间"])
    ws2 = wb.create_sheet("易查分导入")
    ws2.append(["准考证号", "姓名", "考点", "考场地址", "考场地址", "考试时间"])
    wb.save(template)

    data = tmp_path / "exam_data.csv"
    pd.DataFrame(
        [
            {"准考证号": "001", "姓名": "张三"},
            {"准考证号": "002", "姓名": "李四"},
        ]
    ).to_csv(data, index=False)
    output = tmp_path / "exam_out.xlsx"

    fill_template(
        template,
        [data],
        output,
        prompt="两个教室二教E506、二教E504，一个教室65个位置，两天考完",
    )

    result = load_workbook(output)
    assert result["总信息"]["C2"].value in {"二教E506", "二教E504"}
    assert result["总信息"]["D2"].value
    assert result["易查分导入"]["D2"].value == result["易查分导入"]["E2"].value
