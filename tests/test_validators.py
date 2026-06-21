import json

from openpyxl import Workbook

from excel_agent.validators import validate_workbook


def test_validate_workbook_returns_json(tmp_path):
    path = tmp_path / "basic.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "说明"
    data = wb.create_sheet("Data")
    data.append(["日期", "数量", "单价", "金额"])
    data.append(["2026-05-01", 2, 10, "=B2*C2"])
    wb.save(path)

    report_path = tmp_path / "report.json"
    report = validate_workbook(path, report_path)
    assert report["status"] in {"pass", "warn"}
    assert set(["status", "file", "summary", "errors", "warnings", "suggestions"]).issubset(report)
    assert "file_exists" in report["summary"]
    assert "formula_cell_count" in report["summary"]
    loaded = json.loads(report_path.read_text(encoding="utf-8"))
    assert loaded["file"].endswith("basic.xlsx")
