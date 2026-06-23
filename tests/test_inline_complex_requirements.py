from openpyxl import load_workbook

from excel_agent.inline_table_parser import extract_inline_tables
from excel_agent.services.generation_service import generate_from_task_spec
from excel_agent.services.validation_service import validate_generated_workbook
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec_builder import build_task_spec_draft


COMPLEX_PROMPT = """2026年第二季度员工绩效评估及薪酬调整表
请生成参数表和员工明细表。所有计算必须使用公式引用参数表，不得硬编码。

1. 指标权重
指标\t权重
工作质量\t40%
工作效率\t30%
团队协作\t20%
考勤\t10%

2. 绩效等级判定标准
等级\t最低分\t最高分
A\t90\t100
B\t80\t89
C\t70\t79
D\t60\t69
E\t0\t59

3. 薪酬调整系数
等级\t调整比例
A\t+15%
B\t+8%
C\t0%
D\t-5%
E\t-10%

4. 缺勤扣分对照表
缺勤天数范围（天）\t扣分值
0-2\t0
3-4\t2
5-6\t4
≥7\t10

基础数据
部门\t员工编号\t姓名\t当前月薪(元)\t缺勤天数\t工作质量\t工作效率\t团队协作
技术部\tEMP001\t张三\t15000\t1\t92\t88\t85
技术部\tEMP002\t李四\t18000\t4\t85\t90\t80
市场部\tEMP003\t王五\t12000\t0\t78\t82\t90
"""


def test_complex_prompt_detects_inline_tables_and_business_type():
    tables = extract_inline_tables(COMPLEX_PROMPT)
    assert len(tables) == 5
    assert tables[-1]["row_count"] == 3
    draft = build_task_spec_draft(COMPLEX_PROMPT, [])
    plan = draft.task_spec.options["content_plan"]
    assert draft.task_spec.task_type == "finance_model"
    assert plan["layout"] == "multi_sheet"
    assert plan["expected_data_rows"] == 3
    assert len(plan["inline_tables"]) == 5
    assert draft.clarifying_questions == []
    assert "未提供原始数据" not in " ".join(draft.task_spec.assumptions)


def test_model_failure_fallback_preserves_inline_tables(tmp_path):
    draft = build_task_spec_draft(COMPLEX_PROMPT, [])
    spec = draft.task_spec
    spec.include_charts = False
    paths = create_task_paths(
        spec.task_type,
        tmp_path / "tasks",
        output_name=spec.output_name,
    )
    result = generate_from_task_spec(spec, paths)
    assert result.success is True
    assert result.mode == "domain_compiler:performance_compensation"
    wb = load_workbook(paths.output_file, data_only=False)
    assert "说明" in wb.sheetnames
    assert "参数表" in wb.sheetnames
    assert "明细表" in wb.sheetnames
    assert wb["明细表"]["H5"].value.startswith("=IFERROR(10-LOOKUP")
    assert wb["明细表"]["M5"].value.startswith("=IFERROR(D5*(1+L5)")

    validation = validate_generated_workbook(paths.output_file, spec, paths)
    assert validation.status in {"pass", "warn"}
    assert not validation.issues
