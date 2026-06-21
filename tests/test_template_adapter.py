from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from excel_agent.io_utils import read_table
from excel_agent.rich_workbook_builder import build_rich_workbook
from excel_agent.template_adapter import apply_template_mode, inspect_template


def _make_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "固定导入页"
    ws.merge_cells("A1:C1")
    ws["A1"] = "系统导入模板"
    ws["A2"] = "姓名"
    ws["B2"] = "编号"
    ws["C2"] = "成绩"
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.append(["示例人员", "EXAMPLE", 100])
    wb.save(path)


def test_strict_template_ignores_sample_data_and_keeps_structure(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "strict.xlsx"
    _make_template(template)
    blueprint = {
        "title": "成绩导入",
        "sheet_name": "任意名称",
        "columns": [
            {"key": "name", "label": "姓名", "type": "text"},
            {"key": "id", "label": "编号", "type": "text"},
            {"key": "score", "label": "成绩", "type": "number"},
        ],
        "records": [{"name": "张三", "id": "001", "score": 95}],
    }
    build_rich_workbook(blueprint, output)

    apply_template_mode(
        output,
        template,
        mode="strict",
        blueprint=blueprint,
        use_template_data=False,
    )

    wb = load_workbook(output)
    ws = wb["固定导入页"]
    assert wb.sheetnames == ["固定导入页"]
    assert [ws.cell(2, col).value for col in range(1, 4)] == ["姓名", "编号", "成绩"]
    assert [ws.cell(3, col).value for col in range(1, 4)] == ["张三", "001", 95]
    assert "示例人员" not in [cell.value for row in ws.iter_rows() for cell in row]


def test_reference_template_copies_style_but_not_sample_data(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "reference.xlsx"
    _make_template(template)
    blueprint = {
        "title": "新表",
        "sheet_name": "新数据",
        "columns": [
            {"key": "name", "label": "姓名", "type": "text"},
            {"key": "amount", "label": "金额", "type": "money"},
        ],
        "records": [{"name": "李四", "amount": 300}],
    }
    build_rich_workbook(blueprint, output)

    apply_template_mode(output, template, mode="reference", blueprint=blueprint)

    ws = load_workbook(output)["新数据"]
    assert ws["A4"].value == "李四"
    assert "示例人员" not in [cell.value for row in ws.iter_rows() for cell in row]
    assert ws["A3"].fill.fgColor.rgb.endswith("4472C4")


def test_excel_title_row_is_not_mistaken_for_header(tmp_path):
    template = tmp_path / "template.xlsx"
    _make_template(template)

    frame = read_table(template)
    summary = inspect_template(template)

    assert list(frame.columns) == ["姓名", "编号", "成绩"]
    assert frame.iloc[0]["姓名"] == "示例人员"
    assert summary["sheets"][0]["header_row"] == 2
