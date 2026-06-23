from openpyxl import load_workbook

from excel_agent.rich_workbook_builder import build_rich_workbook, inspect_rich_workbook
from excel_agent.validators import validate_workbook


def sales_blueprint():
    records = [
        {"region": "华东", "province": "上海", "salesperson": "张三", "jan": 85000, "feb": 92000, "mar": 105000},
        {"region": "华东", "province": "上海", "salesperson": "李四", "jan": 78000, "feb": 88000, "mar": 96000},
        {"region": "华东", "province": "江苏", "salesperson": "王五", "jan": 120000, "feb": 115000, "mar": 130000},
        {"region": "华东", "province": "江苏", "salesperson": "赵六", "jan": 95000, "feb": 102000, "mar": 99000},
        {"region": "华东", "province": "浙江", "salesperson": "孙七", "jan": 110000, "feb": 108000, "mar": 125000},
        {"region": "华东", "province": "浙江", "salesperson": "周八", "jan": 88000, "feb": 93000, "mar": 87000},
        {"region": "华南", "province": "广东", "salesperson": "吴九", "jan": 150000, "feb": 160000, "mar": 170000},
        {"region": "华南", "province": "广东", "salesperson": "郑十", "jan": 135000, "feb": 142000, "mar": 138000},
        {"region": "华南", "province": "福建", "salesperson": "陈十一", "jan": 52000, "feb": 48000, "mar": 55000},
        {"region": "华南", "province": "福建", "salesperson": "刘十二", "jan": 45000, "feb": 42000, "mar": 48000},
    ]
    return {
        "title": "2026年第一季度销售业绩统计表",
        "sheet_name": "销售业绩",
        "columns": [
            {"key": "region", "label": "区域", "type": "text"},
            {"key": "province", "label": "省份", "type": "text"},
            {"key": "salesperson", "label": "销售员", "type": "text"},
            {"key": "jan", "label": "1月", "type": "money"},
            {"key": "feb", "label": "2月", "type": "money"},
            {"key": "mar", "label": "3月", "type": "money"},
            {
                "key": "quarter_total",
                "label": "季度总销售额(元)",
                "type": "money",
                "formula": "=SUM({jan},{feb},{mar})",
                "number_format": '[>=400000]"🟢 "#,##0;[<150000]"🔴 "#,##0;#,##0',
            },
            {
                "key": "quarter_avg",
                "label": "季度平均销售额(元)",
                "type": "money",
                "formula": "=ROUND({quarter_total}/3,0)",
            },
            {
                "key": "grade",
                "label": "业绩等级",
                "type": "text",
                "formula": '=IF({quarter_total}>=400000,"卓越",IF({quarter_total}>=350000,"优秀",IF({quarter_total}>=250000,"良好",IF({quarter_total}>=150000,"合格","⚠️ 待改进"))))',
            },
        ],
        "header_groups": [
            {"label": "月度销售额", "start_key": "jan", "end_key": "mar"},
            {"label": "季度统计", "start_key": "quarter_total", "end_key": "quarter_avg"},
        ],
        "records": records,
        "sort": [
            {"key": "region", "order": ["华东", "华南"]},
            {"keys": ["jan", "feb", "mar"], "aggregate": "sum", "direction": "desc"},
        ],
        "group_subtotals": {
            "group_key": "region",
            "label_template": "{group}小计",
            "merge_label_keys": ["region", "province"],
            "sum_keys": ["jan", "feb", "mar", "quarter_total"],
            "average_from": {"quarter_avg": "quarter_total"},
            "value_map": {"salesperson": "--", "grade": "--"},
        },
        "grand_total": {
            "label": "总计",
            "merge_label_keys": ["region", "province", "salesperson"],
            "sum_keys": ["jan", "feb", "mar", "quarter_total"],
            "average_from": {"quarter_avg": "quarter_total"},
            "blank_keys": ["grade"],
        },
        "conditional_formats": [
            {"kind": "cell_is", "column_key": "quarter_total", "operator": "greaterThanOrEqual", "value": 400000, "font_bold": True},
            {"kind": "cell_is", "column_key": "quarter_total", "operator": "lessThan", "value": 150000, "font_bold": True},
            {"kind": "formula", "column_key": "grade", "formula": '={cell}="⚠️ 待改进"', "font_bold": True},
        ],
        "charts": [
            {
                "type": "column",
                "title": "销售员季度业绩对比",
                "category_key": "salesperson",
                "value_keys": ["quarter_total"],
                "position": "K3",
            }
        ],
    }


def test_rich_builder_supports_multilevel_headers_subtotals_sort_and_chart(tmp_path):
    path = build_rich_workbook(
        sales_blueprint(),
        tmp_path / "sales.xlsx",
        require_charts=True,
    )
    wb = load_workbook(path, data_only=False)
    ws = wb["销售业绩"]

    assert ws["A1"].value == "2026年第一季度销售业绩统计表"
    assert {"A3:A4", "B3:B4", "C3:C4", "D3:F3", "G3:H3", "I3:I4"}.issubset(
        {str(item) for item in ws.merged_cells.ranges}
    )
    assert ws["C5"].value == "王五"
    assert ws["G5"].value == "=SUM(D5,E5,F5)"
    assert ws["H5"].value == "=ROUND(G5/3,0)"
    assert "卓越" in ws["I5"].value
    assert ws["A11"].value == "华东小计"
    assert ws["C11"].value == "--"
    assert ws["H11"].value == "=ROUND(AVERAGE(G5:G10),0)"
    assert ws["I11"].value == "--"
    assert ws["A16"].value == "华南小计"
    assert ws["A17"].value == "总计"
    assert ws["H17"].value == "=ROUND(AVERAGE(G5:G10,G12:G15),0)"
    assert len(ws._charts) == 1
    assert sum(len(item.rules) for item in ws.conditional_formatting) == 3
    assert ws.freeze_panes == "A5"
    helper = wb["_图表数据"]
    assert helper["A2"].value == "王五"
    assert helper["B2"].value == 365000
    assert not str(helper["B2"].value).startswith("=")

    inspection = inspect_rich_workbook(path)
    assert inspection["sheets"][0]["chart_count"] == 1
    assert inspection["sheets"][0]["conditional_format_count"] == 3


def test_grand_total_default_merge_does_not_overlap_sum_columns(tmp_path):
    blueprint = {
        "title": "预算表",
        "sheet_name": "预算",
        "columns": [
            {"key": "category", "label": "类别", "type": "text"},
            {"key": "item", "label": "项目", "type": "text"},
            {"key": "budget", "label": "预算金额", "type": "money"},
            {"key": "actual", "label": "实际金额", "type": "money"},
            {
                "key": "difference",
                "label": "差额",
                "type": "money",
                "formula": "={budget}-{actual}",
            },
        ],
        "records": [
            {"category": "餐饮", "item": "午餐", "budget": 1000, "actual": 900},
            {"category": "交通", "item": "地铁", "budget": 500, "actual": 450},
        ],
        "grand_total": {
            "label": "合计",
            "sum_keys": ["budget", "actual", "difference"],
            "value_map": {"category": "", "item": ""},
        },
    }

    path = build_rich_workbook(blueprint, tmp_path / "budget.xlsx")
    ws = load_workbook(path, data_only=False)["预算"]

    assert ws["A6"].value == "合计"
    assert "A6:B6" in {str(item) for item in ws.merged_cells.ranges}
    assert ws["C6"].value == "=SUM(C4:C5)"
    assert ws["D6"].value == "=SUM(D4:D5)"
    assert validate_workbook(path)["status"] == "pass"


def test_rich_builder_supports_multi_sheet_cross_references(tmp_path):
    blueprint = {
        "title": "关联测算表",
        "sheets": [
            {
                "sheet_name": "参数表",
                "title": "参数表",
                "columns": [
                    {"key": "item", "label": "项目", "type": "text"},
                    {"key": "value", "label": "参数值", "type": "percentage"},
                ],
                "records": [{"item": "调整比例", "value": 0.1}],
            },
            {
                "sheet_name": "明细表",
                "title": "员工明细表",
                "columns": [
                    {"key": "name", "label": "姓名", "type": "text"},
                    {"key": "salary", "label": "当前薪资", "type": "money"},
                    {
                        "key": "adjusted",
                        "label": "调整后薪资",
                        "type": "money",
                        "formula": "=IFERROR({salary}*(1+'参数表'!$B$4),0)",
                    },
                ],
                "records": [{"name": "张三", "salary": 10000}],
            },
        ],
    }
    path = build_rich_workbook(blueprint, tmp_path / "multi.xlsx")
    wb = load_workbook(path, data_only=False)
    assert wb.sheetnames == ["参数表", "明细表"]
    assert wb["明细表"]["C4"].value == "=IFERROR(B4*(1+'参数表'!$B$4),0)"
    assert validate_workbook(path)["status"] == "pass"
