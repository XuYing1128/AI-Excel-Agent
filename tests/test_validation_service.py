import json

from openpyxl import Workbook

from excel_agent.services.validation_service import validate_generated_workbook
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec import TaskSpec


def test_validation_service_writes_json_report(tmp_path):
    paths = create_task_paths("generic_table", tmp_path / "tasks")
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "说明"
    data = wb.create_sheet("Data")
    data.append(["日期", "数量"])
    data.append(["2026-06-21", 1])
    wb.save(paths.output_file)

    result = validate_generated_workbook(
        paths.output_file,
        TaskSpec(task_type="generic_table", user_goal="测试"),
        paths,
    )
    assert result.status in {"pass", "warn"}
    assert paths.validation_report.exists()
    report = json.loads(paths.validation_report.read_text(encoding="utf-8"))
    assert report["file"].endswith("result.xlsx")
