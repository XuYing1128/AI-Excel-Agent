import json

from openpyxl import load_workbook

from excel_agent.model_registry import ModelSettings, ProviderConfig, save_model_settings
from excel_agent.services.agent import orchestrator
from excel_agent.services.custom_api_service import ToolCall, ToolChatResult
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec import TaskSpec


def _write_model_settings(path):
    save_model_settings(
        ModelSettings(
            providers=[
                ProviderConfig(
                    id="builder",
                    name="构建模型",
                    base_url="https://example.com/v1",
                    api_key="secret",
                    model="builder-model",
                )
            ],
            roles={"builder": "builder"},
            agent_enabled=True,
            run_python_enabled=True,
        ),
        path,
    )


def test_agent_orchestrator_builds_workbook_with_mock_tools(tmp_path, monkeypatch):
    model_path = tmp_path / "model_settings.json"
    _write_model_settings(model_path)
    monkeypatch.setenv("AI_EXCEL_MODEL_SETTINGS_FILE", str(model_path))
    blueprint = {
        "title": "项目清单",
        "sheet_name": "任务",
        "columns": [
            {"key": "task", "label": "任务", "type": "text"},
            {"key": "owner", "label": "负责人", "type": "text"},
            {"key": "progress", "label": "完成率", "type": "percentage"},
        ],
        "records": [
            {"task": "整理需求", "owner": "张三", "progress": 0.5},
            {"task": "生成表格", "owner": "李四", "progress": 0.8},
        ],
    }

    def fake_chat_with_tools(*args, **kwargs):
        return ToolChatResult(
            success=True,
            content="",
            tool_calls=[
                ToolCall("call_build", "build_workbook", {"blueprint": blueprint}),
                ToolCall("call_finish", "finish_task", {"summary": "完成"}),
            ],
            error=None,
            status_code=200,
            latency_ms=10,
            message={"role": "assistant", "content": "", "tool_calls": []},
        )

    monkeypatch.setattr(orchestrator.model_registry, "chat_with_tools", fake_chat_with_tools)
    paths = create_task_paths("project_plan", tmp_path / "tasks", output_name="项目清单.xlsx")
    result = orchestrator.run_agent(
        TaskSpec(task_type="project_plan", user_goal="生成项目清单", output_name="项目清单.xlsx"),
        paths,
        max_steps=3,
    )

    assert result.success is True
    assert paths.output_file.exists()
    assert result.tool_calls == 2
    wb = load_workbook(paths.output_file)
    assert "任务" in wb.sheetnames
    assert wb["任务"]["A1"].value == "项目清单"
    blueprint_payload = json.loads((paths.task_dir / "agent_workbook_blueprint.json").read_text(encoding="utf-8"))
    assert blueprint_payload["title"] == "项目清单"


def test_agent_orchestrator_builds_complex_chart(tmp_path, monkeypatch):
    """The user's core need: the agent must reliably produce a real, non-empty
    chart of the requested type via the robust rich-builder path."""
    model_path = tmp_path / "model_settings.json"
    _write_model_settings(model_path)
    monkeypatch.setenv("AI_EXCEL_MODEL_SETTINGS_FILE", str(model_path))
    blueprint = {
        "title": "各区域季度销售对比",
        "sheet_name": "对比",
        "columns": [
            {"key": "region", "label": "区域", "type": "text"},
            {"key": "q1", "label": "一季度", "type": "money"},
            {"key": "q2", "label": "二季度", "type": "money"},
            {"key": "q3", "label": "三季度", "type": "money"},
        ],
        "records": [
            {"region": "华东", "q1": 120, "q2": 150, "q3": 170},
            {"region": "华南", "q1": 90, "q2": 95, "q3": 110},
            {"region": "华北", "q1": 60, "q2": 70, "q3": 85},
        ],
        "charts": [
            {"type": "column", "title": "区域季度对比", "category_key": "region",
             "value_keys": ["q1", "q2", "q3"]}
        ],
    }

    def fake_chat_with_tools(*args, **kwargs):
        return ToolChatResult(
            success=True,
            content="",
            tool_calls=[
                ToolCall("call_build", "build_workbook", {"blueprint": blueprint}),
                ToolCall("call_finish", "finish_task", {"summary": "完成"}),
            ],
            error=None,
            status_code=200,
            latency_ms=10,
            message={"role": "assistant", "content": "", "tool_calls": []},
        )

    monkeypatch.setattr(orchestrator.model_registry, "chat_with_tools", fake_chat_with_tools)
    paths = create_task_paths("sales_report", tmp_path / "tasks", output_name="对比.xlsx")
    spec = TaskSpec(task_type="sales_report", user_goal="按区域对比各季度销售并出柱状图",
                    output_name="对比.xlsx", include_charts=True)
    result = orchestrator.run_agent(spec, paths, max_steps=3)

    assert result.success is True
    wb = load_workbook(paths.output_file)
    ws = wb["对比"]
    assert ws._charts, "agent output must contain a chart"
    chart = ws._charts[0]
    assert len(chart.series) == 3, "multi-series comparison chart"
    # The chart must read literal values (helper sheet), not un-recalculated formulas.
    helper = [name for name in wb.sheetnames if name.startswith("_图表")]
    assert helper, "chart should use a literal-value helper sheet"
    sample = wb[helper[0]].cell(2, 2).value
    assert isinstance(sample, (int, float)), "chart data must be literal numbers"


def test_agent_orchestrator_returns_failure_when_model_stalls(tmp_path, monkeypatch):
    model_path = tmp_path / "model_settings.json"
    _write_model_settings(model_path)
    monkeypatch.setenv("AI_EXCEL_MODEL_SETTINGS_FILE", str(model_path))

    def fake_chat_with_tools(*args, **kwargs):
        return ToolChatResult(
            success=True,
            content="我先解释一下",
            tool_calls=[],
            error=None,
            status_code=200,
            latency_ms=10,
            message={"role": "assistant", "content": "我先解释一下"},
        )

    monkeypatch.setattr(orchestrator.model_registry, "chat_with_tools", fake_chat_with_tools)
    paths = create_task_paths("generic_table", tmp_path / "tasks")
    result = orchestrator.run_agent(
        TaskSpec(task_type="generic_table", user_goal="生成一个表"),
        paths,
        max_steps=2,
    )

    assert result.success is False
    assert "工具" in (result.error or "")


def test_agent_orchestrator_can_use_run_python_for_long_tail_task(tmp_path, monkeypatch):
    model_path = tmp_path / "model_settings.json"
    _write_model_settings(model_path)
    monkeypatch.setenv("AI_EXCEL_MODEL_SETTINGS_FILE", str(model_path))
    code = r'''
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "去重统计"
ws.append(["姓名", "科目数"])
ws.append(["张三", 2])
ws.append(["李四", 1])
wb.save(OUTPUT_DIR + "/result.xlsx")
'''

    def fake_chat_with_tools(*args, **kwargs):
        return ToolChatResult(
            success=True,
            content="",
            tool_calls=[
                ToolCall("call_python", "run_python", {"code": code, "timeout_seconds": 10}),
                ToolCall("call_finish", "finish_task", {"summary": "完成"}),
            ],
            error=None,
            status_code=200,
            latency_ms=10,
            message={"role": "assistant", "content": "", "tool_calls": []},
        )

    monkeypatch.setattr(orchestrator.model_registry, "chat_with_tools", fake_chat_with_tools)
    paths = create_task_paths("generic_table", tmp_path / "tasks", output_name="result.xlsx")
    result = orchestrator.run_agent(
        TaskSpec(task_type="generic_table", user_goal="合并名单去重并统计每人科目数"),
        paths,
        max_steps=2,
    )

    assert result.success is True
    wb = load_workbook(paths.output_file)
    assert wb["去重统计"]["A1"].value == "姓名"
