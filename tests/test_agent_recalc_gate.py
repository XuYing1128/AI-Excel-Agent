"""真算闭环“修复链”集成测试。

mock 模型分两轮：先让 AI 写出含循环引用的表（真算会报 #VALUE!），再写修复代码。
run_python 与 LibreOffice 真算都真实执行，验证整条链：真算抓错 → 打回 → AI 改对 → 放行。
没装 LibreOffice 时跳过（修复链依赖真算）。
"""

import types

import pytest
from openpyxl import load_workbook

from excel_agent.services import recalc
from excel_agent.services.agent import orchestrator
from excel_agent.services.agent.orchestrator import run_agent
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec import TaskSpec


class _Call:
    def __init__(self, name, arguments, cid):
        self.id = cid
        self.name = name
        self.arguments = arguments


class _Resp:
    def __init__(self, tool_calls):
        self.success = True
        self.error = None
        self.message = None
        self.tool_calls = tool_calls
        self.content = ""


# 第 1 轮“AI”写的代码：B2 = AVERAGE(B2:B3) 把自己圈进范围 → 循环引用 → 真算报 #VALUE!。
WRITE_BAD = (
    "from openpyxl import Workbook\n"
    "wb = Workbook(); ws = wb.active; ws.title = '数据'\n"
    "ws['A1'] = '项目'; ws['A2'] = 10; ws['A3'] = 20\n"
    "ws['B1'] = '均值'; ws['B2'] = '=AVERAGE(B2:B3)'\n"
    "wb.save(OUTPUT_FILE)\n"
)
# 第 2 轮：把 B2 改成引用 A 列、不再自引用。
FIX = (
    "from openpyxl import load_workbook\n"
    "wb = load_workbook(OUTPUT_FILE); ws = wb['数据']\n"
    "ws['B2'] = '=AVERAGE(A2:A3)'\n"
    "wb.save(OUTPUT_FILE)\n"
)


@pytest.mark.skipif(not recalc.recalc_available(), reason="修复链依赖本机 LibreOffice 真算")
def test_recalc_gate_repairs_circular_then_passes(tmp_path, monkeypatch):
    responses = [
        _Resp([_Call("run_python", {"code": WRITE_BAD}, "c1"),
               _Call("finish_task", {"summary": "done"}, "f1")]),
        _Resp([_Call("run_python", {"code": FIX}, "c2"),
               _Call("finish_task", {"summary": "fixed"}, "f2")]),
    ]

    def fake_chat(*args, **kwargs):
        return responses.pop(0) if responses else _Resp([])

    fake_settings = types.SimpleNamespace(agent_enabled=True, run_python_enabled=True)
    monkeypatch.setattr(orchestrator, "load_model_settings", lambda: fake_settings)
    monkeypatch.setattr(
        orchestrator.model_registry,
        "get_provider",
        lambda role, settings: types.SimpleNamespace(name="fake", model="fake"),
    )
    monkeypatch.setattr(orchestrator.model_registry, "chat_with_tools", fake_chat)

    spec = TaskSpec(task_type="generic_table", user_goal="给数据加均值统计")
    paths = create_task_paths("generic_table", tmp_path)
    result = run_agent(spec, paths, max_steps=6)

    assert result.success is True
    assert result.steps >= 2  # 第 1 轮被真算打回，至少走到第 2 轮
    wb = load_workbook(result.output_file)
    assert wb["数据"]["B2"].value == "=AVERAGE(A2:A3)"  # 修复确实生效
    log_text = (paths.task_dir / "run_log.json").read_text(encoding="utf-8")
    assert "agent_recalc_gate" in log_text  # 修复链确被触发，而非第 1 轮直接放行
