from openpyxl import load_workbook

from excel_agent.content_plan import build_local_content_plan, merge_model_content_plan
from excel_agent.custom_workbook_builder import build_custom_workbook
from excel_agent.services.validation_service import validate_generated_workbook
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec_builder import build_task_spec_draft


WEATHER_PROMPT = """请根据以下要求与数据，生成一个格式清晰、内容准确的表格。
主题：2026年6月第三周美国西海岸三城市天气概况
表格需包含以下列：日期、城市、天气状况、最高气温（℃）、最低气温（℃）、降水概率、日均气温（℃）、出行建议
日均气温请根据给出的最高与最低气温自动计算（取平均值）
出行建议规则：若降水概率 ≥ 50%，建议“携带雨具”；若最高气温 ≥ 30℃，建议“注意防暑”；其他情况建议“适宜出行”。
表格最后一行需增加“周平均”统计，计算三个城市各自的周平均日均气温。
数据如下：
6月15日
洛杉矶：晴，18 ～ 29℃，降水概率5%
旧金山：多云，13 ～ 21℃，降水概率10%
西雅图：小雨，12 ～ 18℃，降水概率60%
"""


def test_weather_request_becomes_explicit_content_plan():
    plan = build_local_content_plan(WEATHER_PROMPT, "generic_table")
    assert plan["title"] == "2026年6月第三周美国西海岸三城市天气概况"
    assert len(plan["records"]) == 3
    assert [item["name"] for item in plan["columns"]][-2:] == [
        "日均气温（℃）",
        "出行建议",
    ]
    assert {item["kind"] for item in plan["formula_rules"]} == {
        "average",
        "weather_advice",
    }
    merged = merge_model_content_plan(
        plan,
        {"title": "模型擅自改名", "columns": [{"name": "无关列"}]},
    )
    assert merged["title"] == "2026年6月第三周美国西海岸三城市天气概况"
    assert [item["name"] for item in merged["columns"]] == [
        "日期",
        "城市",
        "天气状况",
        "最高气温（℃）",
        "最低气温（℃）",
        "降水概率",
        "日均气温（℃）",
        "出行建议",
    ]


def test_custom_weather_workbook_matches_requested_columns_and_formulas(tmp_path):
    draft = build_task_spec_draft(WEATHER_PROMPT, [])
    spec = draft.task_spec
    assert draft.clarifying_questions == []
    paths = create_task_paths(
        spec.task_type,
        tmp_path,
        output_name=spec.output_name,
    )
    build_custom_workbook(spec.options["content_plan"], paths.output_file)
    wb = load_workbook(paths.output_file, data_only=False)
    ws = wb.active
    assert wb.sheetnames == ["2026年6月第三周美国西海岸三城市天气概况"]
    assert [ws.cell(3, column).value for column in range(1, 9)] == [
        "日期",
        "城市",
        "天气状况",
        "最高气温（℃）",
        "最低气温（℃）",
        "降水概率",
        "日均气温（℃）",
        "出行建议",
    ]
    assert ws["G4"].value.startswith("=IF(")
    assert "建议携带雨具" in ws["H4"].value

    result = validate_generated_workbook(paths.output_file, spec, paths)
    assert result.status == "pass"
    assert result.summary["error_count"] == 0
    assert result.summary["warning_count"] == 0
