from excel_agent.services.agent.tools import ToolContext, registered_tools
from excel_agent.services.agent.tools.run_python import run_python_code, run_python_tool
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec import TaskSpec


def _ctx(tmp_path, *, enabled=True):
    paths = create_task_paths("generic_table", tmp_path / "tasks")
    return ToolContext(
        task_spec=TaskSpec(task_type="generic_table", user_goal="测试脚本"),
        task_paths=paths,
        run_python_enabled=enabled,
    )


def test_run_python_writes_only_task_output(tmp_path):
    ctx = _ctx(tmp_path)
    result = run_python_code(
        """
import json
with open(OUTPUT_DIR + "/result.json", "w", encoding="utf-8") as f:
    json.dump({"ok": True}, f)
print("done")
""",
        ctx,
        timeout_seconds=10,
    )

    assert result.ok is True
    assert (ctx.task_paths.output_dir / "result.json").exists()
    assert "done" in result.data["stdout"]
    assert any(path.endswith("result.json") for path in result.artifacts)


def test_run_python_allows_import_and_network_after_unlock(tmp_path):
    # 翻转后：信任本地环境，import 与联网都放开（os 可用、socket 可创建）。
    ctx = _ctx(tmp_path)
    result = run_python_code(
        """
import os, socket
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.close()
print("cwd:" + os.path.basename(os.getcwd()))
print("socket-ok")
""",
        ctx,
        timeout_seconds=10,
    )

    assert result.ok is True, result.data
    assert "socket-ok" in result.data["stdout"]


def test_run_python_rejects_path_escape(tmp_path):
    ctx = _ctx(tmp_path)
    outside = tmp_path / "outside.txt"
    result = run_python_code(
        f'open(r"{outside}", "w", encoding="utf-8").write("bad")',
        ctx,
        timeout_seconds=10,
    )

    assert result.ok is False
    assert not outside.exists()
    assert "path_escape" in result.data["stderr"]


def test_run_python_timeout_is_killed(tmp_path):
    ctx = _ctx(tmp_path)
    result = run_python_code("while True:\n    pass", ctx, timeout_seconds=1)

    assert result.ok is False
    assert result.error == "timeout"


def test_run_python_tool_hidden_when_disabled(tmp_path):
    ctx = _ctx(tmp_path, enabled=False)
    names = [tool.name for tool in registered_tools(ctx)]
    assert "run_python" not in names


def test_run_python_edits_existing_workbook_in_place(tmp_path):
    # 就地编辑铁律的机制验证：load 已有工作簿、只填目标列，
    # 原有的其它工作表 / 数据 / 顺序必须原样保留（即用户“填表”案例该有的正确行为）。
    from openpyxl import Workbook, load_workbook

    ctx = _ctx(tmp_path)
    ctx.task_paths.input_dir.mkdir(parents=True, exist_ok=True)
    source = ctx.task_paths.input_dir / "tableB.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    ws["A1"], ws["B1"], ws["C1"] = "姓名", "部门", "工资"
    ws["A2"], ws["B2"], ws["C2"] = "张三", "销售", None
    ws["A3"], ws["B3"], ws["C3"] = "李四", "研发", None
    keep = wb.create_sheet("备注")
    keep["A1"] = "请勿改动"
    wb.save(source)

    code = f'''
from openpyxl import load_workbook
wb = load_workbook(r"{source}")
ws = wb["明细"]
salary = {{"张三": 8000, "李四": 9000}}
for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name in salary:
        ws.cell(row=row, column=3).value = salary[name]
wb.save(OUTPUT_FILE)
print("filled")
'''
    result = run_python_code(code, ctx, timeout_seconds=30)

    assert result.ok is True, result.data
    out = load_workbook(ctx.task_paths.output_file)
    assert out.sheetnames == ["明细", "备注"]
    assert out["备注"]["A1"].value == "请勿改动"
    detail = out["明细"]
    assert detail["A2"].value == "张三"
    assert detail["B3"].value == "研发"
    assert detail["C2"].value == 8000
    assert detail["C3"].value == 9000
