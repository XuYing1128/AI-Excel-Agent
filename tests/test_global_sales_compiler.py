from openpyxl import load_workbook

from excel_agent.services.generation_service import generate_from_task_spec
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec_builder import build_task_spec_draft


GLOBAL_SALES_PROMPT = """2026年上半年全球销售业绩 Excel 工作簿制作（纯表格·多Sheet·含数据透视表）

生成一个包含 5 个工作表的 Excel 工作簿：
Sheet1：销售目标参数表（用于引用）
Sheet2：销售明细表（72行数据，含公式与条件格式）
Sheet3：地区汇总表（含 SUMIF 跨表汇总）
Sheet4：产品汇总表
Sheet5：数据透视表布局（交叉汇总，模拟透视表）

Sheet1 工作表标签名：参数表，标题：「各区域产品月度销售目标（元）」
列：地区、产品、月度销售目标(元)
地区	产品	月度销售目标(元)
北美	电子产品	150,000
北美	家居用品	80,000
北美	服装	60,000
欧洲	电子产品	110,000
欧洲	家居用品	90,000
欧洲	服装	70,000
亚洲	电子产品	130,000
亚洲	家居用品	100,000
亚洲	服装	50,000
其他	电子产品	70,000
其他	家居用品	50,000
其他	服装	40,000

Sheet2 工作表标签名：明细。列：地区、产品、月份、销售额(元)、利润(元)、利润率、目标销售额(元)、达成率
基础数据如下（72行，按地区-产品-月份顺序）：
北美-电子产品：(125000,18000) (130000,19000) (135000,20000) (145000,22500) (155000,25000) (165000,27000)
北美-家居用品：(80000,12000) (82000,13000) (78000,11500) (85000,14000) (88000,15000) (90000,16000)
北美-服装：(60000,3000) (58000,2500) (62000,3500) (65000,4000) (70000,4500) (75000,5000)
欧洲-电子产品：(100000,15000) (105000,16000) (110000,17000) (115000,18000) (120000,19000) (125000,20000)
欧洲-家居用品：(90000,14000) (85000,12000) (88000,13000) (86000,12500) (87000,12800) (89000,13500)
欧洲-服装：(70000,14000) (72000,15000) (75000,16000) (78000,17000) (80000,18000) (82000,19000)
亚洲-电子产品：(140000,28000) (150000,30000) (160000,32000) (170000,34000) (180000,36000) (190000,38000)
亚洲-家居用品：(100000,15000) (95000,14000) (102000,16000) (98000,14800) (105000,17000) (110000,18000)
亚洲-服装：(48000,2000) (50000,2500) (52000,3000) (55000,3500) (53000,2800) (56000,3200)
其他-电子产品：(65000,8000) (68000,8500) (70000,9000) (72000,9500) (71000,9000) (73000,10000)
其他-家居用品：(45000,5000) (48000,5500) (50000,6000) (49000,5800) (51000,6200) (52000,6500)
其他-服装：(38000,1000) (40000,1200) (39000,1100) (41000,1300) (42000,1400) (43000,1500)

利润率 = 利润 / 销售额；目标销售额用 INDEX+MATCH；达成率 = 销售额 / 目标销售额 - 1。
地区汇总按地区，产品汇总按产品，交叉汇总按地区和产品展示总销售额与总利润。
"""


def test_global_sales_prompt_uses_domain_compiler(tmp_path, monkeypatch):
    monkeypatch.setenv("AI_EXCEL_OUTPUTS_DIR", str(tmp_path / "outputs"))
    draft = build_task_spec_draft(GLOBAL_SALES_PROMPT, [])
    spec = draft.task_spec
    spec.output_name = "全球销售分析.xlsx"
    paths = create_task_paths(spec.task_type, tmp_path / "outputs" / "tasks", output_name=spec.output_name)

    result = generate_from_task_spec(spec, paths)

    assert result.success is True
    assert result.mode == "domain_compiler:global_sales_analysis"
    wb = load_workbook(paths.output_file, data_only=False)
    assert wb.sheetnames == ["参数表", "明细", "地区汇总", "产品汇总", "交叉汇总"]
    assert wb["明细"].max_row == 75
    formula_count = sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    assert formula_count >= 250
    assert wb["明细"]["F4"].value == "=IFERROR(E4/D4,0)"
    assert "INDEX" in wb["明细"]["G4"].value
    assert wb["交叉汇总"]["B5"].value.startswith("=SUMIFS(")
