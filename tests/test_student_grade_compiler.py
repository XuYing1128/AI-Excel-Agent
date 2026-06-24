from openpyxl import load_workbook

from excel_agent.content_plan import build_local_content_plan
from excel_agent.inline_table_parser import extract_inline_tables
from excel_agent.services.generation_service import generate_from_task_spec
from excel_agent.services.validation_service import validate_generated_workbook
from excel_agent.task_paths import create_task_paths
from excel_agent.task_spec import TaskSpec


COURSE_ROWS = [
    ("CS101", "高等数学", 4, "必修"),
    ("CS102", "大学英语", 3, "必修"),
    ("CS103", "程序设计基础", 3, "必修"),
    ("CS104", "体育", 1, "必修"),
]

STUDENT_ROWS = [
    ("20240001", "张三", "计算机科学", "计科2101"),
    ("20240002", "李四", "计算机科学", "计科2101"),
    ("20240003", "王五", "计算机科学", "计科2102"),
    ("20240004", "赵六", "计算机科学", "计科2102"),
    ("20240005", "孙七", "计算机科学", "计科2103"),
    ("20240006", "周八", "计算机科学", "计科2103"),
    ("20240007", "吴九", "软件工程", "软件2101"),
    ("20240008", "郑十", "软件工程", "软件2101"),
    ("20240009", "陈十一", "软件工程", "软件2102"),
    ("20240010", "刘十二", "软件工程", "软件2102"),
    ("20240011", "李明", "软件工程", "软件2103"),
    ("20240012", "王芳", "软件工程", "软件2103"),
    ("20240013", "张伟", "信息安全", "信安2101"),
    ("20240014", "刘洋", "信息安全", "信安2101"),
    ("20240015", "陈静", "信息安全", "信安2102"),
    ("20240016", "杨磊", "信息安全", "信安2102"),
    ("20240017", "赵敏", "信息安全", "信安2103"),
    ("20240018", "周杰", "信息安全", "信安2103"),
    ("20240019", "吴昊", "信息安全", "信安2103"),
    ("20240020", "郑丽", "信息安全", "信安2103"),
]


def _student_grade_prompt() -> str:
    course_table = "\n".join(
        ["课程编号\t课程名称\t学分\t课程类别"]
        + [f"{code}\t{name}\t{credit}\t{kind}" for code, name, credit, kind in COURSE_ROWS]
    )
    student_table = "\n".join(
        ["学号\t姓名\t专业\t班级"]
        + [f"{sid}\t{name}\t{major}\t{klass}" for sid, name, major, klass in STUDENT_ROWS]
    )
    score_lines = []
    for index, (student_id, *_rest) in enumerate(STUDENT_ROWS):
        base = 60 + index
        score_lines.append(
            f"{student_id}："
            f"CS101({base},{base + 2}), "
            f"CS102({base + 1},{base + 3}), "
            f"CS103({base + 2},{base + 4}), "
            f"CS104({base + 3},{base + 5})"
        )
    scores = "\n".join(score_lines)
    return f"""2026年春季学期学生成绩管理与分析系统
请生成 6 个工作表：课程参数、学生信息、成绩录入、学期总评、专业汇总、课程统计。
成绩录入必须包含20名学生×4门课程=80行，公式必须是活公式。
课程参数表：
{course_table}

学生信息表：
{student_table}

成绩录入表需要按专业、学号、课程编号排序，公式包含课程名称、学分、总评成绩、绩点、备注。
总评成绩 = 平时成绩*30% + 期末成绩*70%，绩点用 LOOKUP 或 IF 实现。
学期总评要使用 SUMIF、SUMPRODUCT、RANK 计算总修学分、加权平均分、GPA、排名、等级。
专业汇总使用 COUNTIF、AVERAGEIF、MAXIFS、MINIFS、COUNTIFS。
课程统计使用 AVERAGEIF、MAXIFS、MINIFS、COUNTIFS。
专业列相同专业连续单元格合并，设置冻结窗格、自动筛选、条件格式和数据验证。
成绩数据如下：
{scores}
"""


def test_review_feedback_is_not_parsed_as_business_tables():
    prompt = (
        _student_grade_prompt()
        + "\n本次修改要求：\n"
        + "修正问题：生成的表格包含无关工作表。\n"
        + "采用建议：重新生成工作簿，严格包含6个命名工作表。\n"
    )

    tables = extract_inline_tables(prompt)
    plan = build_local_content_plan(prompt, "generic_table")

    assert [table["row_count"] for table in tables] == [4, 20]
    assert [table["row_count"] for table in plan["inline_tables"]] == [4, 20]
    assert all(
        not str(table.get("name", "")).startswith(("修正问题", "采用建议"))
        for table in plan["inline_tables"]
    )


def test_student_grade_compiler_generates_exact_six_sheet_workbook(tmp_path):
    prompt = _student_grade_prompt()
    plan = build_local_content_plan(prompt, "generic_table")
    spec = TaskSpec(
        task_type="generic_table",
        user_goal=prompt,
        output_name="2026春季学期成绩分析.xlsx",
        include_charts=False,
        include_summary=False,
        include_instructions_sheet=False,
        options={
            "generation_policy": "custom_content",
            "content_plan": plan,
            "chart_requested_explicitly": False,
            "chart_requirements": {
                "required": False,
                "types": [],
                "explicit": False,
                "negative": False,
                "reason": "本任务未要求图表",
            },
            "chart_types": [],
        },
    )
    paths = create_task_paths(spec.task_type, tmp_path / "tasks", output_name=spec.output_name)

    result = generate_from_task_spec(spec, paths)
    validation = validate_generated_workbook(paths.output_file, spec, paths)
    wb = load_workbook(paths.output_file, data_only=False)

    assert result.success is True
    assert result.mode == "domain_compiler:student_grade_analysis"
    assert validation.status == "pass"
    assert wb.sheetnames == ["课程参数", "学生信息", "成绩录入", "学期总评", "专业汇总", "课程统计"]
    assert all(not name.startswith(("修正问题", "采用建议")) for name in wb.sheetnames)
    assert wb["成绩录入"].max_row == 83
    assert wb["学期总评"].max_row == 23
    assert wb["成绩录入"]["H4"].value.startswith("=IFERROR(F4*30%+G4*70%")
    assert "LOOKUP" in wb["成绩录入"]["J4"].value
    assert "SUMPRODUCT" in wb["学期总评"]["F4"].value
    assert wb["专业汇总"]["B4"].value.startswith("=COUNTIF")
    assert "COUNTIFS" in wb["课程统计"]["F5"].value
    formula_count = sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    assert formula_count >= 700
