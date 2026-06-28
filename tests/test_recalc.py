"""真算校验模块测试。降级/纯函数用例无外部依赖；真算用例在没装 LibreOffice 时跳过。"""

import pytest
from openpyxl import Workbook

from excel_agent.services import recalc
from excel_agent.services.recalc import (
    describe_error_cells,
    recalc_available,
    recalc_workbook,
)


def test_describe_error_cells_basic():
    text = describe_error_cells([{"sheet": "科目统计", "cell": "C2", "value": "#VALUE!"}])
    assert "科目统计!C2=#VALUE!" in text


def test_describe_error_cells_truncates():
    cells = [{"sheet": "S", "cell": f"A{i}", "value": "#VALUE!"} for i in range(30)]
    text = describe_error_cells(cells, limit=5)
    assert "等共 30 处" in text
    assert text.count("、") == 4  # 只展开 5 个


def test_recalc_degrades_without_soffice(monkeypatch, tmp_path):
    monkeypatch.setattr(recalc, "find_soffice", lambda: None)
    path = tmp_path / "x.xlsx"
    wb = Workbook()
    wb.active["A1"] = "=1+1"
    wb.save(path)
    report = recalc_workbook(path)
    assert report["available"] is False
    assert report["ok"] is True  # 没装 LibreOffice 时绝不阻塞出表
    assert report["error_cells"] == []


def test_recalc_missing_file_does_not_block(monkeypatch, tmp_path):
    monkeypatch.setattr(recalc, "find_soffice", lambda: "soffice")
    report = recalc_workbook(tmp_path / "nope.xlsx")
    assert report["ok"] is True


@pytest.mark.skipif(not recalc_available(), reason="需要本机 LibreOffice")
def test_recalc_detects_circular(tmp_path):
    path = tmp_path / "circular.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["B1"] = "=AVERAGE(B1:B3)"  # 范围含自己 → 循环引用，真算才发作
    wb.save(path)
    report = recalc_workbook(path)
    assert report["available"] is True
    assert report["ok"] is False
    assert any(item["cell"] == "B1" for item in report["error_cells"])


@pytest.mark.skipif(not recalc_available(), reason="需要本机 LibreOffice")
def test_recalc_passes_clean_workbook(tmp_path):
    path = tmp_path / "clean.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)
    report = recalc_workbook(path)
    assert report["ok"] is True
    assert report["error_cells"] == []


def _make_ctx(tmp_path):
    from openpyxl import Workbook

    from excel_agent.services.agent.tools.base import ToolContext
    from excel_agent.task_paths import create_task_paths
    from excel_agent.task_spec import TaskSpec

    paths = create_task_paths("generic_table", tmp_path)
    paths.output_file.parent.mkdir(parents=True, exist_ok=True)
    Workbook().save(paths.output_file)
    ctx = ToolContext(task_spec=TaskSpec(task_type="generic_table", user_goal="x"), task_paths=paths)
    return ctx, paths


def test_recalc_check_tool_reports_errors(monkeypatch, tmp_path):
    import importlib

    et = importlib.import_module("excel_agent.services.agent.tools.excel_tools")

    monkeypatch.setattr(
        et,
        "recalc_workbook",
        lambda path, **kw: {
            "available": True,
            "ok": False,
            "error_cells": [{"sheet": "统计", "cell": "B2", "value": "#VALUE!"}],
            "detail": "x",
        },
    )
    ctx, paths = _make_ctx(tmp_path)
    tool = {t.name: t for t in et.excel_tools()}["recalc_check"]
    result = tool.handler({"path": str(paths.output_file)}, ctx)
    assert result.ok is False
    assert "统计!B2=#VALUE!" in result.summary


def test_recalc_check_tool_degrades_without_soffice(monkeypatch, tmp_path):
    import importlib

    et = importlib.import_module("excel_agent.services.agent.tools.excel_tools")

    monkeypatch.setattr(
        et,
        "recalc_workbook",
        lambda path, **kw: {"available": False, "ok": True, "error_cells": [], "detail": "no soffice"},
    )
    ctx, paths = _make_ctx(tmp_path)
    tool = {t.name: t for t in et.excel_tools()}["recalc_check"]
    result = tool.handler({"path": str(paths.output_file)}, ctx)
    assert result.ok is True
    assert "未安装" in result.summary
